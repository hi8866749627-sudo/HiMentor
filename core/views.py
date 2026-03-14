# ---------- DJANGO ----------
import io
import json
import os
import re
import tempfile
import threading
import uuid
import zipfile
import json
from datetime import datetime, date, timedelta, time
from zoneinfo import ZoneInfo
from django.core.management import call_command

from django.conf import settings
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth import authenticate, login
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.models import User
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.db import close_old_connections
from django.db.models import Count
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse
from django.db.models import Max 
from django.db.models import Q
from urllib.parse import quote
from django.core.paginator import Paginator
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
import pandas as pd
# ---------- LOCAL FORMS ----------
from .forms import UploadFileForm

# ---------- LOCAL MODELS ----------
from .models import (
    AcademicModule,
    Attendance,
    AcademicCalendar,
    AcademicHoliday,
    AttendanceWeekMeta,
    CallRecord,
    CoordinatorModuleAccess,
    Mentor,
    MentorPassword,
    OtherCallRecord,
    PracticalMarkUpload,
    SifMarksLock,
    ResultCallRecord,
    ResultUpload,
    ResultUploadJob,
    Student,
    StudentPracticalMark,
    StudentResult,
    Subject,
    SubjectTemplate,
    SubjectAlias,
    TimetableEntry,
    TimetableUpload,
    LectureSession,
    LectureAbsence,
    Room,
    LectureAdjustment,
    WeekLock,
)

# ---------- LOCAL UTILITIES ----------
from .utils import import_students_from_excel, resolve_mentor_identity
from .attendance_utils import import_attendance
from .lecture_utils import (
    parse_timetable_excel,
    phase_range,
    phase_for_date,
    week_for_date,
    end_date_for_week,
    slot_start_time,
)
from .result_utils import import_compiled_bulk_all, import_compiled_result_sheet, import_result_sheet
from .practical_utils import import_practical_marks, ordered_subjects
from .pdf_report import generate_student_pdf, generate_student_prefilled_pdf
from .module_utils import allowed_modules_for_user, get_current_module, is_superadmin_user

TEST_NAMES = ["T1", "T2", "T3", "T4", "REMEDIAL"]
IST = ZoneInfo("Asia/Kolkata")


def _session_mentor_obj(request):
    mentor_key = request.session.get("mentor")
    mentor = resolve_mentor_identity(mentor_key)
    if mentor and mentor_key != mentor.name:
        request.session["mentor"] = mentor.name
    return mentor


def _active_module(request):
    return get_current_module(request)


def _latest_attendance_calls_map(module, week_no, mentor=None):
    qs = CallRecord.objects.filter(student__module=module, week_no=week_no).select_related("student", "student__mentor")
    if mentor:
        qs = qs.filter(student__mentor=mentor)
    qs = qs.order_by("student_id", "-created_at", "-id")
    latest = {}
    for rec in qs:
        if rec.student_id not in latest:
            latest[rec.student_id] = rec
    return latest


def _latest_result_calls_map(upload, mentor=None, student=None, module=None, student_ids=None):
    qs = ResultCallRecord.objects.filter(upload=upload).select_related("student", "student__mentor", "upload", "upload__subject")
    if module:
        qs = qs.filter(student__module=module)
    if mentor:
        qs = qs.filter(student__mentor=mentor)
    if student:
        qs = qs.filter(student=student)
    if student_ids is not None:
        qs = qs.filter(student_id__in=list(student_ids))
    qs = qs.order_by("student_id", "-created_at", "-id")
    latest = {}
    for rec in qs:
        if rec.student_id not in latest:
            latest[rec.student_id] = rec
    return latest


def _upload_fail_student_ids(upload):
    return set(
        StudentResult.objects.filter(upload=upload, fail_flag=True).values_list("student_id", flat=True)
    )


def _require_superadmin(request):
    if not request.user.is_authenticated or request.session.get("mentor"):
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)
    if not is_superadmin_user(request.user):
        return JsonResponse({"ok": False, "msg": "SuperAdmin only"}, status=403)
    return None


def _normalize_whatsapp_phone(number):
    digits = re.sub(r"\D", "", str(number or ""))
    if not digits:
        return ""
    if len(digits) == 10:
        return f"91{digits}"
    return digits


def _ensure_subject_display_order(module):
    subjects = list(Subject.objects.filter(module=module, is_active=True).order_by("display_order", "name"))
    changed = False
    next_order = 1
    for s in subjects:
        if not s.display_order or s.display_order <= 0:
            s.display_order = next_order
            s.save(update_fields=["display_order"])
            changed = True
        next_order += 1
    return changed


def _latest_result_map_for_student(student, module):
    rows = {}
    subjects = Subject.objects.filter(module=module, is_active=True)
    for s in subjects:
        upload = (
            ResultUpload.objects.filter(module=module, subject=s)
            .order_by("-uploaded_at")
            .first()
        )
        if not upload:
            rows[s.id] = None
            continue
        rows[s.id] = StudentResult.objects.filter(upload=upload, student=student).first()
    return rows


def _fmt_mark(v):
    if v is None:
        return "-"
    try:
        if float(v).is_integer():
            return str(int(v))
        return str(round(float(v), 2))
    except Exception:
        return str(v)


def _module_display_name(variant, semester, batch):
    return f"{variant}_{semester} - Batch {batch}"


def _ensure_active_timetable(module):
    now_ts = timezone.now()
    candidate = (
        TimetableUpload.objects.filter(
            module=module,
            effective_from__isnull=False,
            effective_from__lte=now_ts,
        )
        .order_by("-effective_from", "-uploaded_at")
        .first()
    )
    if not candidate or candidate.is_active:
        return
    TimetableUpload.objects.filter(module=module).update(is_active=False)
    candidate.is_active = True
    candidate.save(update_fields=["is_active"])
    TimetableEntry.objects.filter(module=module).update(is_active=False)
    TimetableEntry.objects.filter(module=module, upload=candidate).update(is_active=True)


def _active_upload_for_module(module):
    return TimetableUpload.objects.filter(module=module, is_active=True).order_by("-uploaded_at").first()


def _dept_matches_module(module, dept_key):
    if not dept_key:
        return True
    dept_key = dept_key.upper()
    if dept_key not in {"FY1", "FY2", "FY3", "FY4", "FY5"}:
        return True
    variant = (getattr(module, "variant", "") or "").upper()
    name = (getattr(module, "name", "") or "").upper()
    return variant.startswith(dept_key) or name.startswith(dept_key) or f"{dept_key}_" in name


def _dept_label_from_module(module):
    for key in ("FY1", "FY2", "FY3", "FY4", "FY5"):
        if _dept_matches_module(module, key):
            return key
    return (getattr(module, "year_level", "") or "").upper() or "NA"


def _attendance_fully_marked_for_date(module, date_val):
    _ensure_active_timetable(module)
    if not _attendance_allowed_for_date(module, date_val):
        return True
    day_of_week = date_val.weekday()
    expected = set(
        TimetableEntry.objects.filter(module=module, day_of_week=day_of_week, is_active=True).values_list(
            "batch", "lecture_no"
        )
    )
    if not expected:
        return True
    marked = set(LectureSession.objects.filter(module=module, date=date_val).values_list("batch", "lecture_no"))
    return expected.issubset(marked)


def _attendance_fully_marked_for_range(module, start_date, end_date):
    cur = start_date
    while cur <= end_date:
        if _attendance_allowed_for_date(module, cur) and not _attendance_fully_marked_for_date(module, cur):
            return False
        cur += timedelta(days=1)
    return True


def _sif_marks_rows_for_student(student, module):
    subjects = ordered_subjects(module)
    practical_map = {
        pm.subject_id: pm
        for pm in StudentPracticalMark.objects.filter(module=module, student=student).select_related("subject")
    }
    result_map = _latest_result_map_for_student(student, module)

    rows = []
    for s in subjects:
        practical = practical_map.get(s.id)
        sr = result_map.get(s.id)
        attendance_val = _fmt_mark(practical.attendance_percentage) if practical and practical.attendance_percentage is not None else "-"
        short = (s.short_name or s.name).strip()
        t1 = sr.marks_t1 if sr and sr.marks_t1 is not None else (sr.marks_current if sr and sr.upload.test_name == "T1" else None)
        t2 = sr.marks_t2 if sr and sr.marks_t2 is not None else (sr.marks_current if sr and sr.upload.test_name == "T2" else None)
        t3 = sr.marks_t3 if sr and sr.marks_t3 is not None else (sr.marks_current if sr and sr.upload.test_name == "T3" else None)
        t4 = sr.marks_t4 if sr and sr.marks_t4 is not None else (sr.marks_current if sr and sr.upload.test_name == "T4" else None)
        total = sr.marks_total if sr else None
        rows.append(
            {
                "label": f"{short}-TH",
                "t1": _fmt_mark(t1),
                "t2": _fmt_mark(t2),
                "t3": _fmt_mark(t3),
                "t4": _fmt_mark(t4),
                "total": _fmt_mark(total),
                "attendance": attendance_val,
            }
        )
        rows.append(
            {
                "label": f"{short}-PR",
                "t1": "-",
                "t2": "-",
                "t3": "-",
                "t4": "-",
                "total": _fmt_mark(practical.pr_marks) if practical and practical.pr_marks is not None else "-",
                "attendance": attendance_val,
            }
        )
    return rows


def _result_report_text(test_name, subject_name, mentor_name, total, received, not_received, message_done):
    if test_name == "T1":
        rule = f"Less than 9 marks in {test_name}"
    elif test_name == "T2":
        rule = "Less than 9 marks in T2 & less than 18 in (T1+T2)"
    elif test_name == "T3":
        rule = "Less than 9 marks in T3 & less than 27 in (T1+T2+T3)"
    elif test_name == "T4":
        rule = "Less than 18 marks in SEE & less than 35 in (T1+T2+T3+SEE)"
    else:
        rule = "Less than 35 marks in REMEDIAL"

    return f"""📞Phone call done regarding failed in {subject_name} ({rule})
Name of Faculty- {mentor_name}
Total no of calls- {total:02d}
Received Calls - {received:02d}
Not received- {not_received:02d}
No of Message done as call not Received - {message_done:02d}"""


def _result_filter_config(test_name):
    test_name = (test_name or "").upper()
    if test_name == "T1":
        return {
            "current_key": "current_fail",
            "current_label": "T1<9",
            "total_key": "total_fail",
            "total_label": "till T1<9",
            "either_key": "either_fail",
            "either_label": "Either (T1<9 OR till T1<9)",
            "current_threshold": 9,
            "total_threshold": 9,
            "exam_col_label": "T1 marks /25",
            "total_col_label": "Total till T1 /25",
            "display_columns": [
                {"key": "marks_current", "label": "T1 marks /25"},
                {"key": "marks_total", "label": "Total till T1 /25"},
            ],
        }
    if test_name == "T2":
        return {
            "current_key": "current_fail",
            "current_label": "T2<9",
            "total_key": "total_fail",
            "total_label": "T1+T2<18",
            "either_key": "either_fail",
            "either_label": "Either (T2<9 OR T1+T2<18)",
            "current_threshold": 9,
            "total_threshold": 18,
            "exam_col_label": "T2 marks /25",
            "total_col_label": "T1+T2 /50",
            "display_columns": [
                {"key": "marks_t1", "label": "T1 marks /25"},
                {"key": "marks_current", "label": "T2 marks /25"},
                {"key": "marks_total", "label": "T1+T2 /50"},
            ],
        }
    if test_name == "T3":
        return {
            "current_key": "current_fail",
            "current_label": "T3<9",
            "total_key": "total_fail",
            "total_label": "T1+T2+T3<27",
            "either_key": "either_fail",
            "either_label": "Either (T3<9 OR T1+T2+T3<27)",
            "current_threshold": 9,
            "total_threshold": 27,
            "exam_col_label": "T3 marks /25",
            "total_col_label": "T1+T2+T3 /75",
            "display_columns": [
                {"key": "marks_t1", "label": "T1 marks /25"},
                {"key": "marks_t2", "label": "T2 marks /25"},
                {"key": "marks_current", "label": "T3 marks /25"},
                {"key": "marks_total", "label": "T1+T2+T3 /75"},
            ],
        }
    if test_name == "T4":
        return {
            "current_key": "current_fail",
            "current_label": "T4<18",
            "total_key": "total_fail",
            "total_label": "T1+T2+T3+T4<35",
            "either_key": "either_fail",
            "either_label": "Either (T4<18 OR T1+T2+T3+T4<35)",
            "current_threshold": 18,
            "total_threshold": 35,
            "exam_col_label": "T4 marks /50",
            "total_col_label": "T1+T2+T3+(T4/2) /100",
            "display_columns": [
                {"key": "marks_t1", "label": "T1 marks /25"},
                {"key": "marks_t2", "label": "T2 marks /25"},
                {"key": "marks_t3", "label": "T3 marks /25"},
                {"key": "marks_current", "label": "T4 marks /50"},
                {"key": "marks_t4_half", "label": "T4/2 /25"},
                {"key": "marks_total", "label": "T1+T2+T3+(T4/2) /100"},
            ],
        }
    return {
        "current_key": "current_fail",
        "current_label": "REM<35",
        "total_key": "total_fail",
        "total_label": "REM<35",
        "either_key": "either_fail",
        "either_label": "Either (REM<35)",
        "current_threshold": 35,
        "total_threshold": 35,
        "exam_col_label": "REM marks /100",
        "total_col_label": "Total till REM /100",
        "display_columns": [
            {"key": "marks_current", "label": "REM marks /100"},
            {"key": "marks_total", "label": "Total till REM /100"},
        ],
    }


def _format_parent_faculty_remark(text):
    raw = (text or "").strip()
    if not raw:
        return "-"
    if "PARENT::" in raw and "||FACULTY::" in raw:
        parent_part, faculty_part = raw.split("||FACULTY::", 1)
        parent_val = parent_part.replace("PARENT::", "", 1).strip() or "-"
        faculty_val = faculty_part.strip() or "-"
        return f"PARENT: {parent_val}\nFACULTY: {faculty_val}"
    return raw

# ---------------- LOGIN ----------------
def login_page(request):
    error = ""

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        # coordinator login
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            request.session.pop("mentor", None)
            _active_module(request)
            if is_superadmin_user(user):
                return redirect("/home/")
            return redirect("/reports/")

        # mentor login supports both legacy and new scheme:
        # - legacy: mentor@LJ123
        # - new: <mentor_short_name>@LJ123
        mentor = resolve_mentor_identity(username)
        if mentor:
            entered_username = (username or "").strip().lower()
            expected_short = f"{(mentor.name or '').strip().lower()}@LJ123".lower()
            expected_entered = f"{entered_username}@LJ123"
            entered_password_raw = (password or "").strip()
            entered_password = entered_password_raw.lower()
            cred = MentorPassword.objects.filter(mentor=mentor).first()
            custom_ok = bool(cred and cred.check_password(entered_password_raw))
            if custom_ok or entered_password in {expected_short, expected_entered, "mentor@lj123"}:
                request.session["mentor"] = mentor.name
                _active_module(request)
                return redirect("/mentor-dashboard/")

        error = "Invalid username or password"

    return render(request, "login.html", {"error": error})


@login_required
def manage_mentors(request):
    if request.session.get("mentor"):
        return redirect("/mentor-dashboard/")

    module = _active_module(request)
    _ensure_active_timetable(module)
    faculty_names = sorted(
        {
            (name or "").strip()
            for name in TimetableEntry.objects.filter(module=module, is_active=True)
            .exclude(faculty="")
            .values_list("faculty", flat=True)
        }
    )
    faculty_names = [n for n in faculty_names if n]
    for name in faculty_names:
        Mentor.objects.get_or_create(name=name)

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        mentor_id = request.POST.get("mentor_id")
        mentor = Mentor.objects.filter(id=mentor_id).first() if mentor_id else None
        module_mentor_ids = set(Student.objects.filter(module=module).values_list("mentor_id", flat=True))
        module_faculty_names = {n.lower() for n in faculty_names}
        if not mentor or (mentor.id not in module_mentor_ids and mentor.name.lower() not in module_faculty_names):
            messages.error(request, "Mentor not found for current module.")
            return redirect("/manage-mentors/")

        if action == "update_password":
            new_password = (request.POST.get("new_password") or "").strip()
            if len(new_password) < 6:
                messages.error(request, "Password must be at least 6 characters.")
            else:
                cred, _ = MentorPassword.objects.get_or_create(mentor=mentor, defaults={"password_hash": ""})
                cred.set_password(new_password)
                cred.save(update_fields=["password_hash", "updated_at"])
                messages.success(request, f"Password updated for mentor {mentor.name}.")
        elif action == "reset_default":
            MentorPassword.objects.filter(mentor=mentor).delete()
            messages.success(request, f"Password reset to default rule for mentor {mentor.name}.")
        else:
            messages.error(request, "Invalid action.")
        return redirect("/manage-mentors/")

    mentors = (
        Mentor.objects.filter(Q(student__module=module) | Q(name__in=faculty_names))
        .distinct()
        .annotate(student_count=Count("student", filter=Q(student__module=module)))
        .order_by("name")
    )
    cred_map = {c.mentor_id: c for c in MentorPassword.objects.filter(mentor__in=mentors)}
    rows = []
    for m in mentors:
        rows.append(
            {
                "mentor": m,
                "student_count": getattr(m, "student_count", 0),
                "has_custom_password": m.id in cred_map,
            }
        )

    return render(
        request,
        "manage_mentors.html",
        {
            "rows": rows,
            "module": module,
        },
    )


@login_required
def superadmin_home(request):
    if request.session.get("mentor"):
        return redirect("/mentor-dashboard/")
    if not is_superadmin_user(request.user):
        return redirect("/reports/")

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "create":
            coordinator_name = (request.POST.get("coordinator_name") or "").strip()
            username = (request.POST.get("username") or "").strip()
            password = (request.POST.get("password") or "").strip()
            module_ids = request.POST.getlist("module_ids")
            if not coordinator_name or not username or not password:
                messages.error(request, "Coordinator name, username and password are required.")
            elif User.objects.filter(username__iexact=username).exists():
                messages.error(request, "Coordinator username already exists.")
            else:
                coordinator = User.objects.create_user(
                    username=username,
                    password=password,
                    first_name=coordinator_name,
                    is_active=True,
                    is_staff=False,
                    is_superuser=False,
                )
                modules = list(AcademicModule.objects.filter(id__in=module_ids))
                if not modules:
                    coordinator.delete()
                    messages.error(request, "Select at least one module.")
                else:
                    CoordinatorModuleAccess.objects.bulk_create(
                        [CoordinatorModuleAccess(coordinator=coordinator, module=m) for m in modules],
                        ignore_conflicts=True,
                    )
                    messages.success(request, f"Coordinator created: {coordinator.username}")

        elif action == "update":
            coord_id = request.POST.get("coordinator_id")
            username = (request.POST.get("username") or "").strip()
            coordinator_name = (request.POST.get("coordinator_name") or "").strip()
            new_password = (request.POST.get("new_password") or "").strip()
            module_ids = request.POST.getlist("module_ids")
            active_val = (request.POST.get("is_active") or "").strip() == "1"
            coordinator = User.objects.filter(id=coord_id, is_superuser=False).first()
            if not coordinator:
                messages.error(request, "Coordinator not found.")
            else:
                if not username:
                    messages.error(request, "Username is required.")
                    return redirect("/home/")
                username_exists = User.objects.filter(username__iexact=username).exclude(id=coordinator.id).exists()
                if username_exists:
                    messages.error(request, "Username already exists.")
                    return redirect("/home/")
                modules = list(AcademicModule.objects.filter(id__in=module_ids))
                if not modules:
                    messages.error(request, "Select at least one module.")
                    return redirect("/home/")
                coordinator.username = username
                coordinator.first_name = coordinator_name
                coordinator.is_active = active_val
                if new_password:
                    coordinator.set_password(new_password)
                coordinator.save()
                CoordinatorModuleAccess.objects.filter(coordinator=coordinator).delete()
                CoordinatorModuleAccess.objects.bulk_create(
                    [CoordinatorModuleAccess(coordinator=coordinator, module=m) for m in modules],
                    ignore_conflicts=True,
                )
                messages.success(request, f"Coordinator updated: {coordinator.username}")

        elif action == "delete":
            coord_id = request.POST.get("coordinator_id")
            coordinator = User.objects.filter(id=coord_id, is_superuser=False).first()
            if not coordinator:
                messages.error(request, "Coordinator not found.")
            else:
                CoordinatorModuleAccess.objects.filter(coordinator=coordinator).delete()
                coordinator.delete()
                messages.success(request, "Coordinator deleted.")
        elif action == "change_super_password":
            current_password = (request.POST.get("current_password") or "").strip()
            new_password = (request.POST.get("new_password") or "").strip()
            if not request.user.check_password(current_password):
                messages.error(request, "Current password is incorrect.")
            elif len(new_password) < 8:
                messages.error(request, "New password must be at least 8 characters.")
            else:
                request.user.set_password(new_password)
                request.user.save(update_fields=["password"])
                update_session_auth_hash(request, request.user)
                messages.success(request, "Password updated.")

        return redirect("/home/")

    modules = list(AcademicModule.objects.all().order_by("-id"))
    active_modules = [m for m in modules if m.is_active]
    active_module_ids = [m.id for m in active_modules]
    coordinators = User.objects.filter(is_superuser=False).order_by("username")
    coordinator_rows = []
    for c in coordinators:
        mapped = list(
            AcademicModule.objects.filter(coordinator_accesses__coordinator=c)
            .distinct()
            .order_by("-id")
        )
        coordinator_rows.append({"user": c, "modules": mapped})

    stats = {
        "total_coordinators": User.objects.filter(
            is_superuser=False,
            is_active=True,
            module_accesses__module__in=active_modules,
        ).distinct().count(),
        "total_modules": len(active_modules),
        "total_mentors": Mentor.objects.filter(student__module_id__in=active_module_ids).distinct().count(),
        "total_students": Student.objects.filter(module_id__in=active_module_ids).count(),
    }

    module_summary = []
    for m in modules:
        student_qs = Student.objects.filter(module=m)
        module_summary.append(
            {
                "module": m,
                "students": student_qs.count(),
                "mentors": Mentor.objects.filter(student__module=m).distinct().count(),
                "coordinators": User.objects.filter(is_superuser=False, module_accesses__module=m).distinct().count(),
            }
        )

    mentor_summary = (
        Mentor.objects.annotate(
            total_students=Count("student", filter=Q(student__module_id__in=active_module_ids), distinct=True),
            total_modules=Count("student__module", filter=Q(student__module_id__in=active_module_ids), distinct=True),
        )
        .filter(total_students__gt=0)
        .order_by("name")
    )

    student_report_rows = []
    for m in modules:
        student_qs = Student.objects.filter(module=m)
        student_report_rows.append(
            {
                "module_name": m.name,
                "students": student_qs.count(),
                "mentors": Mentor.objects.filter(student__module=m).distinct().count(),
                "attendance_rows": Attendance.objects.filter(student__module=m).count(),
                "result_uploads": ResultUpload.objects.filter(module=m).count(),
                "practical_uploads": PracticalMarkUpload.objects.filter(module=m).count(),
            }
        )

    return render(
        request,
        "home.html",
        {
            "modules": modules,
            "coordinator_rows": coordinator_rows,
            "stats": stats,
            "module_summary": module_summary,
            "mentor_summary": mentor_summary,
            "student_report_rows": student_report_rows,
        },
    )


# ---------------- STUDENT MASTER ----------------
@login_required
def upload_students(request):
    module = _active_module(request)

    message = ""
    skipped_rows = []
    form = UploadFileForm()

    if request.method == 'POST':
        if request.POST.get("action") == "clear_module_students":
            deleted_count, _ = Student.objects.filter(module=module).delete()
            message = f"Deleted student master data for module '{module.name}'. Records removed: {deleted_count}"
        else:
            form = UploadFileForm(request.POST, request.FILES)
            if form.is_valid():
                file = request.FILES['file']
                try:
                    added, updated, skipped, skipped_rows = import_students_from_excel(file, module)
                    message = f"Added: {added} | Updated: {updated} | Skipped: {skipped}"
                except Exception as e:
                    message = f"Upload failed: {str(e)}"
            else:
                message = "Please select a file to upload."
    else:
        form = UploadFileForm()

    students = Student.objects.select_related("mentor").filter(module=module).order_by("roll_no")

    return render(request, 'upload.html', {
        'form': form,
        'message': message,
        'students': students,
        'module': module,
        'skipped_rows': skipped_rows[:200],
    })

# ---------------- ATTENDANCE VIEW & UPLOAD ----------------
@require_http_methods(["GET","POST"])
def upload_attendance(request):
    module = _active_module(request)

    # -------- OPEN PAGE --------
    if request.method == "GET":
        return render(request, "upload_attendance.html")

    # -------- AJAX UPLOAD --------
    try:
        week_no = int(request.POST.get('week'))
        rule = request.POST.get('rule')
        weekly_file = request.FILES.get('weekly_file')
        overall_file = request.FILES.get('overall_file')

        # Week-1 has no overall
        if week_no == 1:
            overall_file = None

        # lock check
        if WeekLock.objects.filter(module=module, week_no=week_no, locked=True).exists():
            return JsonResponse({
                "ok": False,
                "msg": f"Week {week_no} is LOCKED. Upload not allowed."
            })

        # import
        count = import_attendance(weekly_file, overall_file, week_no, module, rule)
        AttendanceWeekMeta.objects.update_or_create(
            module=module,
            week_no=week_no,
            defaults={"source": AttendanceWeekMeta.SOURCE_MANUAL},
        )

        # mentor-wise counts
        mentor_stats = list(
            CallRecord.objects.filter(week_no=week_no, student__module=module)
            .values("student__mentor__name")
            .annotate(total=Count("student", distinct=True))
            .order_by("student__mentor__name")
        )

        total_calls = sum(m["total"] for m in mentor_stats)

        return JsonResponse({
            "ok": True,
            "msg": f"{count} students require follow-up calls for Week {week_no}",
            "week": week_no,
            "mentor_stats": mentor_stats,
            "total_calls": total_calls
        })

    except Exception as e:
        return JsonResponse({
            "ok": False,
            "msg": str(e)
        })


def _process_result_upload(module, username, test_name, subject_id, upload_mode, bulk_confirm, file_obj, progress_cb=None, cancel_cb=None):
    is_all_tests = test_name == "ALL_EXAMS"
    is_all_subjects = str(subject_id).upper() == "ALL"

    if is_all_tests and is_all_subjects:
        summary = import_compiled_bulk_all(
            file_obj,
            username,
            module=module,
            progress_cb=progress_cb,
            cancel_cb=cancel_cb,
        )
        return {
            "ok": True,
            "msg": (
                f"Bulk replace completed. Created uploads: {summary['uploads_created']}. "
                f"Rows matched: {summary['rows_matched']}. Failed calls: {summary['rows_failed']}."
            ),
            "test_name": "ALL_EXAMS",
            "subject_name": "ALL",
            "upload_id": "",
            "mentor_stats": [],
            "total_calls": summary["rows_failed"],
            "upload_mode": upload_mode,
            "found_subjects": summary.get("found_subjects", []),
            "used_subject": "ALL",
        }

    subject = Subject.objects.filter(id=subject_id, module=module, is_active=True).first()
    if not subject:
        raise Exception("Invalid subject")
    if subject.result_format == Subject.FORMAT_T4_ONLY and test_name != "T4":
        raise Exception("This subject is configured as Only T4. Please upload in T4.")

    upload, _ = ResultUpload.objects.update_or_create(
        module=module,
        test_name=test_name,
        subject=subject,
        defaults={"uploaded_by": username},
    )

    if upload_mode == "compiled":
        summary = import_compiled_result_sheet(file_obj, upload, progress_cb=progress_cb, cancel_cb=cancel_cb)
    else:
        summary = import_result_sheet(file_obj, upload, progress_cb=progress_cb, cancel_cb=cancel_cb)

    mentor_stats = list(
        ResultCallRecord.objects.filter(upload=upload)
        .values("student__mentor__name")
        .annotate(total=Count("student", distinct=True))
        .order_by("student__mentor__name")
    )
    total_calls = sum(m["total"] for m in mentor_stats)
    return {
        "ok": True,
        "msg": (
            f"Processed {summary['rows_total']} rows. "
            f"Matched: {summary['rows_matched']}. "
            f"Fail calls generated: {summary['rows_failed']}."
        ),
        "test_name": test_name,
        "subject_name": subject.name,
        "upload_id": upload.id,
        "mentor_stats": mentor_stats,
        "total_calls": total_calls,
        "upload_mode": upload_mode,
        "found_subjects": summary.get("found_subjects", []),
        "used_subject": summary.get("used_subject", ""),
    }


def _run_result_upload_job(job_id, module_id, username, test_name, subject_id, upload_mode, bulk_confirm, temp_path):
    close_old_connections()
    try:
        job = ResultUploadJob.objects.filter(job_id=job_id).first()
        if not job:
            return
        job.status = ResultUploadJob.STATUS_RUNNING
        job.message = "Reading marks and preparing result call list..."
        job.save(update_fields=["status", "message", "updated_at"])

        def cancel_cb():
            return ResultUploadJob.objects.filter(job_id=job_id, cancel_requested=True).exists()

        def progress_cb(current, total, enrollment, student_name, message):
            ResultUploadJob.objects.filter(job_id=job_id).update(
                progress_current=current or 0,
                progress_total=total or 0,
                current_enrollment=(enrollment or ""),
                current_student_name=(student_name or ""),
                message=(message or "Processing result upload..."),
                updated_at=timezone.now(),
            )

        module = AcademicModule.objects.filter(id=module_id).first()
        if not module:
            raise Exception("Module not found")

        with open(temp_path, "rb") as f:
            payload = _process_result_upload(
                module=module,
                username=username,
                test_name=test_name,
                subject_id=subject_id,
                upload_mode=upload_mode,
                bulk_confirm=bulk_confirm,
                file_obj=f,
                progress_cb=progress_cb,
                cancel_cb=cancel_cb,
            )

        if cancel_cb():
            ResultUploadJob.objects.filter(job_id=job_id).update(
                status=ResultUploadJob.STATUS_CANCELLED,
                message="Upload cancelled.",
                updated_at=timezone.now(),
            )
            return

        ResultUploadJob.objects.filter(job_id=job_id).update(
            status=ResultUploadJob.STATUS_COMPLETED,
            message="Upload completed.",
            result_payload=payload,
            progress_current=1,
            progress_total=1,
            updated_at=timezone.now(),
        )
    except Exception as exc:
        cancelled = ResultUploadJob.objects.filter(job_id=job_id, cancel_requested=True).exists()
        ResultUploadJob.objects.filter(job_id=job_id).update(
            status=(ResultUploadJob.STATUS_CANCELLED if cancelled else ResultUploadJob.STATUS_FAILED),
            message=("Upload cancelled." if cancelled else str(exc)),
            updated_at=timezone.now(),
        )
    finally:
        try:
            if temp_path and os.path.exists(temp_path):
                os.remove(temp_path)
        except Exception:
            pass
        close_old_connections()


@login_required
@require_http_methods(["GET"])
def upload_results_progress(request, job_id):
    if "mentor" in request.session:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=403)
    module = _active_module(request)
    job = ResultUploadJob.objects.filter(job_id=job_id, module=module).first()
    if not job:
        return JsonResponse({"ok": False, "msg": "Job not found"}, status=404)
    return JsonResponse(
        {
            "ok": True,
            "job_id": job.job_id,
            "status": job.status,
            "message": job.message or "",
            "progress_current": job.progress_current,
            "progress_total": job.progress_total,
            "current_enrollment": job.current_enrollment or "",
            "current_student_name": job.current_student_name or "",
            "result": job.result_payload or {},
        }
    )


def _to_ist_datetime_text(dt):
    if not dt:
        return "-", "-"
    local_dt = dt.astimezone(IST)
    return local_dt.strftime("%d-%m-%Y"), local_dt.strftime("%I:%M %p")


def _call_status_text(status):
    if status == "received":
        return "Received"
    if status == "not_received":
        return "Not Received"
    return "Pending"


def _exam_name_for_sif(test_name):
    test = (test_name or "").upper()
    if test == "T1":
        return "T1"
    if test == "T2":
        return "T2 / (T1+T2)"
    if test == "T3":
        return "T3 / (T1+T2+T3)"
    if test == "T4":
        return "T4 / (T1+T2+T3+T4)"
    if test == "REMEDIAL":
        return "REM"
    return test_name or "-"


def _result_thresholds(test_name):
    test = (test_name or "").upper()
    if test == "T1":
        return 9, 9
    if test == "T2":
        return 9, 18
    if test == "T3":
        return 9, 27
    if test == "T4":
        return 18, 35
    return 35, 35


def _test_sort_key(test_name):
    order = {"T1": 1, "T2": 2, "T3": 3, "T4": 4, "REMEDIAL": 5}
    return order.get((test_name or "").upper(), 99)


def _subject_sort_key(subject_name):
    name = (subject_name or "").strip().lower()
    priority = ["mathematics", "java", "physics", "software engineering"]
    for idx, token in enumerate(priority, start=1):
        if token in name:
            return idx
    return 99


@login_required
@require_http_methods(["POST"])
def upload_results_cancel(request, job_id):
    if "mentor" in request.session:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=403)
    module = _active_module(request)
    updated = ResultUploadJob.objects.filter(
        job_id=job_id,
        module=module,
        status__in=[ResultUploadJob.STATUS_QUEUED, ResultUploadJob.STATUS_RUNNING],
    ).update(cancel_requested=True, message="Cancelling upload...", updated_at=timezone.now())
    if not updated:
        return JsonResponse({"ok": False, "msg": "Upload is already finished."})
    return JsonResponse({"ok": True, "msg": "Cancel requested."})


@login_required
@require_http_methods(["GET", "POST"])
def upload_results(request):
    if "mentor" in request.session:
        return redirect("/mentor-dashboard/")
    module = _active_module(request)

    if request.method == "GET":
        return render(
            request,
            "upload_results.html",
            {
                "tests": TEST_NAMES,
                "subjects": Subject.objects.filter(module=module, is_active=True).order_by("name"),
            },
        )

    try:
        test_name = (request.POST.get("test_name") or "").strip().upper()
        subject_id = request.POST.get("subject_id")
        upload_mode = (request.POST.get("upload_mode") or "subject").strip().lower()
        bulk_confirm = (request.POST.get("bulk_confirm") or "").strip().lower()
        file_obj = request.FILES.get("result_file")

        allowed_tests = set(TEST_NAMES) | {"ALL_EXAMS"}
        if test_name not in allowed_tests:
            return JsonResponse({"ok": False, "msg": "Invalid test name"})
        if not subject_id:
            return JsonResponse({"ok": False, "msg": "Subject is required"})
        if not file_obj:
            return JsonResponse({"ok": False, "msg": "Result file is required"})
        if upload_mode not in {"subject", "compiled"}:
            return JsonResponse({"ok": False, "msg": "Invalid upload mode"})

        is_all_tests = test_name == "ALL_EXAMS"
        is_all_subjects = str(subject_id).upper() == "ALL"
        if is_all_tests != is_all_subjects:
            return JsonResponse({"ok": False, "msg": "Please select BOTH ALL EXAMS and ALL subjects for bulk upload."})
        if is_all_tests and is_all_subjects and upload_mode != "compiled":
            return JsonResponse({"ok": False, "msg": "ALL_EXAMS + ALL subjects is supported only for Compiled sheet mode."})
        if is_all_tests and is_all_subjects and bulk_confirm != "yes":
            return JsonResponse({"ok": False, "msg": "Bulk upload cancelled. Please select YES to replace old uploads."})

        suffix = os.path.splitext(getattr(file_obj, "name", ""))[1] or ".xlsx"
        fd, temp_path = tempfile.mkstemp(prefix="result_upload_", suffix=suffix, dir=tempfile.gettempdir())
        with os.fdopen(fd, "wb") as tmp:
            for chunk in file_obj.chunks():
                tmp.write(chunk)

        job_key = str(uuid.uuid4())
        job = ResultUploadJob.objects.create(
            job_id=job_key,
            module=module,
            created_by=request.user.username,
            status=ResultUploadJob.STATUS_QUEUED,
            message="Upload queued...",
        )

        t = threading.Thread(
            target=_run_result_upload_job,
            kwargs={
                "job_id": job.job_id,
                "module_id": module.id,
                "username": request.user.username,
                "test_name": test_name,
                "subject_id": str(subject_id),
                "upload_mode": upload_mode,
                "bulk_confirm": bulk_confirm,
                "temp_path": temp_path,
            },
            daemon=True,
        )
        t.start()
        return JsonResponse({"ok": True, "job_id": job.job_id})
    except Exception as e:
        return JsonResponse({"ok": False, "msg": str(e)})


@login_required
def view_results(request):
    if "mentor" in request.session:
        return redirect("/mentor-dashboard/")
    module = _active_module(request)

    subjects = list(Subject.objects.filter(module=module, is_active=True).order_by("name"))
    selected_test = (request.GET.get("test") or "").upper()
    selected_subject = request.GET.get("subject")
    selected_filter = request.GET.get("filter", "either_fail")
    mentor_filter = request.GET.get("mentor", "")
    sort = request.GET.get("sort", "roll")
    direction = request.GET.get("dir", "asc")

    latest_upload = ResultUpload.objects.filter(module=module).select_related("subject").order_by("-uploaded_at").first()
    if not selected_test and not selected_subject and latest_upload:
        selected_test = latest_upload.test_name
        selected_subject = str(latest_upload.subject_id)

    if selected_test not in TEST_NAMES:
        selected_test = latest_upload.test_name if latest_upload else "T1"

    if not selected_subject and subjects:
        selected_subject = str(subjects[0].id)

    uploads = ResultUpload.objects.filter(module=module).select_related("subject").order_by("test_name", "subject__name")
    upload_map = {(u.test_name, str(u.subject_id)): u for u in uploads}
    selected_upload = upload_map.get((selected_test, str(selected_subject))) if selected_subject else None
    if not selected_upload and latest_upload and not request.GET.get("test") and not request.GET.get("subject"):
        selected_upload = latest_upload
        selected_test = latest_upload.test_name
        selected_subject = str(latest_upload.subject_id)

    matrix_rows = []
    for test in TEST_NAMES:
        cells = []
        for s in subjects:
            up = upload_map.get((test, str(s.id)))
            applicable = True
            if s.result_format == Subject.FORMAT_T4_ONLY and test != "T4":
                applicable = False
            cells.append({"subject": s, "upload": up, "applicable": applicable})
        matrix_rows.append({"test": test, "cells": cells})

    config = _result_filter_config(selected_test)
    records = []
    rows = []
    total_count = 0
    mentor_counts = []
    upload_waiting = False

    if selected_subject and not selected_upload:
        upload_waiting = True

    if selected_upload:
        base_qs = (
            StudentResult.objects.filter(upload=selected_upload)
            .select_related("student", "student__mentor", "upload", "upload__subject")
        )
        if selected_filter == config["current_key"]:
            base_qs = base_qs.filter(marks_current__lt=config["current_threshold"])
        elif selected_filter == config["total_key"]:
            base_qs = base_qs.filter(marks_total__lt=config["total_threshold"])
        elif selected_filter == config["either_key"]:
            base_qs = base_qs.filter(
                Q(marks_current__lt=config["current_threshold"]) |
                Q(marks_total__lt=config["total_threshold"])
            )

        mentor_counts = (
            base_qs.values("student__mentor__name")
            .annotate(c=Count("id"))
            .order_by("student__mentor__name")
        )
        total_count_all = base_qs.count()

        qs = base_qs
        if mentor_filter:
            qs = qs.filter(student__mentor__name=mentor_filter)

        sort_map = {
            "roll": "student__roll_no",
            "enroll": "student__enrollment",
            "name": "student__name",
            "mentor": "student__mentor__name",
            "exam": "marks_current",
            "total": "marks_total",
        }
        order = sort_map.get(sort, "student__roll_no")
        if direction == "desc":
            order = "-" + order
        records = qs.order_by(order)
        total_count = records.count()

        # Build previous-upload comparison maps for changed historical marks (same subject only).
        prev_mark_map = {}
        for prev_test in ["T1", "T2", "T3"]:
            prev_upload = (
                ResultUpload.objects.filter(module=module, test_name=prev_test, subject_id=selected_upload.subject_id)
                .order_by("-uploaded_at")
                .first()
            )
            if not prev_upload:
                continue
            prev_rows = StudentResult.objects.filter(upload=prev_upload).values("student_id", "marks_current")
            prev_mark_map[prev_test] = {r["student_id"]: r["marks_current"] for r in prev_rows}

        display_columns = config["display_columns"]
        for r in records:
            row_cells = []
            for col in display_columns:
                key = col["key"]
                value = None
                if key == "marks_t4_half":
                    value = (r.marks_current / 2.0) if r.marks_current is not None else None
                else:
                    value = getattr(r, key, None)

                changed = False
                hover = ""
                if selected_test in {"T2", "T3", "T4"} and key in {"marks_t1", "marks_t2", "marks_t3"}:
                    ref_test = "T1" if key == "marks_t1" else ("T2" if key == "marks_t2" else "T3")
                    prev_value = prev_mark_map.get(ref_test, {}).get(r.student_id)
                    if prev_value is not None and value is not None and float(prev_value) != float(value):
                        changed = True
                        hover = f"Previous {ref_test}: {prev_value}"

                row_cells.append(
                    {
                        "key": key,
                        "label": col["label"],
                        "value": value,
                        "is_changed": changed,
                        "hover": hover,
                    }
                )

            rows.append(
                {
                    "roll_no": r.student.roll_no,
                    "enrollment": r.enrollment,
                    "name": r.student.name,
                    "mentor": r.student.mentor.name,
                    "cells": row_cells,
                }
            )
    else:
        total_count_all = 0

    return render(
        request,
        "view_results.html",
        {
            "tests": TEST_NAMES,
            "subjects": subjects,
            "matrix_rows": matrix_rows,
            "selected_test": selected_test,
            "selected_subject": str(selected_subject or ""),
            "selected_upload": selected_upload,
            "upload_waiting": upload_waiting,
            "records": records,
            "rows": rows,
            "mentor_counts": mentor_counts,
            "total_count": total_count,
            "filter": selected_filter,
            "filter_current_key": config["current_key"],
            "filter_current_label": config["current_label"],
            "filter_total_key": config["total_key"],
            "filter_total_label": config["total_label"],
            "filter_either_key": config["either_key"],
            "filter_either_label": config["either_label"],
            "exam_col_label": config["exam_col_label"],
            "total_col_label": config["total_col_label"],
            "display_columns": config["display_columns"],
            "table_colspan": 4 + len(config["display_columns"]),
            "current_threshold": config["current_threshold"],
            "total_threshold": config["total_threshold"],
            "mentor_filter": mentor_filter,
            "total_count_all": total_count_all,
            "sort": sort,
            "dir": direction,
            "dir_roll": next_dir(sort, direction, "roll"),
            "dir_enroll": next_dir(sort, direction, "enroll"),
            "dir_name": next_dir(sort, direction, "name"),
            "dir_mentor": next_dir(sort, direction, "mentor"),
            "dir_exam": next_dir(sort, direction, "exam"),
            "dir_total": next_dir(sort, direction, "total"),
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def view_practical_marks(request):
    if "mentor" in request.session:
        return redirect("/mentor-dashboard/")
    module = _active_module(request)
    msg = ""
    if request.method == "POST":
        try:
            f = request.FILES.get("practical_file")
            if not f:
                raise Exception("Please select practical marks file.")
            summary = import_practical_marks(f, module=module, uploaded_by=request.user.username)
            msg = (
                f"Uploaded practical marks ({summary.get('mode','-')} mode). "
                f"Rows: {summary['rows_total']}, matched: {summary['rows_matched']}."
            )
        except Exception as exc:
            msg = f"Upload failed: {exc}"

    subjects = ordered_subjects(module)
    students = list(Student.objects.filter(module=module).select_related("mentor").order_by("roll_no", "name"))
    marks_qs = StudentPracticalMark.objects.filter(module=module).select_related("subject", "student")
    mark_map = {(m.student_id, m.subject_id): m for m in marks_qs}
    rows = []
    for s in students:
        row = {"student": s, "cells": []}
        for subj in subjects:
            pm = mark_map.get((s.id, subj.id))
            pr = _fmt_mark(pm.pr_marks) if pm and pm.pr_marks is not None else "-"
            att = _fmt_mark(pm.attendance_percentage) if pm and pm.attendance_percentage is not None else "-"
            row["cells"].append({"subject": subj, "pr": pr, "att": att})
        rows.append(row)

    latest_upload = PracticalMarkUpload.objects.filter(module=module).order_by("-uploaded_at").first()
    return render(
        request,
        "view_practical_marks.html",
        {
            "message": msg,
            "subjects": subjects,
            "rows": rows,
            "latest_upload": latest_upload,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def sif_marks_template(request):
    if "mentor" in request.session:
        return redirect("/mentor-dashboard/")
    module = _active_module(request)
    _ensure_subject_display_order(module)
    students = list(Student.objects.filter(module=module).select_related("mentor").order_by("roll_no", "name"))
    selected_enrollment = request.GET.get("enrollment") or (students[0].enrollment if students else "")
    selected_student = Student.objects.filter(module=module, enrollment=selected_enrollment).first() if selected_enrollment else None

    lock_obj, _ = SifMarksLock.objects.get_or_create(module=module)
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "lock":
            lock_obj.locked = True
            lock_obj.locked_by = request.user.username
            lock_obj.locked_at = timezone.now()
            lock_obj.save(update_fields=["locked", "locked_by", "locked_at"])
            messages.success(request, "SIF marks view locked.")
        elif action == "unlock":
            lock_obj.locked = False
            lock_obj.save(update_fields=["locked"])
            messages.success(request, "SIF marks view unlocked.")
        elif action in {"move_up", "move_down"}:
            if lock_obj.locked:
                messages.error(request, "Unlock first to change sequence.")
            else:
                sid = request.POST.get("subject_id")
                subj = Subject.objects.filter(module=module, id=sid, is_active=True).first()
                if subj:
                    subjects = ordered_subjects(module)
                    ids = [s.id for s in subjects]
                    try:
                        idx = ids.index(subj.id)
                    except ValueError:
                        idx = -1
                    if idx >= 0:
                        swap_idx = idx - 1 if action == "move_up" else idx + 1
                        if 0 <= swap_idx < len(subjects):
                            a = subjects[idx]
                            b = subjects[swap_idx]
                            a.display_order, b.display_order = b.display_order, a.display_order
                            a.save(update_fields=["display_order"])
                            b.save(update_fields=["display_order"])
        return redirect(f"/sif-marks-template/?enrollment={selected_enrollment}")

    marks_rows = _sif_marks_rows_for_student(selected_student, module) if selected_student else []
    return render(
        request,
        "sif_marks_template.html",
        {
            "students": students,
            "selected_student": selected_student,
            "selected_enrollment": selected_enrollment,
            "marks_rows": marks_rows,
            "lock_obj": lock_obj,
            "subjects_ordered": ordered_subjects(module),
        },
    )


@login_required
def subjects_page(request):
    if "mentor" in request.session:
        return redirect("/mentor-dashboard/")
    module = _active_module(request)
    subjects = list(Subject.objects.filter(module=module).order_by("name"))
    subject_choices = []
    subject_key_map = {}
    for s in subjects:
        for val in {(s.name or "").strip(), (s.short_name or "").strip()}:
            if not val:
                continue
            subject_choices.append(val)
            subject_key_map[_norm_subject_key(val)] = val
    subject_choices = sorted({v for v in subject_choices if v})
    is_superadmin = is_superadmin_user(getattr(request, "user", None))
    if is_superadmin:
        for s in subjects:
            template, _ = SubjectTemplate.objects.update_or_create(
                name=s.name,
                defaults={
                    "short_name": s.short_name or s.name,
                    "has_theory": s.has_theory,
                    "has_practical": s.has_practical,
                    "result_format": s.result_format,
                    "is_active": True,
                },
            )
            if s.source_template_id != template.id:
                s.source_template = template
                s.save(update_fields=["source_template"])
    templates = list(SubjectTemplate.objects.filter(is_active=True).order_by("name"))
    selected_template_ids = {s.source_template_id for s in subjects if s.source_template_id}
    if not selected_template_ids:
        template_name_map = {t.name.lower(): t.id for t in templates}
        for s in subjects:
            t_id = template_name_map.get((s.name or "").strip().lower())
            if t_id:
                selected_template_ids.add(t_id)
    alias_map = _subject_alias_map(module)
    known_keys = set(subject_key_map.keys()) | set(alias_map.keys())
    seen_subjects = {
        (v or "").strip()
        for v in TimetableEntry.objects.filter(module=module).values_list("subject", flat=True)
        if (v or "").strip()
    }
    seen_subjects |= {
        (v or "").strip()
        for v in LectureSession.objects.filter(module=module).values_list("subject", flat=True)
        if (v or "").strip()
    }
    alias_suggestions = []
    for label in sorted(seen_subjects):
        if _norm_subject_key(label) not in known_keys:
            alias_suggestions.append(label)
    alias_rows = list(SubjectAlias.objects.filter(module=module, is_active=True).order_by("alias"))
    alias_by_subject = {}
    for subject in subjects:
        keys = {
            _norm_subject_key(subject.name),
            _norm_subject_key(subject.short_name),
        }
        keys = {k for k in keys if k}
        matches = [a.alias for a in alias_rows if _norm_subject_key(a.canonical) in keys]
        alias_display = ", ".join(sorted({m for m in matches if m}))
        alias_by_subject[subject.id] = alias_display
        subject.alias_display = alias_display
    return render(
        request,
        "subjects.html",
        {
            "subjects": subjects,
            "available_templates": templates,
            "selected_template_ids": selected_template_ids,
            "is_superadmin": is_superadmin,
            "format_full": Subject.FORMAT_FULL,
            "format_t4_only": Subject.FORMAT_T4_ONLY,
            "aliases": SubjectAlias.objects.filter(module=module).order_by("alias"),
            "alias_suggestions": alias_suggestions,
            "subject_choices": subject_choices,
            "alias_by_subject": alias_by_subject,
        },
    )


@login_required
@require_http_methods(["POST"])
def add_subject_alias(request):
    if "mentor" in request.session:
        return redirect("/mentor-dashboard/")
    module = _active_module(request)
    alias = (request.POST.get("alias") or "").strip()
    canonical = (request.POST.get("canonical") or "").strip()
    apply_all = bool(request.POST.get("apply_all"))
    if not alias or not canonical:
        messages.error(request, "Alias and canonical subject are required.")
        return redirect("/subjects/")
    targets = [module]
    if is_superadmin_user(getattr(request, "user", None)) and apply_all:
        targets = list(AcademicModule.objects.filter(is_active=True))
    for target in targets:
        existing = SubjectAlias.objects.filter(module=target, alias__iexact=alias).first()
        if existing:
            existing.alias = alias
            existing.canonical = canonical
            existing.is_active = True
            existing.save(update_fields=["alias", "canonical", "is_active"])
        else:
            SubjectAlias.objects.create(module=target, alias=alias, canonical=canonical, is_active=True)
    messages.success(request, "Subject alias saved.")
    return redirect("/subjects/")


@login_required
@require_http_methods(["POST"])
def update_subject_alias(request, alias_id):
    if "mentor" in request.session:
        return redirect("/mentor-dashboard/")
    module = _active_module(request)
    alias_obj = SubjectAlias.objects.filter(id=alias_id).first()
    if not alias_obj:
        messages.error(request, "Alias not found.")
        return redirect("/subjects/")
    if not is_superadmin_user(getattr(request, "user", None)) and alias_obj.module_id != module.id:
        messages.error(request, "Alias not allowed for this module.")
        return redirect("/subjects/")
    alias = (request.POST.get("alias") or "").strip()
    canonical = (request.POST.get("canonical") or "").strip()
    is_active = bool(request.POST.get("is_active"))
    if not alias or not canonical:
        messages.error(request, "Alias and canonical subject are required.")
        return redirect("/subjects/")
    conflict = SubjectAlias.objects.filter(module=alias_obj.module, alias__iexact=alias).exclude(id=alias_obj.id).exists()
    if conflict:
        messages.error(request, "Another alias already exists with this name.")
        return redirect("/subjects/")
    alias_obj.alias = alias
    alias_obj.canonical = canonical
    alias_obj.is_active = is_active
    alias_obj.save(update_fields=["alias", "canonical", "is_active"])
    messages.success(request, "Subject alias updated.")
    return redirect("/subjects/")


@login_required
@require_http_methods(["POST"])
def delete_subject_alias(request, alias_id):
    if "mentor" in request.session:
        return redirect("/mentor-dashboard/")
    module = _active_module(request)
    alias_obj = SubjectAlias.objects.filter(id=alias_id).first()
    if not alias_obj:
        messages.error(request, "Alias not found.")
        return redirect("/subjects/")
    if not is_superadmin_user(getattr(request, "user", None)) and alias_obj.module_id != module.id:
        messages.error(request, "Alias not allowed for this module.")
        return redirect("/subjects/")
    alias_obj.delete()
    messages.info(request, "Subject alias removed.")
    return redirect("/subjects/")


@login_required
@require_http_methods(["POST"])
def add_subject(request):
    if "mentor" in request.session:
        return redirect("/mentor-dashboard/")
    module = _active_module(request)
    name = (request.POST.get("name") or "").strip()
    short_name = (request.POST.get("short_name") or "").strip()
    result_format = (request.POST.get("result_format") or Subject.FORMAT_FULL).strip()
    has_theory = bool(request.POST.get("has_theory"))
    has_practical = bool(request.POST.get("has_practical"))
    if not name:
        messages.error(request, "Subject name is required.")
        return redirect("/subjects/")
    if not short_name:
        short_name = name
    if not has_theory:
        # PR-only subject does not use theory cycle format selection.
        result_format = Subject.FORMAT_FULL
    if result_format not in {Subject.FORMAT_FULL, Subject.FORMAT_T4_ONLY}:
        result_format = Subject.FORMAT_FULL
    subject = Subject.objects.filter(module=module, name__iexact=name).first()
    created = False
    if not subject:
        subject = Subject.objects.create(
            module=module,
            name=name,
            is_active=True,
            result_format=result_format,
            short_name=short_name,
            display_order=(Subject.objects.filter(module=module).aggregate(mx=Max("display_order")).get("mx") or 0) + 1,
            has_theory=has_theory,
            has_practical=has_practical,
        )
        created = True
    if not created:
        subject.name = name
        subject.short_name = short_name
        subject.result_format = result_format
        subject.has_theory = has_theory
        subject.has_practical = has_practical
        subject.is_active = True
        subject.save(update_fields=["name", "short_name", "result_format", "has_theory", "has_practical", "is_active"])
    if is_superadmin_user(getattr(request, "user", None)):
        template, _ = SubjectTemplate.objects.update_or_create(
            name=name,
            defaults={
                "short_name": short_name,
                "has_theory": has_theory,
                "has_practical": has_practical,
                "result_format": result_format,
                "is_active": True,
            },
        )
        if subject.source_template_id != template.id:
            subject.source_template = template
            subject.save(update_fields=["source_template"])
    messages.success(request, "Subject saved.")
    return redirect("/subjects/")


@login_required
@require_http_methods(["POST"])
def edit_subject(request, subject_id):
    if "mentor" in request.session:
        return redirect("/mentor-dashboard/")
    module = _active_module(request)
    name = (request.POST.get("name") or "").strip()
    short_name = (request.POST.get("short_name") or "").strip()
    result_format = (request.POST.get("result_format") or Subject.FORMAT_FULL).strip()
    has_theory = bool(request.POST.get("has_theory"))
    has_practical = bool(request.POST.get("has_practical"))
    if not name:
        messages.error(request, "Subject name is required.")
        return redirect("/subjects/")
    subject = Subject.objects.filter(id=subject_id, module=module).first()
    if subject:
        duplicate = Subject.objects.filter(module=module, name__iexact=name).exclude(id=subject.id).first()
        if duplicate:
            messages.error(request, "A subject with this name already exists in the module.")
            return redirect("/subjects/")
        subject.name = name
        subject.short_name = short_name or name
        if not has_theory:
            result_format = Subject.FORMAT_FULL
        if result_format not in {Subject.FORMAT_FULL, Subject.FORMAT_T4_ONLY}:
            result_format = Subject.FORMAT_FULL
        subject.result_format = result_format
        subject.has_theory = has_theory
        subject.has_practical = has_practical
        subject.save(update_fields=["name", "short_name", "result_format", "has_theory", "has_practical"])
        if is_superadmin_user(getattr(request, "user", None)):
            template = subject.source_template
            if not template:
                template, _ = SubjectTemplate.objects.get_or_create(name=name)
            template.name = name
            template.short_name = subject.short_name
            template.has_theory = has_theory
            template.has_practical = has_practical
            template.result_format = result_format
            template.is_active = True
            template.save()
            if subject.source_template_id != template.id:
                subject.source_template = template
                subject.save(update_fields=["source_template"])
        messages.success(request, "Subject updated.")
    return redirect("/subjects/")


@login_required
@require_http_methods(["POST"])
def apply_subject_templates(request):
    if "mentor" in request.session:
        return redirect("/mentor-dashboard/")
    module = _active_module(request)
    selected_ids = {
        int(x)
        for x in request.POST.getlist("template_ids")
        if str(x).isdigit()
    }
    templates = list(SubjectTemplate.objects.filter(is_active=True).order_by("name"))
    template_map = {t.id: t for t in templates}

    for template in templates:
        existing = Subject.objects.filter(module=module, source_template=template).first()
        if template.id in selected_ids:
            if existing:
                changed_fields = []
                if existing.is_active is False:
                    existing.is_active = True
                    changed_fields.append("is_active")
                if changed_fields:
                    existing.save(update_fields=changed_fields)
            else:
                Subject.objects.create(
                    module=module,
                    source_template=template,
                    name=template.name,
                    short_name=template.short_name or template.name,
                    has_theory=template.has_theory,
                    has_practical=template.has_practical,
                    result_format=template.result_format,
                    is_active=True,
                    display_order=(Subject.objects.filter(module=module).aggregate(mx=Max("display_order")).get("mx") or 0) + 1,
                )
        else:
            if existing and existing.is_active:
                existing.is_active = False
                existing.save(update_fields=["is_active"])

    for template_id in selected_ids:
        template = template_map.get(template_id)
        if not template:
            continue
        existing_by_name = Subject.objects.filter(module=module, name=template.name).first()
        if existing_by_name and not existing_by_name.source_template_id:
            existing_by_name.source_template = template
            existing_by_name.save(update_fields=["source_template"])

    messages.success(request, "Subject selection updated for this module.")
    return redirect("/subjects/")


@login_required
@require_http_methods(["POST"])
def delete_subject(request, subject_id):
    if "mentor" in request.session:
        return redirect("/mentor-dashboard/")
    module = _active_module(request)
    subject = Subject.objects.filter(id=subject_id, module=module).first()
    if subject:
        subject.is_active = False
        subject.save(update_fields=["is_active"])
        messages.success(request, "Subject archived.")
    return redirect("/subjects/")

def next_dir(current_sort, current_dir, column):
    if current_sort == column and current_dir == "asc":
        return "desc"
    return "asc"


def view_attendance(request):

    # mentors should not access coordinator view
    if "mentor" in request.session:
        return redirect("/mentor-dashboard/")
    module = _active_module(request)

    # get available weeks
    weeks = Attendance.objects.filter(student__module=module).values_list("week_no", flat=True)\
                              .distinct().order_by("week_no")

    selected_week = request.GET.get("week")
    # If no week selected → auto open latest week
    if not selected_week:
        latest = Attendance.objects.filter(student__module=module).order_by("-week_no").first()
        if latest:
            selected_week = latest.week_no
    filter_type = request.GET.get("filter", "all")
    mentor_filter = request.GET.get("mentor")
    sort = request.GET.get("sort", "roll")
    direction = request.GET.get("dir", "asc")

    records = None
    mentor_counts = []
    total_count = 0

    # load data only when week selected
    if selected_week:
        selected_week = int(selected_week)

        qs = Attendance.objects.filter(week_no=selected_week, student__module=module)\
            .select_related("student", "student__mentor")

        # ---------- FILTERS ----------
        if filter_type == "weekly":
            qs = qs.filter(week_percentage__lt=80)

        elif filter_type == "overall":
            qs = qs.filter(overall_percentage__lt=80)

        elif filter_type == "either":
            qs = qs.filter(call_required=True)
        
        if mentor_filter:
            qs = qs.filter(student__mentor__name=mentor_filter)

        # ---------- SORTING ----------
        sort_map = {
            "roll": "student__roll_no",
            "enroll": "student__enrollment",
            "name": "student__name",
            "mentor": "student__mentor__name",
            "week": "week_percentage",
            "overall": "overall_percentage",
        }

        order = sort_map.get(sort, "student__roll_no")
        if direction == "desc":
            order = "-" + order

        records = qs.order_by(order)

        # ---------- COUNTS ----------
        mentor_counts = (
            records.values("student__mentor__name")
            .annotate(c=Count("id"))
            .order_by("student__mentor__name")
        )

        total_count = records.count()

    # ---------- ALWAYS RETURN ----------
    return render(request, "view_attendance.html", {
        "weeks": weeks,
        "records": records,
        "selected_week": selected_week,
        "filter": filter_type,
        "sort": sort,
        "dir": direction,
        "mentor_filter": mentor_filter,
        
        # sorting toggle directions
        "dir_roll": next_dir(sort, direction, "roll"),
        "dir_enroll": next_dir(sort, direction, "enroll"),
        "dir_name": next_dir(sort, direction, "name"),
        "dir_mentor": next_dir(sort, direction, "mentor"),
        "dir_week": next_dir(sort, direction, "week"),
        "dir_overall": next_dir(sort, direction, "overall"),

        # counts
        "mentor_counts": mentor_counts,
        "total_count": total_count,
    })



# ---------------- DELETE WEEK ----------------
def delete_week(request):
    module = _active_module(request)

    weeks = Attendance.objects.filter(student__module=module).values_list("week_no", flat=True)\
                              .distinct().order_by("week_no")

    message = ""

    # DELETE SINGLE WEEK
    if request.method == "POST" and "delete_week" in request.POST:
        week_no = int(request.POST.get("week"))

        Attendance.objects.filter(week_no=week_no, student__module=module).delete()
        CallRecord.objects.filter(week_no=week_no, student__module=module).delete()

        message = f"Week-{week_no} deleted successfully"

    # DELETE ALL (password protected)
    if request.method == "POST" and "delete_all" in request.POST:

        password = request.POST.get("password")
        user = authenticate(username=request.user.username, password=password)

        if user:
            Attendance.objects.filter(student__module=module).delete()
            CallRecord.objects.filter(student__module=module).delete()
            message = "ALL WEEKS DELETED"
        else:
            message = "Wrong password"

    return render(request, "delete_week.html", {
        "weeks": weeks,
        "message": message
    })


@login_required
def delete_results(request):
    if "mentor" in request.session:
        return redirect("/")
    module = _active_module(request)

    uploads = ResultUpload.objects.filter(module=module).select_related("subject").order_by("-uploaded_at")
    message = ""

    if request.method == "POST" and "delete_upload" in request.POST:
        upload_id = request.POST.get("upload_id")
        upload = ResultUpload.objects.filter(id=upload_id, module=module).select_related("subject").first()
        if upload:
            label = f"{upload.test_name} - {upload.subject.name}"
            upload.delete()
            message = f"Deleted result upload: {label}"
        else:
            message = "Upload not found."

    if request.method == "POST" and "delete_all" in request.POST:
        password = request.POST.get("password")
        user = authenticate(username=request.user.username, password=password)
        if user:
            ResultUpload.objects.filter(module=module).delete()
            message = "ALL RESULT UPLOADS DELETED"
        else:
            message = "Wrong password"

    uploads = ResultUpload.objects.filter(module=module).select_related("subject").order_by("-uploaded_at")
    return render(
        request,
        "delete_results.html",
        {
            "uploads": uploads,
            "message": message,
        },
    )


# ---------------- LOCK WEEK ----------------
def lock_week(request):
    module = _active_module(request)
    if request.method == "POST":
        week = int(request.POST.get("week"))
        WeekLock.objects.update_or_create(
            module=module,
            week_no=week,
            defaults={"locked": True}
        )
        return redirect(f"/reports/?week={week}")
    return redirect("/reports/")


# ---------------- MENTOR DASHBOARD ----------------
def mentor_dashboard(request):
    mentor = _session_mentor_obj(request)
    if not mentor:
        return redirect("/")
    module = _active_module(request)

    # all uploaded weeks
    weeks = sorted(
        Attendance.objects.filter(student__module=module).values_list("week_no", flat=True).distinct()
    )

    # selected week
    selected_week = request.GET.get("week")

    if not selected_week and weeks:
        selected_week = weeks[-1]
    else:
        selected_week = int(selected_week) if selected_week else None

    records = []

    # build attendance map
    attendance_map = {}
    if selected_week:
        atts = Attendance.objects.filter(week_no=selected_week, student__mentor=mentor, student__module=module)
        for a in atts:
            attendance_map[a.student_id] = a
    
    all_done = False
    not_connected = []

    if selected_week:
        latest_by_student = _latest_attendance_calls_map(module, selected_week, mentor=mentor)
        status_weight = {None: 0, "": 0, "not_received": 1, "received": 2}
        records = sorted(
            list(latest_by_student.values()),
            key=lambda c: (status_weight.get(c.final_status, 0), c.student.roll_no or 999999),
        )
        total = len(records)
        finished = len([c for c in records if c.final_status in {"received", "not_received"}])

        if total > 0 and total == finished:
            all_done = True
            not_connected = [c for c in records if c.final_status == "not_received"]

    
    return render(request,"mentor_dashboard.html",{
        "mentor": mentor,
        "weeks": weeks,
        "selected_week": selected_week,
        "records": records,
        "attendance_map": attendance_map,
        "all_done": all_done,
        "not_connected": not_connected
    })
def mentor_other_calls(request):
    mentor = _session_mentor_obj(request)
    if not mentor:
        return redirect("/")
    module = _active_module(request)
    students = Student.objects.filter(module=module, mentor=mentor).order_by("roll_no", "name")

    qs = (
        OtherCallRecord.objects.filter(mentor=mentor, student__module=module, student__in=students)
        .select_related("student")
        .order_by("student_id", "-created_at", "-id")
    )
    latest_by_student = {}
    for rec in qs:
        if rec.student_id not in latest_by_student:
            latest_by_student[rec.student_id] = rec

    status_weight = {None: 0, "": 0, "not_received": 1, "received": 2}
    records = []
    for s in students:
        latest = latest_by_student.get(s.id)
        records.append(
            {
                "id": s.id,  # button uses student id; each save creates a new immutable call record
                "student": s,
                "final_status": (latest.final_status if latest else None),
            }
        )

    records = sorted(records, key=lambda x: (status_weight.get(x.get("final_status"), 0), x["student"].roll_no or 999999))
    return render(
        request,
        "mentor_other_calls.html",
        {
            "mentor": mentor,
            "records": records,
        },
    )


def save_other_call(request):
    if request.method != "POST":
        return JsonResponse({"ok": False})

    mentor = _session_mentor_obj(request)
    if not mentor:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)
    module = _active_module(request)

    raw_id = request.POST.get("id")
    student = Student.objects.select_related("mentor").filter(
        id=raw_id,
        module=module,
        mentor=mentor,
    ).first()
    if not student:
        # Backward compatibility: old UI may still send call record id.
        previous = (
            OtherCallRecord.objects.select_related("student", "mentor")
            .filter(id=raw_id, mentor=mentor, student__module=module)
            .first()
        )
        student = previous.student if previous else None
    if not student:
        return JsonResponse({"ok": False, "msg": "Student not found"}, status=404)

    status = request.POST.get("status")
    talked = request.POST.get("talked")
    duration = request.POST.get("duration")
    remark = request.POST.get("remark")
    call_reason = request.POST.get("call_reason")
    target = request.POST.get("target")
    call_category = (request.POST.get("call_category") or "other").strip().lower()
    week_no_raw = (request.POST.get("week_no") or "").strip()
    day_no_raw = (request.POST.get("day_no") or "").strip()
    exam_name = (request.POST.get("exam_name") or "").strip()
    subject_name = (request.POST.get("subject_name") or "").strip()
    marks_obtained_raw = (request.POST.get("marks_obtained") or "").strip()
    marks_out_of_raw = (request.POST.get("marks_out_of") or "").strip()

    now_ts = timezone.now()
    call = OtherCallRecord(
        student=student,
        mentor=mentor,
        attempt1_time=now_ts,
    )

    if target in {"student", "father"}:
        call.last_called_target = target
    if call_category not in {"less_attendance", "poor_result", "other", "mentor_intro"}:
        call_category = "other"
    call.call_category = call_category

    if call_category == "less_attendance":
        try:
            week_no = int(week_no_raw)
            day_no = int(day_no_raw)
        except (TypeError, ValueError):
            return JsonResponse({"ok": False, "msg": "Week number and day number are required."}, status=400)

        # Store as a new attendance call record (append-only; never overwrite old call history).
        attendance_call = CallRecord(
            student=student,
            week_no=week_no,
            attempt1_time=timezone.now(),
        )

        parent_text = (remark or "Sick").strip()
        faculty_text = (call_reason or f"Absent on WK-{week_no} DAY-{day_no}").strip()
        attendance_call.talked_with = talked or "father"
        attendance_call.duration = duration or ""
        attendance_call.parent_reason = f"PARENT::{parent_text}||FACULTY::{faculty_text}"
        if status in {"received", "not_received"}:
            attendance_call.final_status = status
        attendance_call.save()
        call.exam_name = ""
        call.subject_name = ""
        call.marks_obtained = None
        call.marks_out_of = None

    if call_category == "poor_result":
        if not exam_name or not subject_name:
            return JsonResponse({"ok": False, "msg": "Exam name and subject name are required."}, status=400)
        call.exam_name = exam_name
        call.subject_name = subject_name
        try:
            call.marks_obtained = float(marks_obtained_raw) if marks_obtained_raw else None
            call.marks_out_of = float(marks_out_of_raw) if marks_out_of_raw else None
        except ValueError:
            return JsonResponse({"ok": False, "msg": "Marks must be numeric."}, status=400)
    elif call_category == "other":
        call.exam_name = ""
        call.subject_name = ""
        call.marks_obtained = None
        call.marks_out_of = None

    if status == "received":
        call.final_status = "received"
        call.talked_with = talked
        call.duration = duration
        if call_category == "poor_result":
            call.parent_remark = (remark or "Student will Study more").strip()
        else:
            call.parent_remark = remark or ""
        call.call_done_reason = call_reason or ""
    elif status == "not_received":
        call.final_status = "not_received"
        call.call_done_reason = call_reason or call.call_done_reason

    call.save()
    return JsonResponse({"ok": True})


# ---------------- SAVE CALL ----------------
def save_call(request):

    if request.method == "POST":
        module = _active_module(request)
        raw_id = request.POST.get("id")
        week_no = request.GET.get("week") or request.POST.get("week") or request.session.get("selected_week")
        try:
            week_no = int(week_no)
        except Exception:
            week_no = None

        student = Student.objects.filter(id=raw_id, module=module).first()
        if not student:
            # Backward compatibility: old payload may still send call record id.
            prev_call = CallRecord.objects.select_related("student").filter(id=raw_id, student__module=module).first()
            student = prev_call.student if prev_call else None
            if not week_no and prev_call:
                week_no = prev_call.week_no
        if not student or not week_no:
            return JsonResponse({"ok": False, "msg": "Invalid student/week"}, status=400)

        status = request.POST.get("status")
        talked = request.POST.get("talked")
        duration = request.POST.get("duration")
        reason = request.POST.get("reason")

        call = CallRecord(
            student=student,
            week_no=week_no,
            attempt1_time=timezone.now(),
        )

        if status == "received":
            call.final_status = "received"
            call.talked_with = talked
            call.duration = duration
            call.parent_reason = reason
        elif status == "not_received":
            call.final_status = "not_received"

        call.save()
        return JsonResponse({"ok": True})


# ---------------- MESSAGE SENT ----------------
def mark_message(request):
    if request.method=="POST":
        module = _active_module(request)
        raw_id = request.POST.get("id")
        week_no = request.GET.get("week") or request.POST.get("week") or request.session.get("selected_week")
        try:
            week_no = int(week_no)
        except Exception:
            week_no = None

        student = Student.objects.filter(id=raw_id, module=module).first()
        source_call = None
        if not student:
            source_call = CallRecord.objects.select_related("student").filter(id=raw_id, student__module=module).first()
            student = source_call.student if source_call else None
            if not week_no and source_call:
                week_no = source_call.week_no
        if not student or not week_no:
            return JsonResponse({"ok": False, "msg": "Invalid student/week"}, status=400)

        if not source_call:
            source_call = (
                CallRecord.objects.filter(student=student, week_no=week_no)
                .order_by("-created_at", "-id")
                .first()
            )
        call = CallRecord(
            student=student,
            week_no=week_no,
            attempt1_time=timezone.now(),
            final_status=(source_call.final_status if source_call else None),
            talked_with=(source_call.talked_with if source_call else None),
            duration=(source_call.duration if source_call else ""),
            parent_reason=(source_call.parent_reason if source_call else ""),
            message_sent=True,
        )
        call.save()
        return JsonResponse({"ok":True})


# ---------------- MENTOR REPORT ----------------
def mentor_report(request):
    mentor_obj = _session_mentor_obj(request)
    if not mentor_obj:
        return redirect("/")
    module = _active_module(request)

    week = request.GET.get("week")
    if not week:
        return render(request,"mentor_report.html")

    week = int(week)

    students = Student.objects.filter(module=module, mentor=mentor_obj).count()

    below80 = Attendance.objects.filter(
        week_no=week, student__mentor=mentor_obj, student__module=module, call_required=True
    ).count()

    latest_calls = _latest_attendance_calls_map(module, week, mentor=mentor_obj)
    calls_done = len([c for c in latest_calls.values() if c.final_status in {"received", "not_received"}])
    received = len([c for c in latest_calls.values() if c.final_status == "received"])
    not_received = len([c for c in latest_calls.values() if c.final_status == "not_received"])
    message_done = len([c for c in latest_calls.values() if c.message_sent])

    not_done = below80 - calls_done

    report = f"""
Follow up Attendance < 80% (Week-{week} only & Overall Week-01 to {week}):

Mentor Name: {mentor_obj.name}
Total no. Of students under mentorship: {students}
No. Of students under mentorship whose attendance < 80%: {below80}
No. Of call done: {calls_done}
No. Of call received: {received}
No. Of call not received: {not_received}
No. Of message done when call not received: {message_done}
Call not done: {not_done}
"""

    return render(request,"mentor_report.html",{"report":report,"week":week})


def mentor_result_calls(request):
    mentor = _session_mentor_obj(request)
    if not mentor:
        return redirect("/")
    module = _active_module(request)
    uploads = list(
        ResultUpload.objects.filter(module=module, calls__student__mentor=mentor, calls__student__module=module)
        .distinct()
        .order_by("-uploaded_at")
    )

    selected_upload = None
    upload_id = request.GET.get("upload")
    if upload_id:
        selected_upload = ResultUpload.objects.filter(id=upload_id, module=module).first()
    if not selected_upload and uploads:
        selected_upload = uploads[0]

    records = []
    all_done = False
    not_connected = []
    if selected_upload:
        fail_student_ids = _upload_fail_student_ids(selected_upload)
        latest_map = _latest_result_calls_map(
            selected_upload,
            mentor=mentor,
            module=module,
            student_ids=fail_student_ids,
        )
        records = sorted(
            list(latest_map.values()),
            key=lambda x: (x.student.roll_no or 999999, x.student.name or ""),
        )
        total = len(records)
        finished = len([c for c in records if c.final_status in {"received", "not_received"}])
        if total > 0 and total == finished:
            all_done = True
            not_connected = [c for c in records if c.final_status == "not_received"]

    return render(
        request,
        "mentor_result_calls.html",
        {
            "mentor": mentor,
            "uploads": uploads,
            "selected_upload": selected_upload,
            "records": records,
            "all_done": all_done,
            "not_connected": not_connected,
        },
    )


def save_result_call(request):
    if request.method != "POST":
        return JsonResponse({"ok": False})

    mentor = _session_mentor_obj(request)
    if not mentor:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)
    module = _active_module(request)

    raw_id = request.POST.get("id")
    upload_id = request.POST.get("upload_id")
    upload = ResultUpload.objects.filter(id=upload_id, module=module).first() if upload_id else None

    student = Student.objects.select_related("mentor").filter(
        id=raw_id,
        mentor=mentor,
        module=module,
    ).first()
    source_call = None
    if not student:
        source_call = ResultCallRecord.objects.select_related("student", "upload").filter(
            id=raw_id,
            student__mentor=mentor,
            student__module=module,
        ).first()
        student = source_call.student if source_call else None
        if not upload and source_call:
            upload = source_call.upload
    if not student or not upload:
        return JsonResponse({"ok": False, "msg": "Student/upload not found"}, status=404)

    if not source_call:
        source_call = (
            ResultCallRecord.objects.filter(upload=upload, student=student)
            .order_by("-created_at", "-id")
            .first()
        )
    if not source_call:
        return JsonResponse({"ok": False, "msg": "No result call context found"}, status=404)

    status = request.POST.get("status")
    talked = request.POST.get("talked")
    duration = request.POST.get("duration")
    reason = request.POST.get("reason")

    call = ResultCallRecord(
        upload=upload,
        student=student,
        attempt1_time=timezone.now(),
        fail_reason=(source_call.fail_reason if source_call else ""),
        marks_current=(source_call.marks_current if source_call else 0),
        marks_total=(source_call.marks_total if source_call else None),
        message_sent=(source_call.message_sent if source_call else False),
    )

    if status == "received":
        call.final_status = "received"
        call.talked_with = talked
        call.duration = duration
        call.parent_reason = reason
    elif status == "not_received":
        call.final_status = "not_received"

    call.save()
    return JsonResponse({"ok": True})


def mark_result_message(request):
    if request.method != "POST":
        return JsonResponse({"ok": False})

    mentor = _session_mentor_obj(request)
    if not mentor:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)
    module = _active_module(request)

    raw_id = request.POST.get("id")
    upload_id = request.POST.get("upload_id")
    upload = ResultUpload.objects.filter(id=upload_id, module=module).first() if upload_id else None

    student = Student.objects.select_related("mentor").filter(
        id=raw_id,
        mentor=mentor,
        module=module,
    ).first()
    source_call = None
    if not student:
        source_call = ResultCallRecord.objects.select_related("student", "upload").filter(
            id=raw_id,
            student__mentor=mentor,
            student__module=module,
        ).first()
        student = source_call.student if source_call else None
        if not upload and source_call:
            upload = source_call.upload
    if not student or not upload:
        return JsonResponse({"ok": False, "msg": "Student/upload not found"}, status=404)

    if not source_call:
        source_call = (
            ResultCallRecord.objects.filter(upload=upload, student=student)
            .order_by("-created_at", "-id")
            .first()
        )
    if not source_call:
        return JsonResponse({"ok": False, "msg": "No result call context found"}, status=404)
    call = ResultCallRecord(
        upload=upload,
        student=student,
        attempt1_time=timezone.now(),
        final_status=(source_call.final_status if source_call else None),
        talked_with=(source_call.talked_with if source_call else None),
        duration=(source_call.duration if source_call else ""),
        parent_reason=(source_call.parent_reason if source_call else ""),
        message_sent=True,
        fail_reason=(source_call.fail_reason if source_call else ""),
        marks_current=(source_call.marks_current if source_call else 0),
        marks_total=(source_call.marks_total if source_call else None),
    )
    call.save()
    return JsonResponse({"ok": True})


def mentor_result_report(request):
    mentor = _session_mentor_obj(request)
    if not mentor:
        return redirect("/")
    module = _active_module(request)

    uploads = list(
        ResultUpload.objects.filter(module=module, calls__student__mentor=mentor, calls__student__module=module)
        .distinct()
        .order_by("-uploaded_at")
    )

    selected_upload = None
    upload_id = request.GET.get("upload")
    if upload_id:
        selected_upload = ResultUpload.objects.filter(id=upload_id, module=module).first()
    if not selected_upload and uploads:
        selected_upload = uploads[0]

    report = ""
    if selected_upload:
        fail_student_ids = _upload_fail_student_ids(selected_upload)
        latest_calls = _latest_result_calls_map(
            selected_upload,
            mentor=mentor,
            module=module,
            student_ids=fail_student_ids,
        )
        total = len(latest_calls)
        received = len([c for c in latest_calls.values() if c.final_status == "received"])
        not_received = len([c for c in latest_calls.values() if c.final_status == "not_received"])
        message_done = len([c for c in latest_calls.values() if c.message_sent])
        report = _result_report_text(
            selected_upload.test_name,
            selected_upload.subject.name,
            mentor.name,
            total,
            received,
            not_received,
            message_done,
        )

    return render(
        request,
        "mentor_result_report.html",
        {
            "uploads": uploads,
            "selected_upload": selected_upload,
            "report": report,
        },
    )


# ---------------- PDF PRINT ----------------
def print_student(request, enrollment):
    if not request.user.is_authenticated and "mentor" not in request.session:
        return redirect("/")

    module = _active_module(request)
    student = Student.objects.select_related("mentor").get(module=module, enrollment=enrollment)
    mentor = _session_mentor_obj(request)
    if mentor and student.mentor_id != mentor.id:
        return HttpResponse("Unauthorized", status=403)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{student.name}.pdf"'

    generate_student_pdf(response, student)
    return response


def _safe_pdf_name(text):
    cleaned = re.sub(r"[^A-Za-z0-9 _-]+", "", str(text or "")).strip()
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned or "Student"


def mentor_prefilled_sif_pdf(request, enrollment):
    mentor = _session_mentor_obj(request)
    if not mentor:
        return redirect("/")
    module = _active_module(request)
    student = Student.objects.select_related("mentor").filter(module=module, enrollment=enrollment, mentor=mentor).first()
    if not student:
        return HttpResponse("Unauthorized", status=403)

    roll = student.roll_no if student.roll_no is not None else "NA"
    filename = f"{roll}_{_safe_pdf_name(student.name)}_SIF.pdf"
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    generate_student_prefilled_pdf(response, student)
    return response


def mentor_prefilled_sif_zip(request):
    mentor = _session_mentor_obj(request)
    if not mentor:
        return redirect("/")
    module = _active_module(request)

    students = list(
        Student.objects.select_related("mentor")
        .filter(module=module, mentor=mentor)
        .order_by("roll_no", "name")
    )

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for s in students:
            pdf_bytes = io.BytesIO()
            generate_student_prefilled_pdf(pdf_bytes, s)
            roll = s.roll_no if s.roll_no is not None else "NA"
            pdf_name = f"{roll}_{_safe_pdf_name(s.name)}_SIF.pdf"
            zf.writestr(pdf_name, pdf_bytes.getvalue())

    zip_name = f"{_safe_pdf_name(mentor.name)}_Prefilled_SIF.zip"
    response = HttpResponse(buffer.getvalue(), content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="{zip_name}"'
    return response


# ---------------- COORDINATOR DASHBOARD ----------------
def coordinator_dashboard(request):

    if "mentor" in request.session:
        return redirect("/mentor-dashboard/")
    module = _active_module(request)

    week = request.GET.get("week")
    if not week:
        return render(request,"coordinator_dashboard.html")

    week = int(week)
    mentors = Mentor.objects.filter(student__module=module).distinct()
    data = []

    for m in mentors:

        total_students = Student.objects.filter(module=module, mentor=m).count()

        need_call = Attendance.objects.filter(
            week_no=week, student__mentor=m, student__module=module, call_required=True
        ).count()

        latest_calls = _latest_attendance_calls_map(module, week, mentor=m)
        received = len([c for c in latest_calls.values() if c.final_status == "received"])
        not_received = len([c for c in latest_calls.values() if c.final_status == "not_received"])

        done = received + not_received
        not_done = max(need_call - done, 0)

        message_sent = len([c for c in latest_calls.values() if c.message_sent])

        percent = round((done/need_call)*100,1) if need_call else 0

        data.append({
            "mentor":m.name,
            "students":total_students,
            "need_call":need_call,
            "done":done,
            "received":received,
            "not_received":not_received,
            "not_done":not_done,
            "msg_sent":message_sent,
            "percent":percent
        })

    return render(request,"coordinator_dashboard.html",{"data":data,"week":week})


def _build_live_followup_rows(module, selected_mentor="", selected_type="all", selected_week=None, selected_exam="all"):
    mentor_names = list(
        Mentor.objects.filter(student__module=module)
        .distinct()
        .order_by("name")
        .values_list("name", flat=True)
    )

    rows = []

    def _fmt_num(v):
        if v is None:
            return "-"
        try:
            f = float(v)
            return f"{f:.2f}".rstrip("0").rstrip(".")
        except Exception:
            return str(v)

    def _status_text(status):
        if status == "received":
            return "Received"
        if status == "not_received":
            return "Not Received"
        return "Pending"

    def _result_denom(test_name):
        test = (test_name or "").upper()
        if test == "T1":
            return ("T1/T1", 25, 25)
        if test == "T2":
            return ("T2/T1+T2", 25, 50)
        if test == "T3":
            return ("T3/T1+T2+T3", 25, 75)
        if test == "T4":
            return ("T4/T1+T2+T3+T4", 50, 100)
        return ("REMEDIAL/REMEDIAL", 100, 100)

    attendance_weeks = list(
        Attendance.objects.filter(student__module=module)
        .order_by("week_no")
        .values_list("week_no", flat=True)
        .distinct()
    )

    attendance_calls = (
        CallRecord.objects.select_related("student", "student__mentor")
        .filter(student__module=module)
    )
    if selected_week:
        attendance_calls = attendance_calls.filter(week_no=selected_week)
    attendance_calls = attendance_calls.order_by("student_id", "week_no", "-created_at", "-id")
    latest_attendance_calls = {}
    for c in attendance_calls:
        key = (c.student_id, c.week_no)
        if key not in latest_attendance_calls:
            latest_attendance_calls[key] = c
    attendance_pairs = set(latest_attendance_calls.keys())
    attendance_call_weeks = set(attendance_pairs)
    attendance_map = {
        (a.student_id, a.week_no): a
        for a in Attendance.objects.filter(
            student__module=module,
            student_id__in=[p[0] for p in attendance_pairs] if attendance_pairs else [],
            week_no__in=[p[1] for p in attendance_pairs] if attendance_pairs else [],
        )
    }
    for c in latest_attendance_calls.values():
        status = _status_text(c.final_status)
        dt = c.attempt2_time or c.attempt1_time or c.created_at
        if c.final_status:
            date_text, time_text = _to_ist_datetime_text(dt)
            duration = c.duration or "-"
        else:
            date_text, time_text, duration = "-", "-", "-"
        a = attendance_map.get((c.student_id, c.week_no))
        wk = _fmt_num(getattr(a, "week_percentage", None))
        ov = _fmt_num(getattr(a, "overall_percentage", None))
        rows.append(
            {
                "mentor": c.student.mentor.name if c.student.mentor_id else "",
                "roll_no": c.student.roll_no,
                "enrollment": c.student.enrollment,
                "student_name": c.student.name,
                "followup_type": "less_attendance",
                "followup_label": "Less Attendance",
                "exam_key": "",
                "status": status,
                "date_text": date_text,
                "time_text": time_text,
                "duration": duration,
                "reason": f"Week-{c.week_no} W: {wk}%, O: {ov}",
                "remarks": _format_parent_faculty_remark(c.parent_reason),
                "sort_dt": dt or timezone.now(),
            }
        )

    result_calls = (
        ResultCallRecord.objects.select_related("student", "student__mentor", "upload", "upload__subject")
        .filter(student__module=module)
    )
    active_fail_pairs = set(
        StudentResult.objects.filter(upload__module=module, fail_flag=True).values_list("upload_id", "student_id")
    )
    result_calls = result_calls.order_by("upload_id", "student_id", "-created_at", "-id")
    latest_result_calls = {}
    for c in result_calls:
        key = (c.upload_id, c.student_id)
        if key not in active_fail_pairs:
            continue
        if key not in latest_result_calls:
            latest_result_calls[key] = c
    for c in latest_result_calls.values():
        status = _status_text(c.final_status)
        dt = c.attempt2_time or c.attempt1_time or c.created_at
        if c.final_status:
            date_text, time_text = _to_ist_datetime_text(dt)
            duration = c.duration or "-"
        else:
            date_text, time_text, duration = "-", "-", "-"
        test_name = c.upload.test_name if c.upload_id else "-"
        test_key = (test_name or "").upper()
        subject_name = c.upload.subject.name if c.upload_id and c.upload.subject_id else "-"
        exam_label, cur_total, overall_total = _result_denom(test_name)
        current_marks = _fmt_num(c.marks_current)
        overall_marks = _fmt_num(c.marks_total)
        rows.append(
            {
                "mentor": c.student.mentor.name if c.student.mentor_id else "",
                "roll_no": c.student.roll_no,
                "enrollment": c.student.enrollment,
                "student_name": c.student.name,
                "followup_type": "poor_result",
                "followup_label": "Poor Result",
                "exam_key": test_key,
                "status": status,
                "date_text": date_text,
                "time_text": time_text,
                "duration": duration,
                "reason": f"{exam_label} : {subject_name} : ({current_marks}/{cur_total}, {overall_marks}/{overall_total})",
                "remarks": _format_parent_faculty_remark(c.parent_reason),
                "sort_dt": dt or timezone.now(),
            }
        )

    other_calls = (
        OtherCallRecord.objects.select_related("student", "student__mentor")
        .filter(student__module=module)
    )
    for c in other_calls:
        status = _status_text(c.final_status)
        dt = c.attempt2_time or c.attempt1_time or c.updated_at or c.created_at
        if c.final_status:
            date_text, time_text = _to_ist_datetime_text(dt)
            duration = c.duration or "-"
        else:
            date_text, time_text, duration = "-", "-", "-"
        reason_text = (c.call_done_reason or "").strip()
        remarks_text = (c.parent_remark or "").strip()
        lower_blob = f"{reason_text} {remarks_text}".lower()

        if c.call_category == "less_attendance":
            f_type = "less_attendance"
            f_label = "Less Attendance"
            reason = reason_text or "Less attendance follow-up"
            exam_key = ""
            # De-duplicate with primary attendance call rows:
            # if same student + week already exists in CallRecord, skip OtherCall row.
            wk_match = re.search(r"wk[-\s:]*([0-9]+)", f"{reason_text} {remarks_text}", flags=re.IGNORECASE)
            if wk_match:
                try:
                    wk_no = int(wk_match.group(1))
                except Exception:
                    wk_no = None
                if wk_no is not None and (c.student_id, wk_no) in attendance_call_weeks:
                    continue
        elif c.call_category == "poor_result":
            f_type = "poor_result"
            f_label = "Poor Result"
            if c.exam_name or c.subject_name:
                reason = f"Poor result ({c.exam_name or '-'} - {c.subject_name or '-'})"
            else:
                reason = reason_text or "Poor result follow-up"
            exam_raw = (c.exam_name or "").upper()
            if "T1" in exam_raw and "T2" not in exam_raw and "T3" not in exam_raw and "T4" not in exam_raw:
                exam_key = "T1"
            elif "T2" in exam_raw:
                exam_key = "T2"
            elif "T3" in exam_raw:
                exam_key = "T3"
            elif "T4" in exam_raw or "SEE" in exam_raw or "TOTAL" in exam_raw:
                exam_key = "T4"
            else:
                exam_key = ""
        elif c.call_category == "mentor_intro":
            f_type = "mentor_intro"
            f_label = "Mentor Intro Call"
            reason = reason_text or "Mentor introduction call"
            exam_key = ""
        else:
            if "intro" in lower_blob:
                f_type = "mentor_intro"
                f_label = "Mentor Intro Call"
            else:
                f_type = "other_direct"
                f_label = "Other (Direct)"
            reason = reason_text or "Direct call"
            exam_key = ""

        rows.append(
            {
                "mentor": c.student.mentor.name if c.student.mentor_id else "",
                "roll_no": c.student.roll_no,
                "enrollment": c.student.enrollment,
                "student_name": c.student.name,
                "followup_type": f_type,
                "followup_label": f_label,
                "exam_key": exam_key,
                "status": status,
                "date_text": date_text,
                "time_text": time_text,
                "duration": duration,
                "reason": reason,
                "remarks": _format_parent_faculty_remark(remarks_text),
                "sort_dt": dt or timezone.now(),
            }
        )

    if selected_mentor:
        rows = [r for r in rows if r["mentor"] == selected_mentor]

    if selected_type != "all":
        rows = [r for r in rows if r["followup_type"] == selected_type]

    if selected_exam in {"T1", "T2", "T3", "T4"}:
        rows = [r for r in rows if r.get("exam_key") == selected_exam]

    rows.sort(key=lambda r: (r["sort_dt"], r["mentor"] or "", r["roll_no"] or 999999), reverse=True)
    return mentor_names, attendance_weeks, rows


def live_followup_sheet(request):
    mentor_mode = bool(request.session.get("mentor"))
    if not mentor_mode and not request.user.is_authenticated:
        return redirect("/")

    module = _active_module(request)
    selected_mentor = (request.GET.get("mentor") or "").strip()
    if mentor_mode:
        selected_mentor = (request.session.get("mentor") or "").strip()
    selected_type = (request.GET.get("type") or "all").strip().lower()
    selected_week_raw = (request.GET.get("week") or "").strip()
    selected_exam = (request.GET.get("exam") or "ALL").strip().upper()
    selected_sort = (request.GET.get("sort") or "date").strip().lower()
    selected_dir = (request.GET.get("dir") or "desc").strip().lower()
    page_number = request.GET.get("page", "1")

    selected_week = None
    week_param_present = "week" in request.GET
    if selected_week_raw.isdigit():
        selected_week = int(selected_week_raw)

    if selected_exam not in {"ALL", "T1", "T2", "T3", "T4"}:
        selected_exam = "ALL"
    mentor_names, attendance_weeks, rows = _build_live_followup_rows(
        module,
        selected_mentor,
        selected_type,
        selected_week,
        selected_exam if selected_exam != "ALL" else "all",
    )
    if (not week_param_present) and selected_week is None and attendance_weeks:
        selected_week = max(attendance_weeks)
        mentor_names, attendance_weeks, rows = _build_live_followup_rows(
            module,
            selected_mentor,
            selected_type,
            selected_week,
            selected_exam if selected_exam != "ALL" else "all",
        )
    selected_week_q = str(selected_week) if selected_week is not None else "ALL"

    key_map = {
        "mentor": "mentor",
        "roll": "roll_no",
        "enrollment": "enrollment",
        "name": "student_name",
        "type": "followup_label",
        "status": "status",
        "date": "sort_dt",
        "duration": "duration",
        "reason": "reason",
        "remarks": "remarks",
    }
    if selected_sort not in key_map:
        selected_sort = "date"
    if selected_dir not in {"asc", "desc"}:
        selected_dir = "desc"

    sort_key = key_map[selected_sort]
    reverse = selected_dir == "desc"

    if sort_key == "roll_no":
        rows.sort(key=lambda r: (r.get("roll_no") is None, r.get("roll_no") or 0), reverse=reverse)
    elif sort_key == "sort_dt":
        rows.sort(key=lambda r: r.get("sort_dt") or timezone.now(), reverse=reverse)
    else:
        rows.sort(key=lambda r: str(r.get(sort_key) or "").lower(), reverse=reverse)

    paginator = Paginator(rows, 120)
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "live_followup_sheet.html",
        {
            "rows": page_obj.object_list,
            "page_obj": page_obj,
            "mentor_names": mentor_names,
            "attendance_weeks": attendance_weeks,
            "selected_mentor": selected_mentor,
            "selected_type": selected_type,
            "selected_week": selected_week,
            "selected_week_q": selected_week_q,
            "selected_exam": selected_exam,
            "selected_sort": selected_sort,
            "selected_dir": selected_dir,
            "module": module,
            "mentor_mode": mentor_mode,
            "is_superadmin_view": (not mentor_mode and bool(request.user.is_authenticated and is_superadmin_user(request.user))),
        },
    )


def live_followup_sheet_excel(request):
    mentor_mode = bool(request.session.get("mentor"))
    if not mentor_mode and not request.user.is_authenticated:
        return redirect("/")

    module = _active_module(request)
    selected_mentor = (request.GET.get("mentor") or "").strip()
    if mentor_mode:
        selected_mentor = (request.session.get("mentor") or "").strip()
    selected_type = (request.GET.get("type") or "all").strip().lower()
    selected_week_raw = (request.GET.get("week") or "").strip()
    selected_exam = (request.GET.get("exam") or "ALL").strip().upper()
    selected_week = int(selected_week_raw) if selected_week_raw.isdigit() else None
    if selected_exam not in {"ALL", "T1", "T2", "T3", "T4"}:
        selected_exam = "ALL"
    _, _, rows = _build_live_followup_rows(
        module, selected_mentor, selected_type, selected_week, selected_exam if selected_exam != "ALL" else "all"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Live Followup"

    headers = [
        "Mentor",
        "Roll No.",
        "Enrollment No.",
        "Name of Students",
        "Followup Type",
        "Status",
        "Date of Phone Call",
        "Call Duration (Mins)",
        "Reason of Phone Call",
        "Remarks by Parents",
    ]
    ws.append(headers)

    for r in rows:
        ws.append(
            [
                r["mentor"],
                r["roll_no"],
                r["enrollment"],
                r["student_name"],
                r["followup_label"],
                r["status"],
                f'{r["date_text"]} {r["time_text"]}',
                r["duration"],
                r["reason"],
                r["remarks"],
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="live_followup_sheet.xlsx"'
    wb.save(response)
    return response


def live_followup_sheet_pdf(request):
    mentor_mode = bool(request.session.get("mentor"))
    if not mentor_mode and not request.user.is_authenticated:
        return redirect("/")

    module = _active_module(request)
    selected_mentor = (request.GET.get("mentor") or "").strip()
    if mentor_mode:
        selected_mentor = (request.session.get("mentor") or "").strip()
    selected_type = (request.GET.get("type") or "all").strip().lower()
    selected_week_raw = (request.GET.get("week") or "").strip()
    selected_exam = (request.GET.get("exam") or "ALL").strip().upper()
    selected_week = int(selected_week_raw) if selected_week_raw.isdigit() else None
    if selected_exam not in {"ALL", "T1", "T2", "T3", "T4"}:
        selected_exam = "ALL"
    _, _, rows = _build_live_followup_rows(
        module, selected_mentor, selected_type, selected_week, selected_exam if selected_exam != "ALL" else "all"
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="live_followup_sheet.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=12,
        rightMargin=12,
        topMargin=14,
        bottomMargin=12,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"Live Followup Sheet - {module.name}", styles["Heading3"]),
        Spacer(1, 6),
    ]

    table_data = [
        [
            "Mentor",
            "Roll No.",
            "Enrollment No.",
            "Name of Students",
            "Followup Type",
            "Status",
            "Date of Phone Call",
            "Call Duration (Mins)",
            "Reason of Phone Call",
            "Remarks by Parents",
        ]
    ]

    for r in rows:
        table_data.append(
            [
                r["mentor"],
                str(r["roll_no"] or "-"),
                r["enrollment"] or "-",
                r["student_name"] or "-",
                r["followup_label"] or "-",
                r["status"] or "-",
                f'{r["date_text"]} {r["time_text"]}',
                str(r["duration"] or "-"),
                r["reason"] or "-",
                r["remarks"] or "-",
            ]
        )

    tbl = Table(
        table_data,
        repeatRows=1,
        colWidths=[66, 38, 85, 110, 70, 62, 88, 56, 110, 80],
    )
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ]
        )
    )
    story.append(tbl)
    doc.build(story)
    return response


@login_required
def live_followup_sheet_db_backup_json(request):
    if request.session.get("mentor"):
        return HttpResponse("Forbidden", status=403)
    if not is_superadmin_user(request.user):
        return HttpResponse("Forbidden", status=403)

    buffer = io.StringIO()
    call_command(
        "dumpdata",
        "auth.user",
        "core",
        indent=2,
        stdout=buffer,
    )
    payload = buffer.getvalue()
    ts = timezone.now().astimezone(IST).strftime("%Y%m%d_%H%M%S")
    response = HttpResponse(payload, content_type="application/json")
    response["Content-Disposition"] = f'attachment; filename="easymentor_db_backup_{ts}.json"'
    return response


@login_required
def coordinator_result_report(request):
    if "mentor" in request.session:
        return redirect("/mentor-dashboard/")
    module = _active_module(request)

    uploads = ResultUpload.objects.filter(module=module).order_by("-uploaded_at")
    selected_upload = None
    upload_id = request.GET.get("upload")
    if upload_id:
        selected_upload = ResultUpload.objects.filter(id=upload_id, module=module).first()
    if not selected_upload:
        selected_upload = uploads.first()

    data = []
    if selected_upload:
        fail_student_ids = _upload_fail_student_ids(selected_upload)
        mentors = (
            Mentor.objects.filter(student__module=module, student__resultcallrecord__upload=selected_upload)
            .distinct()
            .order_by("name")
        )
        for m in mentors:
            latest_calls = _latest_result_calls_map(
                selected_upload,
                mentor=m,
                module=module,
                student_ids=fail_student_ids,
            )
            need_call = len(latest_calls)
            received = len([c for c in latest_calls.values() if c.final_status == "received"])
            not_received = len([c for c in latest_calls.values() if c.final_status == "not_received"])
            done = received + not_received
            not_done = max(need_call - done, 0)
            msg_sent = len([c for c in latest_calls.values() if c.message_sent])
            percent = round((done / need_call) * 100, 1) if need_call else 0
            data.append(
                {
                    "mentor": m.name,
                    "need_call": need_call,
                    "done": done,
                    "received": received,
                    "not_received": not_received,
                    "not_done": not_done,
                    "msg_sent": msg_sent,
                    "percent": percent,
                }
            )

    return render(
        request,
        "coordinator_result_report.html",
        {
            "uploads": uploads,
            "selected_upload": selected_upload,
            "data": data,
        },
    )

def update_mobile(request):

    if request.method == "POST":
        if not request.user.is_authenticated and "mentor" not in request.session:
            return JsonResponse({"ok": False, "error": "Unauthorized"}, status=401)

        module = _active_module(request)
        enrollment = request.POST.get("enrollment")
        field = request.POST.get("field")
        value = request.POST.get("value")

        student = Student.objects.get(module=module, enrollment=enrollment)
        mentor = _session_mentor_obj(request)
        is_mentor_update = bool(mentor)
        if is_mentor_update and student.mentor_id != mentor.id:
            return JsonResponse({"ok": False, "error": "Unauthorized"}, status=403)

        if field == "father":
            student.father_mobile = value
            student.father_mobile_updated_by_mentor = is_mentor_update
        elif field == "mother":
            student.mother_mobile = value
        elif field == "student":
            student.student_mobile = value
            student.student_mobile_updated_by_mentor = is_mentor_update

        student.save()

        return JsonResponse({"ok": True})

    
# ---------------- CONTROL PANEL ----------------
def control_panel(request):

    if "mentor" in request.session:
        return redirect("/")

    module = _active_module(request)
    students = Student.objects.select_related("mentor").filter(module=module).order_by("roll_no")

    return render(request,"control_panel.html",{"students":students})


def mentor_print_sif(request):
    mentor = _session_mentor_obj(request)
    if not mentor:
        return redirect("/")
    module = _active_module(request)

    students = Student.objects.select_related("mentor").filter(module=module, mentor=mentor).order_by("roll_no", "name")
    return render(
        request,
        "mentor_print_sif.html",
        {
            "mentor": mentor,
            "students": students,
        },
    )


def mentor_view_sif(request):
    mentor = _session_mentor_obj(request)
    if not mentor:
        return redirect("/")

    module = _active_module(request)
    students = list(
        Student.objects.select_related("mentor")
        .filter(module=module, mentor=mentor)
        .order_by("roll_no", "name")
    )

    selected_enrollment = (request.GET.get("enrollment") or "").strip()
    selected_student = None
    if students:
        if selected_enrollment:
            selected_student = next((s for s in students if s.enrollment == selected_enrollment), None)
        if not selected_student:
            selected_student = students[0]

    sem_value = "1"
    if module and module.semester:
        sem_value = (module.semester.split("-")[-1] or "1").strip()

    attendance_rows = []
    result_rows = []
    other_rows = []

    if selected_student:
        calls_by_week = {}
        for c in CallRecord.objects.filter(student=selected_student).order_by("week_no", "-created_at", "-id"):
            if c.week_no not in calls_by_week:
                calls_by_week[c.week_no] = c
        attendance_qs = Attendance.objects.filter(student=selected_student).order_by("week_no")
        sr = 1
        for att in attendance_qs:
            if (att.week_percentage or 0) >= 80 and (att.overall_percentage or 0) >= 80:
                continue
            call = calls_by_week.get(att.week_no)
            call_done = bool(call and call.final_status in ("received", "not_received"))
            if call_done:
                call_dt = (call.attempt2_time or call.attempt1_time or call.created_at)
                date, time = _to_ist_datetime_text(call_dt)
                duration = (call.duration or "").strip() or "-"
                time_duration = f"{time} ({duration})" if time != "-" else "-"
            else:
                date, time, duration, time_duration = "-", "-", "-", "-"
            discussed = ((call.talked_with or "-").title() if call else "-")
            parent_remark = "-"
            faculty_remark = "-"
            if call and call.parent_reason:
                reason_text = (call.parent_reason or "").strip()
                if "PARENT::" in reason_text and "||FACULTY::" in reason_text:
                    p, f = reason_text.split("||FACULTY::", 1)
                    parent_remark = p.replace("PARENT::", "", 1).strip() or "-"
                    faculty_remark = f.strip() or "-"
                else:
                    parent_remark = reason_text

            attendance_rows.append(
                {
                    "sr": sr,
                    "sem": sem_value,
                    "week_no": att.week_no,
                    "attend_text": f"W:{round(att.week_percentage, 2)} / O:{round(att.overall_percentage, 2)}",
                    "status": _call_status_text(call.final_status if call else None),
                    "date": date,
                    "time_duration": time_duration,
                    "discussed": discussed,
                    "parent_remark": parent_remark,
                    "faculty_remark": faculty_remark,
                }
            )
            sr += 1

        result_rows_qs = StudentResult.objects.filter(student=selected_student).select_related("upload", "upload__subject")
        result_rows_sorted = sorted(
            result_rows_qs,
            key=lambda r: (
                _test_sort_key(r.upload.test_name if r.upload else ""),
                _subject_sort_key(r.upload.subject.name if r.upload and r.upload.subject else ""),
                r.upload.uploaded_at if r.upload else timezone.now(),
                r.id,
            ),
        )
        result_call_map = {}
        for c in ResultCallRecord.objects.filter(student=selected_student).select_related("upload").order_by("upload_id", "-created_at", "-id"):
            if c.upload_id not in result_call_map:
                result_call_map[c.upload_id] = c
        sr = 1
        for row in result_rows_sorted:
            if not row.upload:
                continue
            cur_thr, total_thr = _result_thresholds(row.upload.test_name)
            current_fail = row.marks_current is not None and row.marks_current < cur_thr
            total_fail = row.marks_total is not None and row.marks_total < total_thr
            if not (current_fail or total_fail):
                continue

            call = result_call_map.get(row.upload_id)
            call_done = bool(call and call.final_status in ("received", "not_received"))
            if call_done:
                call_dt = (call.attempt2_time or call.attempt1_time or call.created_at)
                date, time = _to_ist_datetime_text(call_dt)
                duration = (call.duration or "").strip() or "-"
                time_duration = f"{time} ({duration})" if time != "-" else "-"
            else:
                date, time, duration, time_duration = "-", "-", "-", "-"
            discussed = ((call.talked_with or "-").title() if call else "-")
            subject_name = row.upload.subject.name if row.upload.subject else "-"
            obtained = "-" if row.marks_current is None else row.marks_current
            total = "-" if row.marks_total is None else row.marks_total
            result_rows.append(
                {
                    "sr": sr,
                    "sem": sem_value,
                    "status": _call_status_text(call.final_status if call else None),
                    "date": date,
                    "time_duration": time_duration,
                    "discussed": discussed,
                    "exam": _exam_name_for_sif(row.upload.test_name),
                    "subject": f"{subject_name} ({obtained}/{total})",
                    "remark": (call.parent_reason if call and call.parent_reason else "-"),
                }
            )
            sr += 1

        direct_call = (
            OtherCallRecord.objects.filter(student=selected_student)
            .order_by("-created_at", "-id")
            .first()
        )
        if direct_call and direct_call.call_category != "poor_result":
            call_done = direct_call.final_status in ("received", "not_received")
            if call_done:
                call_dt = direct_call.attempt2_time or direct_call.attempt1_time or direct_call.updated_at or direct_call.created_at
                date, time = _to_ist_datetime_text(call_dt)
                duration = (direct_call.duration or "").strip() or "-"
                time_duration = f"{time} ({duration})" if time != "-" else "-"
            else:
                time_duration, date = "-", "-"
            discussed = (direct_call.talked_with or "-").title()
            other_rows.append(
                {
                    "sr": 1,
                    "sem": sem_value,
                    "status": _call_status_text(direct_call.final_status),
                    "date": date,
                    "time_duration": time_duration,
                    "discussed": discussed,
                    "reason": direct_call.call_done_reason or "-",
                    "remark": direct_call.parent_remark or "-",
                }
            )
        else:
            other_rows.append(
                {
                    "sr": 1,
                    "sem": sem_value,
                    "status": "Pending",
                    "date": "-",
                    "time_duration": "-",
                    "discussed": "-",
                    "reason": "-",
                    "remark": "-",
                }
            )

    return render(
        request,
        "mentor_view_sif.html",
        {
            "mentor": mentor,
            "students": students,
            "selected_student": selected_student,
            "attendance_rows": attendance_rows,
            "result_rows": result_rows,
            "other_rows": other_rows,
        },
    )


def mentor_student_data(request):
    mentor = _session_mentor_obj(request)
    if not mentor:
        return redirect("/")

    module = _active_module(request)
    qs = Student.objects.select_related("mentor").filter(module=module)
    batch_choices = sorted({(v or "").strip() for v in qs.values_list("batch", flat=True) if (v or "").strip()})
    division_choices = sorted({(v or "").strip() for v in qs.values_list("division", flat=True) if (v or "").strip()})
    mentor_choices = sorted({(v or "").strip() for v in qs.values_list("mentor__name", flat=True) if (v or "").strip()})

    filters = {
        "roll": (request.GET.get("f_roll") or "").strip(),
        "batch": (request.GET.get("f_batch") or "").strip(),
        "division": (request.GET.get("f_division") or "").strip(),
        "enrollment": (request.GET.get("f_enrollment") or "").strip(),
        "name": (request.GET.get("f_name") or "").strip(),
        "mentor": (request.GET.get("f_mentor") or "").strip(),
        "student_mobile": (request.GET.get("f_student_mobile") or "").strip(),
        "father_mobile": (request.GET.get("f_father_mobile") or "").strip(),
    }

    if filters["roll"]:
        qs = qs.filter(roll_no__icontains=filters["roll"])
    if filters["batch"]:
        qs = qs.filter(batch__icontains=filters["batch"])
    if filters["division"]:
        qs = qs.filter(division__icontains=filters["division"])
    if filters["enrollment"]:
        qs = qs.filter(enrollment__icontains=filters["enrollment"])
    if filters["name"]:
        qs = qs.filter(name__icontains=filters["name"])
    if filters["mentor"]:
        qs = qs.filter(mentor__name__icontains=filters["mentor"])
    if filters["student_mobile"]:
        qs = qs.filter(student_mobile__icontains=filters["student_mobile"])
    if filters["father_mobile"]:
        qs = qs.filter(father_mobile__icontains=filters["father_mobile"])

    sort = (request.GET.get("sort") or "roll").strip()
    direction = (request.GET.get("dir") or "asc").strip().lower()
    sort_map = {
        "roll": "roll_no",
        "batch": "batch",
        "division": "division",
        "enrollment": "enrollment",
        "name": "name",
        "mentor": "mentor__name",
        "student_mobile": "student_mobile",
        "father_mobile": "father_mobile",
    }
    sort_field = sort_map.get(sort, "roll_no")
    if direction == "desc":
        sort_field = f"-{sort_field}"
    students = qs.order_by(sort_field, "name")

    def next_dir(key):
        if sort == key and direction == "asc":
            return "desc"
        return "asc"

    return render(
        request,
        "mentor_student_data.html",
        {
            "students": students,
            "filters": filters,
            "batch_choices": batch_choices,
            "division_choices": division_choices,
            "mentor_choices": mentor_choices,
            "sort": sort,
            "direction": direction,
            "dir_roll": next_dir("roll"),
            "dir_batch": next_dir("batch"),
            "dir_division": next_dir("division"),
            "dir_enrollment": next_dir("enrollment"),
            "dir_name": next_dir("name"),
            "dir_mentor": next_dir("mentor"),
            "dir_student_mobile": next_dir("student_mobile"),
            "dir_father_mobile": next_dir("father_mobile"),
        },
    )


def mentor_whatsapp_panel(request):
    mentor = _session_mentor_obj(request)
    if not mentor:
        return redirect("/")

    module = _active_module(request)
    students = list(
        Student.objects.filter(module=module, mentor=mentor).order_by("roll_no", "name")
    )

    message_text = ""
    target = ""
    recipients = []

    if request.method == "POST":
        message_text = (request.POST.get("message") or "").strip()
        target = (request.POST.get("target") or "").strip()
        if not message_text:
            messages.error(request, "Message is required.")
        elif target not in {"student", "father"}:
            messages.error(request, "Choose a valid target.")
        else:
            for s in students:
                raw_number = s.student_mobile if target == "student" else s.father_mobile
                phone = _normalize_whatsapp_phone(raw_number)
                if not phone:
                    continue
                recipients.append(
                    {
                        "name": s.name,
                        "enrollment": s.enrollment,
                        "phone": phone,
                    }
                )
            if not recipients:
                messages.warning(request, f"No valid {target} numbers found for your mentees.")

    return render(
        request,
        "mentor_whatsapp_panel.html",
        {
            "message_text": message_text,
            "target": target,
            "recipients": recipients,
            "recipient_count": len(recipients),
            "message_encoded": quote(message_text),
        },
    )


def download_whatsapp_extension(request):
    mentor = _session_mentor_obj(request)
    if not mentor and not request.user.is_authenticated:
        return redirect("/")

    ext_dir = os.path.join(settings.BASE_DIR, "whatsapp_automation_extension")
    if not os.path.isdir(ext_dir):
        return HttpResponse("Extension files not found.", status=404)

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for name in ["manifest.json", "whatsapp_content.js", "README.md"]:
            path = os.path.join(ext_dir, name)
            if os.path.exists(path):
                zf.write(path, arcname=name)
        zf.writestr(
            "INSTALL_AND_SUPPORT.txt",
            (
                "EasyMentor WhatsApp Extension\n\n"
                "How to use:\n"
                "1. Go to chrome://extensions\n"
                "2. Enable Developer mode\n"
                "3. Load unpacked -> select folder whatsapp_automation_extension\n"
                "4. Login once to WhatsApp Web\n"
                "5. In portal mentor page, enter message and click Start Auto (Extension)\n\n"
                "Buy me a chai with UPI link:\n"
                "upi://pay?pa=8866749627@upi&pn=Hardik%20Shah&am=15&cu=INR&tn=Cutting%20Chai\n"
            ),
        )

    response = HttpResponse(buffer.getvalue(), content_type="application/zip")
    response["Content-Disposition"] = 'attachment; filename="easymentor_whatsapp_extension.zip"'
    return response


def mentor_sif_marks(request):
    mentor = _session_mentor_obj(request)
    if not mentor:
        return redirect("/")
    module = _active_module(request)
    lock_obj = SifMarksLock.objects.filter(module=module).first()
    if not lock_obj or not lock_obj.locked:
        return render(request, "mentor_sif_marks.html", {"locked": False})

    students = list(
        Student.objects.select_related("mentor")
        .filter(module=module, mentor=mentor)
        .order_by("roll_no", "name")
    )
    selected_enrollment = request.GET.get("enrollment") or (students[0].enrollment if students else "")
    selected_student = Student.objects.filter(module=module, mentor=mentor, enrollment=selected_enrollment).first() if selected_enrollment else None
    marks_rows = _sif_marks_rows_for_student(selected_student, module) if selected_student else []
    return render(
        request,
        "mentor_sif_marks.html",
        {
            "locked": True,
            "students": students,
            "selected_student": selected_student,
            "selected_enrollment": selected_enrollment,
            "marks_rows": marks_rows,
        },
    )


def mentor_sif_marks_pdf(request, enrollment):
    mentor = _session_mentor_obj(request)
    if not mentor:
        return redirect("/")
    module = _active_module(request)
    lock_obj = SifMarksLock.objects.filter(module=module).first()
    if not lock_obj or not lock_obj.locked:
        return HttpResponse("Marks not yet locked.", status=403)
    student = Student.objects.filter(module=module, mentor=mentor, enrollment=enrollment).first()
    if not student:
        return HttpResponse("Unauthorized", status=403)
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{_safe_pdf_name(student.name)}_SIF.pdf"'
    generate_student_pdf(response, student)
    return response


def mentor_sif_marks_pdf_all(request):
    mentor = _session_mentor_obj(request)
    if not mentor:
        return redirect("/")
    module = _active_module(request)
    lock_obj = SifMarksLock.objects.filter(module=module).first()
    if not lock_obj or not lock_obj.locked:
        return HttpResponse("Marks not yet locked.", status=403)

    students = list(
        Student.objects.select_related("mentor")
        .filter(module=module, mentor=mentor)
        .order_by("roll_no", "name")
    )
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for s in students:
            pdf_bytes = io.BytesIO()
            generate_student_pdf(pdf_bytes, s)
            roll = s.roll_no if s.roll_no is not None else "NA"
            zf.writestr(f"{roll}_{_safe_pdf_name(s.name)}_SIF.pdf", pdf_bytes.getvalue())

    response = HttpResponse(buffer.getvalue(), content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="{_safe_pdf_name(mentor.name)}_SIF_Marks.zip"'
    return response


def healthz(request):
    return JsonResponse({"ok": True, "service": "easymentor", "ts": timezone.now().isoformat()})


@login_required
@require_http_methods(["POST"])
def rbac_create_coordinator(request):
    guard = _require_superadmin(request)
    if guard:
        return guard

    username = (request.POST.get("username") or "").strip()
    password = request.POST.get("password") or ""
    module_ids_raw = (request.POST.get("module_ids") or "").strip()

    if not username or not password:
        return JsonResponse({"ok": False, "msg": "username and password required"}, status=400)
    if User.objects.filter(username__iexact=username).exists():
        return JsonResponse({"ok": False, "msg": "username already exists"}, status=400)

    module_ids = [x.strip() for x in module_ids_raw.split(",") if x.strip().isdigit()]
    modules = list(AcademicModule.objects.filter(id__in=module_ids, is_active=True)) if module_ids else []
    if not modules:
        return JsonResponse({"ok": False, "msg": "at least one valid module is required"}, status=400)

    user = User.objects.create_user(username=username, password=password, is_active=True, is_staff=False, is_superuser=False)
    CoordinatorModuleAccess.objects.bulk_create(
        [CoordinatorModuleAccess(coordinator=user, module=m) for m in modules],
        ignore_conflicts=True,
    )
    return JsonResponse({"ok": True, "coordinator_id": user.id, "modules": [m.id for m in modules]})


@login_required
@require_http_methods(["POST"])
def rbac_update_coordinator_modules(request):
    guard = _require_superadmin(request)
    if guard:
        return guard

    coordinator_id = request.POST.get("coordinator_id")
    module_ids_raw = (request.POST.get("module_ids") or "").strip()
    if not str(coordinator_id or "").isdigit():
        return JsonResponse({"ok": False, "msg": "valid coordinator_id required"}, status=400)

    coordinator = User.objects.filter(id=int(coordinator_id)).first()
    if not coordinator:
        return JsonResponse({"ok": False, "msg": "coordinator not found"}, status=404)
    if is_superadmin_user(coordinator):
        return JsonResponse({"ok": False, "msg": "cannot remap superadmin"}, status=400)

    module_ids = [x.strip() for x in module_ids_raw.split(",") if x.strip().isdigit()]
    modules = list(AcademicModule.objects.filter(id__in=module_ids, is_active=True)) if module_ids else []
    if not modules:
        return JsonResponse({"ok": False, "msg": "at least one valid module is required"}, status=400)

    CoordinatorModuleAccess.objects.filter(coordinator=coordinator).delete()
    CoordinatorModuleAccess.objects.bulk_create(
        [CoordinatorModuleAccess(coordinator=coordinator, module=m) for m in modules],
        ignore_conflicts=True,
    )
    return JsonResponse({"ok": True, "coordinator_id": coordinator.id, "modules": [m.id for m in modules]})


@login_required
@require_http_methods(["POST"])
def superadmin_change_password(request):
    guard = _require_superadmin(request)
    if guard:
        return guard

    current_password = request.POST.get("current_password") or ""
    new_password = request.POST.get("new_password") or ""
    if not request.user.check_password(current_password):
        return JsonResponse({"ok": False, "msg": "current password is incorrect"}, status=400)
    if len(new_password) < 8:
        return JsonResponse({"ok": False, "msg": "new password must be at least 8 chars"}, status=400)

    request.user.set_password(new_password)
    request.user.save(update_fields=["password"])
    update_session_auth_hash(request, request.user)
    return JsonResponse({"ok": True})


# ---------------- MODULE SWITCH ----------------
@require_http_methods(["POST"])
def switch_module(request):
    if not request.user.is_authenticated and "mentor" not in request.session:
        return redirect("/")
    module_id = request.POST.get("module_id")
    allowed_qs = allowed_modules_for_user(request)
    module = allowed_qs.filter(id=module_id).first()
    if module:
        request.session["current_module_id"] = module.id
    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER") or "/reports/"
    return redirect(next_url)


@login_required
def manage_modules(request):
    if "mentor" in request.session:
        return redirect("/mentor-dashboard/")
    if not is_superadmin_user(request.user):
        return HttpResponse("Forbidden", status=403)

    if request.method == "POST":
        action = (request.POST.get("action") or "create").strip()
        if action == "delete":
            module_id = request.POST.get("module_id")
            module = AcademicModule.objects.filter(id=module_id).first()
            if not module:
                messages.error(request, "Module not found.")
            else:
                module.delete()
                messages.success(request, "Module deleted.")
            return redirect("/modules/")

        module_id = request.POST.get("module_id")
        batch = (request.POST.get("academic_batch") or "").strip()
        year_level = (request.POST.get("year_level") or "FY").strip()
        variant = (request.POST.get("variant") or "FY2-CE").strip()
        semester = (request.POST.get("semester") or "Sem-1").strip()
        is_active = (request.POST.get("is_active") or "1").strip() == "1"

        if not batch:
            messages.error(request, "Batch is required.")
            return redirect("/modules/")
        if year_level not in {x[0] for x in AcademicModule.YEAR_CHOICES}:
            year_level = "FY"
        if variant not in {x[0] for x in AcademicModule.VARIANT_CHOICES}:
            variant = "FY2-CE"
        if semester not in {x[0] for x in AcademicModule.SEM_CHOICES}:
            semester = "Sem-1"

        name = _module_display_name(variant, semester, batch)

        if action == "update":
            module = AcademicModule.objects.filter(id=module_id).first()
            if not module:
                messages.error(request, "Module not found.")
                return redirect("/modules/")
            conflict = AcademicModule.objects.filter(name=name).exclude(id=module.id).exists()
            if conflict:
                messages.error(request, "Another module already exists with same name.")
                return redirect("/modules/")
            module.name = name
            module.academic_batch = batch
            module.year_level = year_level
            module.variant = variant
            module.semester = semester
            module.is_active = is_active
            module.save(
                update_fields=[
                    "name",
                    "academic_batch",
                    "year_level",
                    "variant",
                    "semester",
                    "is_active",
                ]
            )
            request.session["current_module_id"] = module.id
            messages.success(request, f"Module updated: {module.name}")
            return redirect("/modules/")

        module, created = AcademicModule.objects.get_or_create(
            name=name,
            defaults={
                "academic_batch": batch,
                "year_level": year_level,
                "variant": variant,
                "semester": semester,
                "is_active": is_active,
            },
        )
        if not created and action == "create":
            module.academic_batch = batch
            module.year_level = year_level
            module.variant = variant
            module.semester = semester
            module.is_active = is_active
            module.save(update_fields=["academic_batch", "year_level", "variant", "semester", "is_active"])
        request.session["current_module_id"] = module.id
        if created:
            messages.success(request, f"Module created: {module.name}")
        else:
            messages.info(request, f"Module already exists and updated: {module.name}")
        return redirect("/modules/")

    return render(
        request,
        "modules.html",
        {
            "modules": AcademicModule.objects.all().order_by("-id"),
            "year_choices": AcademicModule.YEAR_CHOICES,
            "variant_choices": AcademicModule.VARIANT_CHOICES,
            "sem_choices": AcademicModule.SEM_CHOICES,
        },
    )


# ---------------- SEM REGISTER ----------------

def _semester_register_fallback_from_daily(module, students):
    calendar = _calendar_for_module(module)
    if not calendar:
        return [], {}

    today = timezone.localdate()
    sessions = list(
        LectureSession.objects.filter(module=module, date__lte=today).order_by("date", "lecture_no", "batch")
    )
    sessions = [session for session in sessions if _attendance_allowed_for_date(module, session.date)]
    if not sessions:
        return [], {}

    student_keys = {student.id: _student_batch_keys(student) for student in students}
    absences = LectureAbsence.objects.filter(session_id__in=[session.id for session in sessions]).values_list("session_id", "student_id")
    absent_map = {}
    for session_id, student_id in absences:
        absent_map.setdefault(session_id, set()).add(student_id)

    per_student_week = {student.id: {} for student in students}
    weeks_seen = set()

    for session in sessions:
        phase, week_no = week_for_date(calendar, session.date)
        if not phase or not week_no:
            continue
        normalized_week = _normalize_week_no(phase, week_no)
        weeks_seen.add(normalized_week)
        batch_key = _norm_batch_key(session.batch)
        if not batch_key:
            continue
        absent_students = absent_map.get(session.id, set())
        for student in students:
            if batch_key not in student_keys.get(student.id, set()):
                continue
            stats = per_student_week[student.id].setdefault(normalized_week, {"held": 0, "attended": 0})
            stats["held"] += 1
            if student.id not in absent_students:
                stats["attended"] += 1

    weeks = sorted(weeks_seen)
    fallback = {}
    for student in students:
        fallback[student.id] = {}
        cumulative_held = 0
        cumulative_attended = 0
        for week in weeks:
            stats = per_student_week.get(student.id, {}).get(week)
            if not stats or not stats["held"]:
                continue
            cumulative_held += stats["held"]
            cumulative_attended += stats["attended"]
            fallback[student.id][week] = {
                "week_percentage": round((stats["attended"] / stats["held"]) * 100, 2) if stats["held"] else None,
                "overall_percentage": round((cumulative_attended / cumulative_held) * 100, 2) if cumulative_held else None,
            }
    return weeks, fallback


def semester_register(request):
    if "mentor" in request.session:
        return redirect("/mentor-semester-register/")

    module = _active_module(request)
    students = list(Student.objects.select_related("mentor").filter(module=module).order_by("roll_no"))
    attendance_map = {}
    weeks = set()
    for rec in Attendance.objects.filter(student__module=module).select_related("student"):
        attendance_map[(rec.student_id, rec.week_no)] = rec
        weeks.add(rec.week_no)
    fallback_weeks, fallback_map = _semester_register_fallback_from_daily(module, students)
    weeks.update(fallback_weeks)
    weeks = sorted(weeks)

    table = []

    for s in students:

        row = {
            "roll": s.roll_no,
            "enrollment": s.enrollment,
            "name": s.name,
            "mentor": s.mentor.name
        }

        overall = None

        for w in weeks:
            rec = attendance_map.get((s.id, w))
            if rec:
                row[f"week_{w}"] = rec.week_percentage
                overall = rec.overall_percentage
            else:
                fallback = fallback_map.get(s.id, {}).get(w)
                row[f"week_{w}"] = fallback.get("week_percentage") if fallback else None
                if fallback and fallback.get("overall_percentage") is not None:
                    overall = fallback["overall_percentage"]

        row["overall"] = overall
        table.append(row)

    return render(request, "semester_register.html", {
        "title": "Overall Attendance Register",
        "weeks": weeks,
        "rows": table
    })


def mentor_semester_register(request):
    mentor = _session_mentor_obj(request)
    if not mentor:
        return redirect("/")

    module = _active_module(request)
    students = list(
        Student.objects.select_related("mentor")
        .filter(module=module, mentor=mentor)
        .order_by("roll_no")
    )
    attendance_map = {}
    weeks = set()
    for rec in Attendance.objects.filter(student__module=module, student__mentor=mentor).select_related("student"):
        attendance_map[(rec.student_id, rec.week_no)] = rec
        weeks.add(rec.week_no)
    fallback_weeks, fallback_map = _semester_register_fallback_from_daily(module, students)
    weeks.update(fallback_weeks)
    weeks = sorted(weeks)

    table = []
    for s in students:
        row = {
            "roll": s.roll_no,
            "enrollment": s.enrollment,
            "name": s.name,
            "mentor": s.mentor.name,
        }
        overall = None
        for w in weeks:
            rec = attendance_map.get((s.id, w))
            if rec:
                row[f"week_{w}"] = rec.week_percentage
                overall = rec.overall_percentage
            else:
                fallback = fallback_map.get(s.id, {}).get(w)
                row[f"week_{w}"] = fallback.get("week_percentage") if fallback else None
                if fallback and fallback.get("overall_percentage") is not None:
                    overall = fallback["overall_percentage"]
        row["overall"] = overall
        table.append(row)

    return render(
        request,
        "semester_register.html",
        {
            "title": "My Mentees Attendance Register",
            "weeks": weeks,
            "rows": table,
        },
    )


# ---------------- DAILY ATTENDANCE (TIMETABLE) ----------------

def _calendar_for_module(module):
    return AcademicCalendar.objects.filter(module=module).first()


def _calendar_has_values(calendar):
    if not calendar:
        return False
    return any(
        [
            calendar.is_active,
            calendar.t1_start,
            calendar.t1_end,
            calendar.t2_start,
            calendar.t2_end,
            calendar.t3_start,
            calendar.t3_end,
            calendar.t4_start,
            calendar.t4_end,
        ]
    )


def _holiday_set(module):
    return set(
        AcademicHoliday.objects.filter(module=module, is_active=True).values_list("date", flat=True)
    )


def _attendance_block_reason(module, date_val):
    calendar = _calendar_for_module(module)
    if not calendar or not calendar.is_active:
        return "Academic calendar is inactive."
    if date_val > timezone.localdate():
        return "Attendance can only be filled up to today."
    if not phase_for_date(calendar, date_val):
        return "Date is outside academic calendar ranges."
    if date_val.weekday() == 6 or date_val in _holiday_set(module):
        return "Selected date is a holiday or Sunday."
    return ""


def _attendance_allowed_for_date(module, date_val):
    return _attendance_block_reason(module, date_val) == ""


def _parse_date_param(raw, fallback=None):
    if not raw:
        return fallback
    try:
        return datetime.strptime(raw, "%Y-%m-%d").date()
    except Exception:
        return fallback


def _slot_has_started(slot_date, slot_time):
    start = slot_start_time(slot_time)
    if not start:
        return False
    if slot_date < timezone.localdate():
        return True
    if slot_date > timezone.localdate():
        return False
    return timezone.localtime().time() >= time(start[0], start[1])


def _normalize_week_no(phase, week_no):
    phase = (phase or "").upper()
    if week_no is None:
        return None
    if phase == "T1":
        return week_no
    if phase == "T2":
        return 100 + week_no
    if phase == "T3":
        return 200 + week_no
    if phase == "T4":
        return 300 + week_no
    return week_no


def _attendance_lock_for_module_week(module, week_no):
    return WeekLock.objects.filter(module=module, week_no=week_no, locked=True).exists()


def _has_manual_week(module, week_no):
    return AttendanceWeekMeta.objects.filter(
        module=module, week_no=week_no, source=AttendanceWeekMeta.SOURCE_MANUAL
    ).exists()


def _create_calls_for_week(module, week_no):
    qs = Attendance.objects.filter(student__module=module, week_no=week_no, call_required=True)
    created = 0
    for att in qs.select_related("student"):
        if not CallRecord.objects.filter(student=att.student, week_no=week_no).exists():
            CallRecord.objects.create(student=att.student, week_no=week_no)
            created += 1
    return created


def recompute_weekly_attendance_from_daily(module, phase, week_no):
    calendar = _calendar_for_module(module)
    start, end = phase_range(calendar, phase)
    if not start or not end:
        return 0, "Academic calendar not configured."

    end_date = end_date_for_week(calendar, phase, week_no)
    if end_date and end_date > end:
        end_date = end
    if not end_date:
        return 0, "Week range invalid."

    start_monday = start - timedelta(days=start.weekday())
    week_start = start_monday + timedelta(days=(week_no - 1) * 7)
    week_end = week_start + timedelta(days=5)
    if week_start < start:
        week_start = start
    if week_end > end_date:
        week_end = end_date

    weekly_sessions = list(
        LectureSession.objects.filter(
            module=module,
            date__gte=week_start,
            date__lte=week_end,
        )
    )
    weekly_sessions = [s for s in weekly_sessions if _attendance_allowed_for_date(module, s.date)]

    overall_sessions = list(
        LectureSession.objects.filter(
            module=module,
            date__gte=start,
            date__lte=end_date,
        )
    )
    overall_sessions = [s for s in overall_sessions if _attendance_allowed_for_date(module, s.date)]

    if not weekly_sessions and not overall_sessions:
        return 0, "No lecture sessions found for selected week."

    week_session_ids = {s.id for s in weekly_sessions}
    overall_session_ids = [s.id for s in overall_sessions]
    if not overall_session_ids:
        return 0, "No lecture sessions found for selected week."

    absences = (
        LectureAbsence.objects.filter(session_id__in=overall_session_ids)
        .select_related("student", "session")
    )

    held_week_by_batch = {}
    for s in weekly_sessions:
        batch_key = _norm_batch_key(s.batch)
        if not batch_key:
            continue
        held_week_by_batch[batch_key] = held_week_by_batch.get(batch_key, 0) + 1

    held_overall_by_batch = {}
    for s in overall_sessions:
        batch_key = _norm_batch_key(s.batch)
        if not batch_key:
            continue
        held_overall_by_batch[batch_key] = held_overall_by_batch.get(batch_key, 0) + 1

    absent_week_by_student = {}
    absent_overall_by_student = {}
    for a in absences:
        absent_overall_by_student[a.student_id] = absent_overall_by_student.get(a.student_id, 0) + 1
        if a.session_id in week_session_ids:
            absent_week_by_student[a.student_id] = absent_week_by_student.get(a.student_id, 0) + 1

    normalized_week = _normalize_week_no(phase, week_no)
    updated = 0
    for student in Student.objects.filter(module=module):
        batch_keys = _student_batch_keys(student)
        held_week = sum(held_week_by_batch.get(key, 0) for key in batch_keys)
        held_overall = sum(held_overall_by_batch.get(key, 0) for key in batch_keys)

        absent_week = absent_week_by_student.get(student.id, 0)
        absent_overall = absent_overall_by_student.get(student.id, 0)

        attended_week = max(held_week - absent_week, 0)
        attended_overall = max(held_overall - absent_overall, 0)

        week_pct = round((attended_week / held_week) * 100, 2) if held_week else 0
        overall_pct = round((attended_overall / held_overall) * 100, 2) if held_overall else 0
        call_required = False
        if held_week and week_pct < 80:
            call_required = True
        if held_overall and overall_pct < 80:
            call_required = True
        Attendance.objects.update_or_create(
            week_no=normalized_week,
            student=student,
            defaults={
                "week_percentage": week_pct,
                "overall_percentage": overall_pct,
                "call_required": call_required,
            },
        )
        updated += 1

    AttendanceWeekMeta.objects.update_or_create(
        module=module,
        week_no=normalized_week,
        defaults={"source": AttendanceWeekMeta.SOURCE_AUTO},
    )
    _create_calls_for_week(module, normalized_week)
    return updated, None


def _recompute_weekly_attendance_async(module_id, phase, week_no):
    close_old_connections()
    try:
        module = AcademicModule.objects.filter(id=module_id, is_active=True).first()
        if not module:
            return
        normalized_week = _normalize_week_no(phase, week_no)
        if _attendance_lock_for_module_week(module, normalized_week):
            return
        if _has_manual_week(module, normalized_week):
            return
        recompute_weekly_attendance_from_daily(module, phase, week_no)
    except Exception:
        pass
    finally:
        close_old_connections()


@login_required
def upload_timetable(request):
    if request.session.get("mentor"):
        return redirect("/mentor-dashboard/")

    module = _active_module(request)
    _ensure_active_timetable(module)
    form = UploadFileForm(request.POST or None, request.FILES or None)
    uploads = TimetableUpload.objects.filter(module=module).order_by("-uploaded_at")[:5]

    if request.method == "POST" and request.POST.get("action") == "delete_all":
        TimetableEntry.objects.filter(module=module).delete()
        TimetableUpload.objects.filter(module=module).delete()
        messages.success(request, "All timetable entries deleted.")
        return redirect("/upload-timetable/")

    if request.method == "POST" and form.is_valid():
        file = form.cleaned_data["file"]
        try:
            entries, sheet_name = parse_timetable_excel(file)
        except Exception as exc:
            messages.error(request, f"Unable to parse timetable: {exc}")
            return render(
                request,
                "upload_timetable.html",
                {"form": form, "uploads": uploads, "module": module},
            )

        activate_mode = (request.POST.get("activate_mode") or "now").strip().lower()
        schedule_date = (request.POST.get("schedule_date") or "").strip()
        schedule_time = (request.POST.get("schedule_time") or "00:00").strip()
        effective_from = timezone.now()
        if activate_mode == "later":
            if not schedule_date:
                messages.error(request, "Select a schedule date to activate later.")
                return redirect("/upload-timetable/")
            try:
                date_val = _parse_date_param(schedule_date)
                hour, minute = [int(x) for x in schedule_time.split(":")]
                effective_from = timezone.make_aware(datetime.combine(date_val, time(hour, minute)))
            except Exception:
                messages.error(request, "Invalid schedule date/time.")
                return redirect("/upload-timetable/")

        created = 0
        skipped = 0
        upload = TimetableUpload.objects.create(
            module=module,
            uploaded_by=request.user.username,
            source_name=getattr(file, "name", "")[:255],
            rows_total=len(entries),
            rows_created=0,
            rows_skipped=0,
            is_active=False,
            effective_from=effective_from,
        )
        for entry in entries:
            batch = (entry.get("batch") or "").strip()
            subject = (entry.get("subject") or "").strip()
            if not batch or not subject:
                skipped += 1
                continue
            faculty = (entry.get("faculty") or "").strip().upper()
            if faculty:
                Mentor.objects.get_or_create(name=faculty)
            TimetableEntry.objects.update_or_create(
                module=module,
                day_of_week=entry["day_of_week"],
                lecture_no=entry["lecture_no"],
                batch=batch,
                upload=upload,
                defaults={
                    "time_slot": entry.get("time_slot") or "",
                    "subject": subject,
                    "faculty": faculty,
                    "room": (entry.get("room") or "").strip(),
                    "is_active": True,
                },
            )
            created += 1

        upload.rows_created = created
        upload.rows_skipped = skipped
        upload.save(update_fields=["rows_created", "rows_skipped"])

        if activate_mode != "later":
            TimetableUpload.objects.filter(module=module).update(is_active=False)
            upload.is_active = True
            upload.save(update_fields=["is_active"])
            TimetableEntry.objects.filter(module=module).update(is_active=False)
            TimetableEntry.objects.filter(module=module, upload=upload).update(is_active=True)
            _sync_subjects_from_timetable(module)
            messages.success(
                request,
                f"Timetable uploaded ({sheet_name}). Entries created: {created}, skipped: {skipped}. Activated now.",
            )
        else:
            TimetableEntry.objects.filter(module=module, upload=upload).update(is_active=False)
            _sync_subjects_from_timetable(module)
            messages.success(
                request,
                f"Timetable uploaded ({sheet_name}). Entries created: {created}, skipped: {skipped}. Scheduled from {effective_from:%d %b %Y %H:%M}.",
            )
        return redirect("/upload-timetable/")

    return render(
        request,
        "upload_timetable.html",
        {"form": form, "uploads": uploads, "module": module, "now": timezone.now()},
    )


def view_timetable(request):
    mentor = _session_mentor_obj(request)
    if not mentor and not request.user.is_authenticated:
        return redirect("/")
    module = _active_module(request)
    _ensure_active_timetable(module)
    choice_lists = _timetable_choice_lists(module)
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "edit":
            day_val = request.POST.get("day")
            lecture_raw = request.POST.get("lecture_no")
            batch = (request.POST.get("batch") or "").strip()
            subject = (request.POST.get("subject") or "").strip()
            faculty = (request.POST.get("faculty") or "").strip().upper()
            room = (request.POST.get("room") or "").strip()
            time_slot = (request.POST.get("time_slot") or "").strip()
            if not (day_val and lecture_raw and batch and subject):
                messages.error(request, "Day, lecture, batch, and subject are required.")
            else:
                try:
                    day_val = int(day_val)
                    lecture_no = int(lecture_raw)
                except Exception:
                    messages.error(request, "Invalid day or lecture number.")
                else:
                    active_upload = _active_upload_for_module(module)
                    if not active_upload:
                        messages.error(request, "No active timetable found. Upload and activate a timetable first.")
                    else:
                        entry, _ = TimetableEntry.objects.update_or_create(
                            module=module,
                            day_of_week=day_val,
                            lecture_no=lecture_no,
                            batch=batch,
                            upload=active_upload,
                            defaults={
                                "time_slot": time_slot,
                                "subject": subject,
                                "faculty": faculty,
                                "room": room,
                                "is_active": True,
                            },
                        )
                        TimetableEntry.objects.filter(
                            module=module,
                            day_of_week=day_val,
                            lecture_no=lecture_no,
                            batch=batch,
                        ).exclude(id=entry.id).update(is_active=False)
                        _sync_subjects_from_timetable(module)
                        messages.success(request, "Timetable updated.")
            return redirect("/view-timetable/")
    day_filter = request.GET.get("day")
    lecture_filter = (request.GET.get("lecture_no") or "").strip()
    batch_filter = (request.GET.get("batch") or "").strip()
    subject_filter = (request.GET.get("subject") or "").strip()
    faculty_filter = (request.GET.get("faculty") or "").strip()
    room_filter = (request.GET.get("room") or "").strip()

    qs = TimetableEntry.objects.filter(module=module, is_active=True).order_by("day_of_week", "lecture_no", "batch")
    if day_filter and str(day_filter).isdigit():
        qs = qs.filter(day_of_week=int(day_filter))
    if lecture_filter and str(lecture_filter).isdigit():
        qs = qs.filter(lecture_no=int(lecture_filter))
    if batch_filter:
        qs = qs.filter(batch__iexact=batch_filter)
    if subject_filter:
        qs = qs.filter(subject__iexact=subject_filter)
    if faculty_filter:
        qs = qs.filter(faculty__iexact=faculty_filter)
    if room_filter:
        qs = qs.filter(room__iexact=room_filter)

    day_choices = TimetableEntry.DAY_CHOICES
    grouped = {}
    for entry in qs:
        day_key = entry.day_of_week
        grouped.setdefault(day_key, []).append(entry)

    day_tables = []
    for day_key, entries in grouped.items():
        batch_set = sorted({e.batch for e in entries})
        lectures = sorted({e.lecture_no for e in entries})
        time_map = {}
        cell_map = {}
        for e in entries:
            time_map[e.lecture_no] = e.time_slot
            cell_map.setdefault(e.lecture_no, {})[e.batch] = e
        day_tables.append(
            {
                "day": dict(day_choices).get(day_key, str(day_key)),
                "day_key": day_key,
                "batches": batch_set,
                "lectures": lectures,
                "time_map": time_map,
                "cell_map": cell_map,
            }
        )

    day_tables.sort(key=lambda d: d["day_key"])

    return render(
        request,
        "view_timetable.html",
        {
            "day_filter": day_filter,
            "lecture_filter": lecture_filter,
            "batch_filter": batch_filter,
            "subject_filter": subject_filter,
            "faculty_filter": faculty_filter,
            "room_filter": room_filter,
            "day_choices": day_choices,
            "lecture_choices": choice_lists["lecture_choices"],
            "batches": choice_lists["batch_choices"],
            "subjects": choice_lists["subject_choices"],
            "faculties": choice_lists["faculty_choices"],
            "rooms": choice_lists["room_choices"],
            "day_tables": day_tables,
            "module": module,
        },
    )


def download_timetable_excel(request):
    mentor = _session_mentor_obj(request)
    if not mentor and not request.user.is_authenticated:
        return redirect("/")

    module = _active_module(request)
    _ensure_active_timetable(module)
    entries = list(
        TimetableEntry.objects.filter(module=module, is_active=True).order_by("day_of_week", "lecture_no", "batch")
    )
    if not entries:
        return HttpResponse("No timetable entries found.", status=404)

    day_labels = dict(TimetableEntry.DAY_CHOICES)
    batches = sorted({e.batch for e in entries if e.batch})
    day_lecture_map = {}
    time_map = {}
    cell_map = {}
    for e in entries:
        day_lecture_map.setdefault(e.day_of_week, set()).add(e.lecture_no)
        if e.time_slot:
            time_map.setdefault((e.day_of_week, e.lecture_no), e.time_slot)
        cell_map.setdefault((e.day_of_week, e.lecture_no), {})[e.batch] = e

    header = ["Day", "Lecture", "Time"]
    division_row = ["", "", ""]
    for b in batches:
        division_row.extend([b, "", ""])
        header.extend(["Subject", "Faculty", "Room"])

    rows = [division_row, header]
    for day in sorted(day_lecture_map.keys()):
        lectures = sorted(day_lecture_map.get(day, []))
        for lec in lectures:
            row = [day_labels.get(day, str(day)), lec, time_map.get((day, lec), "")]
            for b in batches:
                entry = cell_map.get((day, lec), {}).get(b)
                if entry:
                    row.extend([entry.subject, entry.faculty, entry.room])
                else:
                    row.extend(["", "", ""])
            rows.append(row)

    df = pd.DataFrame(rows)
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, header=False, sheet_name="Timetable")
    buffer.seek(0)

    filename = f"Timetable_{module.name.replace(' ', '_')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def academic_calendar(request):
    if request.session.get("mentor"):
        return redirect("/mentor-dashboard/")

    is_superadmin = is_superadmin_user(request.user)
    module = _active_module(request)
    calendar, _ = AcademicCalendar.objects.get_or_create(module=module)
    holidays = AcademicHoliday.objects.filter(module=module).order_by("-date")
    selected_year = (request.GET.get("year") or request.POST.get("year") or module.year_level or "FY").strip().upper()
    year_choices = [code for code, _ in AcademicModule.YEAR_CHOICES]
    if selected_year not in year_choices:
        selected_year = "FY"
    year_modules = list(
        AcademicModule.objects.filter(is_active=True, year_level=selected_year).order_by("variant", "semester", "name")
    )

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "calendar":
            calendar.is_active = bool(request.POST.get("is_active"))
            calendar.t1_start = _parse_date_param(request.POST.get("t1_start"))
            calendar.t1_end = _parse_date_param(request.POST.get("t1_end"))
            calendar.t2_start = _parse_date_param(request.POST.get("t2_start"))
            calendar.t2_end = _parse_date_param(request.POST.get("t2_end"))
            calendar.t3_start = _parse_date_param(request.POST.get("t3_start"))
            calendar.t3_end = _parse_date_param(request.POST.get("t3_end"))
            calendar.t4_start = _parse_date_param(request.POST.get("t4_start"))
            calendar.t4_end = _parse_date_param(request.POST.get("t4_end"))
            calendar.save()
            messages.success(request, "Academic calendar updated.")
            return redirect("/academic-calendar/")

        if action == "bulk_apply" and is_superadmin:
            module_ids = request.POST.getlist("module_ids")
            apply_is_active = bool(request.POST.get("is_active"))
            payload = {
                "is_active": apply_is_active,
                "t1_start": _parse_date_param(request.POST.get("t1_start")),
                "t1_end": _parse_date_param(request.POST.get("t1_end")),
                "t2_start": _parse_date_param(request.POST.get("t2_start")),
                "t2_end": _parse_date_param(request.POST.get("t2_end")),
                "t3_start": _parse_date_param(request.POST.get("t3_start")),
                "t3_end": _parse_date_param(request.POST.get("t3_end")),
                "t4_start": _parse_date_param(request.POST.get("t4_start")),
                "t4_end": _parse_date_param(request.POST.get("t4_end")),
            }
            if not module_ids:
                messages.error(request, "Select at least one module.")
            else:
                target_modules = AcademicModule.objects.filter(id__in=module_ids, is_active=True)
                applied = []
                skipped = []
                for target in target_modules:
                    target_calendar = AcademicCalendar.objects.filter(module=target).first()
                    if _calendar_has_values(target_calendar):
                        skipped.append(target.name)
                        continue
                    if not target_calendar:
                        target_calendar = AcademicCalendar(module=target)
                    target_calendar.is_active = payload["is_active"]
                    target_calendar.t1_start = payload["t1_start"]
                    target_calendar.t1_end = payload["t1_end"]
                    target_calendar.t2_start = payload["t2_start"]
                    target_calendar.t2_end = payload["t2_end"]
                    target_calendar.t3_start = payload["t3_start"]
                    target_calendar.t3_end = payload["t3_end"]
                    target_calendar.t4_start = payload["t4_start"]
                    target_calendar.t4_end = payload["t4_end"]
                    target_calendar.save()
                    applied.append(target.name)
                if applied:
                    messages.success(request, f"Applied calendar to {len(applied)} module(s).")
                if skipped:
                    messages.info(request, f"Skipped existing module calendars: {', '.join(skipped[:6])}{' ...' if len(skipped) > 6 else ''}")
            return redirect(f"/academic-calendar/?year={selected_year}")

        if action == "holiday_add":
            holiday_date = _parse_date_param(request.POST.get("holiday_date"))
            label = (request.POST.get("holiday_label") or "").strip()
            if holiday_date:
                target_modules = [module]
                if is_superadmin:
                    dept_keys = [d.strip().upper() for d in request.POST.getlist("holiday_depts") if d.strip()]
                    if dept_keys:
                        target_modules = [
                            m for m in AcademicModule.objects.filter(is_active=True)
                            if any(_dept_matches_module(m, key) for key in dept_keys)
                        ]
                    else:
                        target_modules = list(AcademicModule.objects.filter(is_active=True))
                for target in target_modules:
                    AcademicHoliday.objects.update_or_create(
                        module=target,
                        date=holiday_date,
                        defaults={"label": label, "is_active": True},
                    )
                messages.success(request, "Holiday added.")
            else:
                messages.error(request, "Select a valid holiday date.")
            return redirect("/academic-calendar/")

        if action == "holiday_update":
            holiday_id = request.POST.get("holiday_id")
            holiday_date = _parse_date_param(request.POST.get("holiday_date"))
            label = (request.POST.get("holiday_label") or "").strip()
            holiday = AcademicHoliday.objects.filter(id=holiday_id, module=module).first()
            if not holiday:
                messages.error(request, "Holiday not found.")
            elif not holiday_date:
                messages.error(request, "Select a valid holiday date.")
            else:
                conflict = AcademicHoliday.objects.filter(module=module, date=holiday_date).exclude(id=holiday.id).exists()
                if conflict:
                    messages.error(request, "Another holiday already exists on that date.")
                else:
                    holiday.date = holiday_date
                    holiday.label = label
                    holiday.is_active = True
                    holiday.save(update_fields=["date", "label", "is_active"])
                    messages.success(request, "Holiday updated.")
            return redirect("/academic-calendar/")

        if action == "holiday_delete":
            holiday_id = request.POST.get("holiday_id")
            AcademicHoliday.objects.filter(id=holiday_id, module=module).delete()
            messages.info(request, "Holiday removed.")
            return redirect("/academic-calendar/")

    return render(
        request,
        "academic_calendar.html",
        {
            "calendar": calendar,
            "holidays": holidays,
            "module": module,
            "is_superadmin": is_superadmin,
            "selected_year": selected_year,
            "year_choices": year_choices,
            "year_modules": year_modules,
        },
    )


def mentor_schedule(request):
    mentor = _session_mentor_obj(request)
    if not mentor:
        return redirect("/")

    dept_filter = (request.GET.get("dept") or "").strip().upper()
    modules = [m for m in allowed_modules_for_user(request) if m.is_active and _dept_matches_module(m, dept_filter)]
    if not modules:
        return render(
            request,
            "mentor_schedule.html",
            {
                "mentor": mentor,
                "entries": [],
                "selected_date": _parse_date_param(request.GET.get("date"), timezone.localdate()),
                "has_proxy_alert": False,
                "has_inactive_calendar": False,
                "dept_filter": dept_filter or "ALL",
            },
        )
    selected_date = _parse_date_param(request.GET.get("date"), timezone.localdate())
    day_of_week = selected_date.weekday()

    entries = []
    has_inactive_calendar = False
    for module in modules:
        _ensure_active_timetable(module)
        calendar = _calendar_for_module(module)
        if calendar and not calendar.is_active:
            has_inactive_calendar = True

        adjustments = _active_adjustments_for_date(module, selected_date)
        adj_map = {(a.batch, a.lecture_no): a for a in adjustments}

        merge_room_map = {
            (a.date, a.lecture_no): a.merge_room
            for a in adjustments
            if a.proxy_faculty and a.proxy_faculty.name.lower() == mentor.name.lower() and a.merge_room
        }
        original_entries = TimetableEntry.objects.filter(
            module=module, day_of_week=day_of_week, faculty__iexact=mentor.name, is_active=True
        )
        for entry in original_entries:
            merge_room = merge_room_map.get((selected_date, entry.lecture_no), "")
            adj = adj_map.get((entry.batch, entry.lecture_no))
            subject_val = entry.subject
            time_slot_val = entry.time_slot
            room_val = merge_room or entry.room
            if adj:
                subject_val = adj.subject or subject_val
                time_slot_val = adj.time_slot or time_slot_val
                room_val = adj.room or room_val
            entries.append(
                {
                    "module": module,
                    "module_name": module.name,
                    "dept_label": _dept_label_from_module(module),
                    "lecture_no": entry.lecture_no,
                    "time_slot": time_slot_val,
                    "batch": entry.batch,
                    "subject": subject_val,
                    "room": room_val,
                    "status": "original",
                    "proxy_faculty": adj.proxy_faculty.name if adj and adj.proxy_faculty else "",
                    "has_proxy": bool(adj),
                    "adjustment_type": adj.adjustment_type if adj else "",
                    "swap_with": f"{adj.swap_batch} L{adj.swap_lecture_no}" if adj and adj.adjustment_type == LectureAdjustment.TYPE_SWAP and adj.swap_batch and adj.swap_lecture_no else "",
                }
            )

        for adj in adjustments:
            if adj.adjustment_type == LectureAdjustment.TYPE_PROXY and adj.proxy_faculty and adj.proxy_faculty.name.lower() == mentor.name.lower():
                entries.append(
                    {
                        "module": module,
                        "module_name": module.name,
                        "dept_label": _dept_label_from_module(module),
                        "lecture_no": adj.lecture_no,
                        "time_slot": adj.time_slot,
                        "batch": adj.batch,
                        "subject": adj.subject,
                        "room": adj.room,
                        "status": "proxy",
                        "proxy_for": adj.original_faculty,
                        "has_proxy": True,
                        "adjustment_type": adj.adjustment_type,
                    }
                )

    has_proxy_alert = any(e.get("status") == "proxy" for e in entries)
    entries.sort(key=lambda x: (_slot_sort_key(x.get("time_slot")), x["lecture_no"], x["module_name"], x["batch"]))
    return render(
        request,
        "mentor_schedule.html",
        {
            "mentor": mentor,
            "entries": entries,
            "selected_date": selected_date,
            "has_proxy_alert": has_proxy_alert,
            "has_inactive_calendar": has_inactive_calendar,
            "dept_filter": dept_filter or "ALL",
        },
    )


def mentor_mark_attendance(request):
    mentor = _session_mentor_obj(request)
    if not mentor:
        return redirect("/")

    dept_filter = (request.GET.get("dept") or "").strip().upper()
    modules = [m for m in allowed_modules_for_user(request) if m.is_active and _dept_matches_module(m, dept_filter)]
    selected_date = _parse_date_param(request.GET.get("date"), timezone.localdate())
    is_holiday = selected_date.weekday() == 6
    has_inactive_calendar = False
    attendance_block_reason = ""
    allowed_modules = []
    for module in modules:
        _ensure_active_timetable(module)
        calendar = _calendar_for_module(module)
        if calendar and not calendar.is_active:
            has_inactive_calendar = True
        if selected_date in _holiday_set(module):
            is_holiday = True
        if _attendance_allowed_for_date(module, selected_date):
            allowed_modules.append(module)
        else:
            if not attendance_block_reason:
                attendance_block_reason = _attendance_block_reason(module, selected_date)

    batch_rows = []
    for module in allowed_modules:
        module_rows = _build_attendance_batch_rows(module, selected_date, mentor=mentor, allow_override=False)
        for row in module_rows:
            row["module"] = module
            row["module_name"] = module.name
            row["dept_label"] = _dept_label_from_module(module)
        batch_rows.extend(module_rows)
    batch_rows.sort(key=lambda row: (row.get("module_name", ""), row.get("batch", "")))

    return render(
        request,
        "mentor_mark_attendance.html",
        {
            "mentor": mentor,
            "batch_rows": batch_rows,
            "selected_date": selected_date,
            "is_holiday": is_holiday,
            "has_inactive_calendar": has_inactive_calendar,
            "dept_filter": dept_filter or "ALL",
            "attendance_block_reason": attendance_block_reason if not batch_rows else "",
        },
    )


def _norm_batch_key(value):
    return "".join(ch for ch in str(value or "").upper() if ch.isalnum())


def _norm_subject_key(value):
    return "".join(ch for ch in str(value or "").upper() if ch.isalnum())


def _subject_alias_map(module):
    alias_map = {}
    global_aliases = SubjectAlias.objects.filter(module__isnull=True, is_active=True)
    for alias in global_aliases:
        key = _norm_subject_key(alias.alias)
        if key:
            alias_map[key] = (alias.canonical or "").strip()
    module_aliases = SubjectAlias.objects.filter(module=module, is_active=True)
    for alias in module_aliases:
        key = _norm_subject_key(alias.alias)
        if key:
            alias_map[key] = (alias.canonical or "").strip()
    return alias_map


def _canonical_subject_name(module, raw_value, alias_map=None):
    raw = (raw_value or "").strip()
    if not raw:
        return ""
    if alias_map is None:
        alias_map = _subject_alias_map(module)
    key = _norm_subject_key(raw)
    mapped = alias_map.get(key)
    return mapped or raw


def _slot_sort_key(value):
    text = (value or "").strip()
    match = re.search(r"(\d{1,2})[:.]?(\d{2})", text)
    if not match:
        return (99, 99, text)
    return (int(match.group(1)), int(match.group(2)), text)


def _timetable_choice_lists(module):
    qs = TimetableEntry.objects.filter(module=module, is_active=True)
    return {
        "lecture_choices": sorted({int(n) for n in qs.values_list("lecture_no", flat=True) if n}),
        "batch_choices": sorted({(v or "").strip() for v in qs.values_list("batch", flat=True) if (v or "").strip()}),
        "subject_choices": sorted({(v or "").strip() for v in qs.values_list("subject", flat=True) if (v or "").strip()}),
        "faculty_choices": sorted({(v or "").strip() for v in qs.values_list("faculty", flat=True) if (v or "").strip()}),
        "room_choices": sorted({(v or "").strip() for v in qs.values_list("room", flat=True) if (v or "").strip()}),
    }


def _sync_subjects_from_timetable(module):
    timetable_subjects = sorted(
        {
            (name or "").strip()
            for name in TimetableEntry.objects.filter(module=module, is_active=True).values_list("subject", flat=True)
            if (name or "").strip()
        }
    )
    if not timetable_subjects:
        return 0

    existing_by_name = {
        (s.name or "").strip().lower(): s
        for s in Subject.objects.filter(module=module)
    }
    templates_by_name = {
        (t.name or "").strip().lower(): t
        for t in SubjectTemplate.objects.filter(is_active=True)
    }
    next_order = (Subject.objects.filter(module=module).aggregate(mx=Max("display_order")).get("mx") or 0) + 1
    created_count = 0

    for subject_name in timetable_subjects:
        key = subject_name.lower()
        existing = existing_by_name.get(key)
        template = templates_by_name.get(key)
        if existing:
            update_fields = []
            if not existing.is_active:
                existing.is_active = True
                update_fields.append("is_active")
            if template and existing.source_template_id != template.id:
                existing.source_template = template
                update_fields.append("source_template")
            if not existing.short_name:
                existing.short_name = template.short_name if template and template.short_name else subject_name
                update_fields.append("short_name")
            if update_fields:
                existing.save(update_fields=update_fields)
            continue

        Subject.objects.create(
            module=module,
            source_template=template,
            name=subject_name,
            short_name=(template.short_name if template and template.short_name else subject_name),
            display_order=next_order,
            has_theory=(template.has_theory if template else True),
            has_practical=(template.has_practical if template else True),
            result_format=(template.result_format if template else Subject.FORMAT_FULL),
            is_active=True,
        )
        next_order += 1
        created_count += 1
    return created_count


def _active_adjustments_for_date(module, selected_date):
    return list(
        LectureAdjustment.objects.filter(
            module=module,
            date=selected_date,
            status=LectureAdjustment.STATUS_ACTIVE,
        ).select_related("proxy_faculty", "created_by", "timetable_entry")
    )


def _cancel_adjustment_with_pair(adjustment, cancelled_by):
    pair_key = (adjustment.swap_pair_key or "").strip()
    qs = LectureAdjustment.objects.filter(id=adjustment.id)
    if pair_key:
        qs = LectureAdjustment.objects.filter(
            module=adjustment.module,
            date=adjustment.date,
            swap_pair_key=pair_key,
            status=LectureAdjustment.STATUS_ACTIVE,
        )
    qs.update(
        status=LectureAdjustment.STATUS_CANCELLED,
        cancelled_by=cancelled_by,
        cancelled_at=timezone.now(),
    )


def _swap_partner_choices(module, selected_date, source_entry, active_adjustments, include_same_faculty=False):
    adjusted_keys = {(a.batch, a.lecture_no) for a in active_adjustments}
    entries = (
        TimetableEntry.objects.filter(module=module, day_of_week=selected_date.weekday(), is_active=True)
        .exclude(id=source_entry.id)
        .exclude(faculty="")
        .order_by("lecture_no", "batch", "faculty")
    )
    partners = []
    source_key = (source_entry.batch, source_entry.lecture_no)
    for candidate in entries:
        candidate_key = (candidate.batch, candidate.lecture_no)
        if source_key in adjusted_keys or candidate_key in adjusted_keys:
            continue
        if not include_same_faculty and (candidate.faculty or "").strip().lower() == (source_entry.faculty or "").strip().lower():
            continue
        if _slot_has_started(selected_date, candidate.time_slot):
            continue
        partners.append(
            {
                "id": candidate.id,
                "label": f"L{candidate.lecture_no} · {candidate.time_slot} · {candidate.batch} · {candidate.faculty} · {candidate.subject}",
                "faculty": candidate.faculty,
                "batch": candidate.batch,
                "lecture_no": candidate.lecture_no,
                "time_slot": candidate.time_slot,
                "subject": candidate.subject,
            }
        )
    return partners


def _create_swap_adjustments(module, selected_date, entry_a, entry_b, created_by, remarks):
    swap_key = uuid.uuid4().hex
    defaults_common = {
        "adjustment_type": LectureAdjustment.TYPE_SWAP,
        "proxy_faculty": None,
        "merge_room": "",
        "remarks": remarks,
        "status": LectureAdjustment.STATUS_ACTIVE,
        "created_by": created_by,
        "cancelled_by": "",
        "cancelled_at": None,
        "swap_pair_key": swap_key,
    }
    LectureAdjustment.objects.update_or_create(
        module=module,
        date=selected_date,
        batch=entry_a.batch,
        lecture_no=entry_a.lecture_no,
        defaults={
            **defaults_common,
            "timetable_entry": entry_a,
            "time_slot": entry_b.time_slot,
            "subject": entry_a.subject,
            "original_faculty": entry_a.faculty,
            "room": entry_a.room,
            "swap_batch": entry_b.batch,
            "swap_lecture_no": entry_b.lecture_no,
            "swap_time_slot": entry_b.time_slot,
        },
    )


def _build_adjustment_rows(module, selected_date, faculty_filter="", exclude_proxy_name=""):
    day_of_week = selected_date.weekday()
    active_adjustments = _active_adjustments_for_date(module, selected_date)
    adjustment_map = {(a.batch, a.lecture_no): a for a in active_adjustments}
    active_modules = AcademicModule.objects.filter(is_active=True)
    faculty_choices = sorted(
        {
            (name or "").strip()
            for name in TimetableEntry.objects.filter(module=module, is_active=True)
            .exclude(faculty="")
            .values_list("faculty", flat=True)
        }
    )
    entries_qs = TimetableEntry.objects.filter(module=module, day_of_week=day_of_week, is_active=True)
    if faculty_filter:
        entries_qs = entries_qs.filter(faculty__iexact=faculty_filter)
    entries = entries_qs.order_by("lecture_no", "batch", "faculty")

    rooms_base = set(
        list(Room.objects.filter(module=module, is_active=True).values_list("name", flat=True))
        + list(
            TimetableEntry.objects.filter(module=module, is_active=True).exclude(room="").values_list("room", flat=True)
        )
    )

    rows = []
    for entry in entries:
        adj = adjustment_map.get((entry.batch, entry.lecture_no))
        slot_started = _slot_has_started(selected_date, (adj.time_slot if adj and adj.time_slot else entry.time_slot))
        slot_faculty_subjects = {}
        slot_entries = TimetableEntry.objects.filter(
            module=module, day_of_week=day_of_week, lecture_no=entry.lecture_no, is_active=True
        ).exclude(faculty="")
        for se in slot_entries:
            slot_faculty_subjects.setdefault(se.faculty, set()).add(se.subject)

        conflict_entries = list(
            TimetableEntry.objects.filter(
                module__in=active_modules,
                day_of_week=day_of_week,
                lecture_no=entry.lecture_no,
                is_active=True,
            )
            .exclude(faculty="")
        )
        conflict_adjustments = [
            item
            for item in LectureAdjustment.objects.filter(
                module__in=active_modules,
                date=selected_date,
                lecture_no=entry.lecture_no,
                status=LectureAdjustment.STATUS_ACTIVE,
            ).select_related("proxy_faculty")
        ]
        conflict_faculties = set(e.faculty for e in conflict_entries if e.faculty)
        conflict_depts = {
            str(e.faculty).strip().lower(): _dept_label_from_module(e.module)
            for e in conflict_entries
            if e.faculty
        }
        conflict_rooms = {
            str(e.faculty).strip().lower(): e.room
            for e in conflict_entries
            if e.faculty and e.room
        }
        for item in conflict_adjustments:
            faculty_name = (item.proxy_faculty.name if item.proxy_faculty else item.original_faculty or "").strip()
            if not faculty_name:
                continue
            conflict_faculties.add(faculty_name)
            conflict_depts.setdefault(faculty_name.lower(), _dept_label_from_module(item.module))
            if item.room:
                conflict_rooms[faculty_name.lower()] = item.room

        batch_faculties = []
        for fac in faculty_choices:
            if exclude_proxy_name and fac.lower() == exclude_proxy_name.lower():
                continue
            if fac.lower() == (entry.faculty or "").strip().lower():
                continue
            subject_label = ""
            subjects = slot_faculty_subjects.get(fac, set())
            if subjects:
                subject_label = sorted(s for s in subjects if s)[0] if subjects else ""
            batch_faculties.append(
                {
                    "name": fac,
                    "subject": subject_label,
                    "has_conflict": fac in conflict_faculties,
                    "conflict_dept": conflict_depts.get(fac.lower(), ""),
                }
            )
        batch_faculties.sort(key=lambda x: x["name"])

        used_rooms = set(
            TimetableEntry.objects.filter(
                module=module, day_of_week=day_of_week, lecture_no=entry.lecture_no, is_active=True
            )
            .exclude(room="")
            .values_list("room", flat=True)
        )
        used_rooms |= set(
            LectureAdjustment.objects.filter(
                module=module, date=selected_date, lecture_no=entry.lecture_no, status=LectureAdjustment.STATUS_ACTIVE
            ).exclude(room="").values_list("room", flat=True)
        )
        available_rooms = [r for r in sorted(r for r in rooms_base if r and r not in used_rooms)]
        if entry.room and entry.room not in available_rooms:
            available_rooms.insert(0, entry.room)

        rows.append(
            {
                "entry": entry,
                "adjustment": adj,
                "faculties": batch_faculties,
                "available_rooms": available_rooms,
                "slot_started": slot_started,
                "conflict_rooms": conflict_rooms,
                "available_rooms_json": json.dumps(available_rooms),
                "conflict_rooms_json": json.dumps(conflict_rooms),
                "swap_choices": _swap_partner_choices(module, selected_date, entry, active_adjustments),
            }
        )
    return rows, faculty_choices
    LectureAdjustment.objects.update_or_create(
        module=module,
        date=selected_date,
        batch=entry_b.batch,
        lecture_no=entry_b.lecture_no,
        defaults={
            **defaults_common,
            "timetable_entry": entry_b,
            "time_slot": entry_a.time_slot,
            "subject": entry_b.subject,
            "original_faculty": entry_b.faculty,
            "room": entry_b.room,
            "swap_batch": entry_a.batch,
            "swap_lecture_no": entry_a.lecture_no,
            "swap_time_slot": entry_a.time_slot,
        },
    )


def _build_attendance_batch_rows(module, selected_date, mentor=None, allow_override=False, prefill_absent=True):
    _ensure_active_timetable(module)
    day_of_week = selected_date.weekday()
    entries_qs = TimetableEntry.objects.filter(module=module, day_of_week=day_of_week, is_active=True)
    if mentor:
        entries_qs = entries_qs.filter(faculty__iexact=mentor.name)
    entries_qs = entries_qs.order_by("batch", "lecture_no")

    alias_map = _subject_alias_map(module)
    adjustments = _active_adjustments_for_date(module, selected_date)
    adj_by_key = {(a.batch, a.lecture_no): a for a in adjustments}
    merge_room_by_proxy = {
        (a.date, a.lecture_no, a.proxy_faculty.name.lower() if a.proxy_faculty else ""): a.merge_room
        for a in adjustments
        if a.merge_room and a.proxy_faculty
    }

    batch_map = {}
    for entry in entries_qs:
        batch_map.setdefault(entry.batch, []).append(
            {"entry": entry, "adjustment": adj_by_key.get((entry.batch, entry.lecture_no))}
        )

    if mentor:
        for adj in adjustments:
            if adj.proxy_faculty and adj.proxy_faculty.name.lower() == mentor.name.lower():
                proxy_entry = TimetableEntry(
                    module=module,
                    day_of_week=day_of_week,
                    lecture_no=adj.lecture_no,
                    time_slot=adj.time_slot,
                    batch=adj.batch,
                    subject=adj.subject,
                    faculty=mentor.name,
                    room=adj.room,
                )
                batch_map.setdefault(adj.batch, []).append({"entry": proxy_entry, "adjustment": adj})
    else:
        for adj in adjustments:
            if (adj.batch, adj.lecture_no) not in adj_by_key:
                continue

    students_all = list(Student.objects.filter(module=module).order_by("roll_no", "name"))
    students_by_batch = {}
    for student in students_all:
        for key in {_norm_batch_key(student.batch), _norm_batch_key(student.division)}:
            if key:
                students_by_batch.setdefault(key, []).append(student)

    batch_rows = []
    for batch, batch_entries in batch_map.items():
        students = students_by_batch.get(_norm_batch_key(batch), [])
        slots = []
        for item in batch_entries:
            entry = item["entry"]
            adj = item.get("adjustment")
            if adj:
                entry.subject = adj.subject or entry.subject
                entry.time_slot = adj.time_slot or entry.time_slot
                entry.room = adj.room or entry.room
                if adj.adjustment_type == LectureAdjustment.TYPE_PROXY and adj.proxy_faculty:
                    entry.faculty = adj.proxy_faculty.name
            if mentor:
                merge_room = merge_room_by_proxy.get((selected_date, entry.lecture_no, mentor.name.lower()), "")
                if merge_room and entry.faculty.lower() == mentor.name.lower():
                    entry.room = merge_room
            session = LectureSession.objects.filter(
                module=module, date=selected_date, lecture_no=entry.lecture_no, batch=batch
            ).first()
            absent_rolls = set()
            if session and prefill_absent:
                absences = LectureAbsence.objects.filter(session=session).select_related("student")
                absent_rolls = {a.student.roll_no for a in absences if a.student.roll_no is not None}
            slots.append(
                {
                    "entry": entry,
                    "subject_display": _canonical_subject_name(module, entry.subject, alias_map),
                    "session": session,
                    "absent_rolls": absent_rolls,
                    "form_id": f"form-{_norm_batch_key(batch)}-{entry.lecture_no}",
                    "adjustment": adj,
                    "can_edit": allow_override or not (adj and adj.adjustment_type == LectureAdjustment.TYPE_PROXY and adj.proxy_faculty and mentor and adj.proxy_faculty.name != mentor.name),
                }
            )
        slots.sort(key=lambda s: (_slot_sort_key(s["entry"].time_slot), s["entry"].lecture_no))
        previous_slot = None
        for slot in slots:
            can_copy_prev = False
            if previous_slot:
                prev_lecture = previous_slot["entry"].lecture_no
                curr_lecture = slot["entry"].lecture_no
                can_copy_prev = bool(prev_lecture and curr_lecture and curr_lecture == prev_lecture + 1)
            slot["can_copy_prev"] = can_copy_prev
            previous_slot = slot
        batch_rows.append({"batch": batch, "students": students, "slots": slots})
    batch_rows.sort(key=lambda row: row["batch"])
    return batch_rows


@login_required
def coordinator_mark_attendance(request):
    if request.session.get("mentor"):
        return redirect("/mentor-dashboard/")

    module = _active_module(request)
    _ensure_active_timetable(module)
    selected_date = _parse_date_param(request.GET.get("date"), timezone.localdate())
    calendar = _calendar_for_module(module)
    holiday_dates = _holiday_set(module)
    is_holiday = selected_date.weekday() == 6 or selected_date in holiday_dates
    attendance_block_reason = _attendance_block_reason(module, selected_date)
    batch_rows = []
    if not attendance_block_reason:
        batch_rows = _build_attendance_batch_rows(
            module, selected_date, mentor=None, allow_override=True, prefill_absent=True
        )

    return render(
        request,
        "coordinator_mark_attendance.html",
        {
            "batch_rows": batch_rows,
            "selected_date": selected_date,
            "calendar": calendar,
            "is_holiday": is_holiday,
            "module": module,
            "attendance_block_reason": attendance_block_reason,
        },
    )


def mentor_load_adjustment(request):
    mentor = _session_mentor_obj(request)
    if not mentor:
        return redirect("/")

    module = _active_module(request)
    _ensure_active_timetable(module)
    selected_date = _parse_date_param(request.GET.get("date"), timezone.localdate())
    day_of_week = selected_date.weekday()

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "create_proxy":
            entry_id = request.POST.get("entry_id")
            proxy_name = (request.POST.get("proxy_faculty") or "").strip()
            proxy_subject = (request.POST.get("proxy_subject") or "").strip()
            room_select = (request.POST.get("room_select") or "").strip()
            room_custom = (request.POST.get("room_custom") or "").strip()
            merge_room = (request.POST.get("merge_room") or "").strip()
            remarks = (request.POST.get("remarks") or "").strip()
            entry = TimetableEntry.objects.filter(id=entry_id, module=module, is_active=True).first()
            if not entry:
                messages.error(request, "Lecture not found.")
                return redirect(f"/mentor-load-adjustment/?date={selected_date:%Y-%m-%d}")

            if _slot_has_started(selected_date, entry.time_slot):
                messages.error(request, "Lecture already started. Adjustment not allowed.")
                return redirect(f"/mentor-load-adjustment/?date={selected_date:%Y-%m-%d}")

            proxy = Mentor.objects.filter(name__iexact=proxy_name).first()
            if not proxy:
                messages.error(request, "Select a valid proxy faculty.")
                return redirect(f"/mentor-load-adjustment/?date={selected_date:%Y-%m-%d}")

            room = merge_room or room_custom or room_select or entry.room
            active_modules = AcademicModule.objects.filter(is_active=True)
            conflict = TimetableEntry.objects.filter(
                module__in=active_modules,
                day_of_week=day_of_week,
                lecture_no=entry.lecture_no,
                faculty__iexact=proxy.name,
                is_active=True,
            ).exists()
            if conflict and not merge_room:
                messages.error(request, "Proxy faculty already has a lecture. Merge room required.")
                return redirect(f"/mentor-load-adjustment/?date={selected_date:%Y-%m-%d}")
            proxy_slot_subject = (
                TimetableEntry.objects.filter(
                    module__in=active_modules,
                    day_of_week=day_of_week,
                    lecture_no=entry.lecture_no,
                    faculty__iexact=proxy.name,
                    is_active=True,
                )
                .values_list("subject", flat=True)
                .first()
            )
            LectureAdjustment.objects.update_or_create(
                module=module,
                date=selected_date,
                batch=entry.batch,
                lecture_no=entry.lecture_no,
                defaults={
                    "timetable_entry": entry,
                    "adjustment_type": LectureAdjustment.TYPE_PROXY,
                    "time_slot": entry.time_slot,
                    "subject": proxy_slot_subject or proxy_subject or entry.subject,
                    "original_faculty": entry.faculty,
                    "proxy_faculty": proxy,
                    "room": room,
                    "merge_room": merge_room,
                    "swap_pair_key": "",
                    "swap_batch": "",
                    "swap_lecture_no": None,
                    "swap_time_slot": "",
                    "remarks": remarks,
                    "status": LectureAdjustment.STATUS_ACTIVE,
                    "created_by": mentor,
                    "cancelled_by": "",
                    "cancelled_at": None,
                },
            )
            messages.success(request, "Adjustment saved.")
            return redirect(f"/mentor-load-adjustment/?date={selected_date:%Y-%m-%d}")

        if action == "create_swap":
            entry_id = request.POST.get("entry_id")
            partner_id = request.POST.get("swap_entry_id")
            remarks = (request.POST.get("remarks") or "").strip()
            entry = TimetableEntry.objects.filter(
                id=entry_id,
                module=module,
                day_of_week=day_of_week,
                faculty__iexact=mentor.name,
                is_active=True,
            ).first()
            partner = TimetableEntry.objects.filter(
                id=partner_id,
                module=module,
                day_of_week=day_of_week,
                is_active=True,
            ).exclude(id=entry_id).first()
            if not entry or not partner:
                messages.error(request, "Select valid lectures to swap.")
                return redirect(f"/mentor-load-adjustment/?date={selected_date:%Y-%m-%d}")
            if _slot_has_started(selected_date, entry.time_slot) or _slot_has_started(selected_date, partner.time_slot):
                messages.error(request, "Lecture already started. Swap not allowed.")
                return redirect(f"/mentor-load-adjustment/?date={selected_date:%Y-%m-%d}")
            active_keys = {
                (a.batch, a.lecture_no)
                for a in _active_adjustments_for_date(module, selected_date)
            }
            if (entry.batch, entry.lecture_no) in active_keys or (partner.batch, partner.lecture_no) in active_keys:
                messages.error(request, "One of the selected lectures already has an active adjustment.")
                return redirect(f"/mentor-load-adjustment/?date={selected_date:%Y-%m-%d}")
            _create_swap_adjustments(module, selected_date, entry, partner, mentor, remarks)
            messages.success(request, "Lecture swap saved.")
            return redirect(f"/mentor-load-adjustment/?date={selected_date:%Y-%m-%d}")

        if action == "cancel":
            adj_id = request.POST.get("adjustment_id")
            adj = LectureAdjustment.objects.filter(id=adj_id, created_by=mentor, status=LectureAdjustment.STATUS_ACTIVE).first()
            if not adj:
                messages.error(request, "Adjustment not found or not allowed.")
                return redirect(f"/mentor-load-adjustment/?date={selected_date:%Y-%m-%d}")
            if _slot_has_started(adj.date, adj.time_slot):
                messages.error(request, "Lecture already started. Cancellation not allowed.")
                return redirect(f"/mentor-load-adjustment/?date={selected_date:%Y-%m-%d}")
            _cancel_adjustment_with_pair(adj, mentor.name)
            messages.success(request, "Adjustment cancelled.")
            return redirect(f"/mentor-load-adjustment/?date={selected_date:%Y-%m-%d}")

    rows, _ = _build_adjustment_rows(module, selected_date, faculty_filter=mentor.name, exclude_proxy_name=mentor.name)

    return render(
        request,
        "mentor_load_adjustment.html",
        {
            "mentor": mentor,
            "is_coordinator": False,
            "rows": rows,
            "selected_date": selected_date,
            "faculty_choices": [],
            "selected_faculty": "",
        },
    )


@login_required
def coordinator_load_adjustment(request):
    if request.session.get("mentor"):
        return redirect("/mentor-dashboard/")

    module = _active_module(request)
    _ensure_active_timetable(module)
    selected_date = _parse_date_param(request.GET.get("date"), timezone.localdate())
    selected_faculty = (request.GET.get("faculty") or request.POST.get("faculty") or "").strip()
    day_of_week = selected_date.weekday()

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "create_proxy":
            entry_id = request.POST.get("entry_id")
            proxy_name = (request.POST.get("proxy_faculty") or "").strip()
            room_select = (request.POST.get("room_select") or "").strip()
            room_custom = (request.POST.get("room_custom") or "").strip()
            merge_room = (request.POST.get("merge_room") or "").strip()
            remarks = (request.POST.get("remarks") or "").strip()
            entry = TimetableEntry.objects.filter(id=entry_id, module=module, is_active=True).first()
            if not entry:
                messages.error(request, "Lecture not found.")
                return redirect(f"/coordinator-load-adjustment/?date={selected_date:%Y-%m-%d}")

            if _slot_has_started(selected_date, entry.time_slot):
                messages.error(request, "Lecture already started. Adjustment not allowed.")
                return redirect(f"/coordinator-load-adjustment/?date={selected_date:%Y-%m-%d}")

            proxy = Mentor.objects.filter(name__iexact=proxy_name).first()
            if not proxy:
                messages.error(request, "Select a valid proxy faculty.")
                return redirect(f"/coordinator-load-adjustment/?date={selected_date:%Y-%m-%d}")

            room = merge_room or room_custom or room_select or entry.room
            active_modules = AcademicModule.objects.filter(is_active=True)
            conflict = TimetableEntry.objects.filter(
                module__in=active_modules,
                day_of_week=day_of_week,
                lecture_no=entry.lecture_no,
                faculty__iexact=proxy.name,
                is_active=True,
            ).exists()
            if conflict and not merge_room:
                messages.error(request, "Proxy faculty already has a lecture. Merge room required.")
                return redirect(f"/coordinator-load-adjustment/?date={selected_date:%Y-%m-%d}")
            proxy_slot_subject = (
                TimetableEntry.objects.filter(
                    module__in=active_modules,
                    day_of_week=day_of_week,
                    lecture_no=entry.lecture_no,
                    faculty__iexact=proxy.name,
                    is_active=True,
                )
                .values_list("subject", flat=True)
                .first()
            )
            created_by = Mentor.objects.filter(name__iexact=request.user.username).first()
            LectureAdjustment.objects.update_or_create(
                module=module,
                date=selected_date,
                batch=entry.batch,
                lecture_no=entry.lecture_no,
                defaults={
                    "timetable_entry": entry,
                    "adjustment_type": LectureAdjustment.TYPE_PROXY,
                    "time_slot": entry.time_slot,
                    "subject": proxy_slot_subject or entry.subject,
                    "original_faculty": entry.faculty,
                    "proxy_faculty": proxy,
                    "room": room,
                    "merge_room": merge_room,
                    "swap_pair_key": "",
                    "swap_batch": "",
                    "swap_lecture_no": None,
                    "swap_time_slot": "",
                    "remarks": remarks,
                    "status": LectureAdjustment.STATUS_ACTIVE,
                    "created_by": created_by,
                    "cancelled_by": "",
                    "cancelled_at": None,
                },
            )
            messages.success(request, "Proxy assigned.")
        elif action == "create_swap":
            entry_id = request.POST.get("entry_id")
            partner_id = request.POST.get("swap_entry_id")
            remarks = (request.POST.get("remarks") or "").strip()
            entry = TimetableEntry.objects.filter(
                id=entry_id,
                module=module,
                day_of_week=day_of_week,
                is_active=True,
            ).first()
            partner = TimetableEntry.objects.filter(
                id=partner_id,
                module=module,
                day_of_week=day_of_week,
                is_active=True,
            ).exclude(id=entry_id).first()
            if not entry or not partner:
                messages.error(request, "Select valid lectures to swap.")
                return redirect(f"/coordinator-load-adjustment/?date={selected_date:%Y-%m-%d}")
            if _slot_has_started(selected_date, entry.time_slot) or _slot_has_started(selected_date, partner.time_slot):
                messages.error(request, "Lecture already started. Swap not allowed.")
                return redirect(f"/coordinator-load-adjustment/?date={selected_date:%Y-%m-%d}")
            active_keys = {
                (a.batch, a.lecture_no)
                for a in _active_adjustments_for_date(module, selected_date)
            }
            if (entry.batch, entry.lecture_no) in active_keys or (partner.batch, partner.lecture_no) in active_keys:
                messages.error(request, "One of the selected lectures already has an active adjustment.")
                return redirect(f"/coordinator-load-adjustment/?date={selected_date:%Y-%m-%d}")
            created_by = Mentor.objects.filter(name__iexact=request.user.username).first()
            _create_swap_adjustments(module, selected_date, entry, partner, created_by, remarks)
            messages.success(request, "Lecture swap saved.")
        elif action == "cancel":
            adj_id = request.POST.get("adjustment_id")
            adj = LectureAdjustment.objects.filter(id=adj_id, status=LectureAdjustment.STATUS_ACTIVE).first()
            if not adj:
                messages.error(request, "Adjustment not found.")
            elif _slot_has_started(adj.date, adj.time_slot):
                messages.error(request, "Lecture already started. Cancellation not allowed.")
            else:
                _cancel_adjustment_with_pair(adj, request.user.username)
                messages.success(request, "Adjustment cancelled.")
        return redirect(f"/coordinator-load-adjustment/?date={selected_date:%Y-%m-%d}")

    rows, faculty_choices = _build_adjustment_rows(module, selected_date, faculty_filter=selected_faculty)

    return render(
        request,
        "mentor_load_adjustment.html",
        {
            "mentor": None,
            "is_coordinator": True,
            "rows": rows,
            "selected_date": selected_date,
            "faculty_choices": faculty_choices,
            "selected_faculty": selected_faculty,
        },
    )


@login_required
def coordinator_adjustments(request):
    if request.session.get("mentor"):
        return redirect("/mentor-dashboard/")

    module = _active_module(request)
    _ensure_active_timetable(module)
    start_date = _parse_date_param(request.GET.get("start_date"), timezone.localdate())
    week_start = start_date - timedelta(days=start_date.weekday())
    week_end = week_start + timedelta(days=6)

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "cancel":
            adj_id = request.POST.get("adjustment_id")
            adj = LectureAdjustment.objects.filter(id=adj_id, status=LectureAdjustment.STATUS_ACTIVE).first()
            if not adj:
                messages.error(request, "Adjustment not found.")
            elif _slot_has_started(adj.date, adj.time_slot):
                messages.error(request, "Lecture already started. Cancellation not allowed.")
            else:
                adj.status = LectureAdjustment.STATUS_CANCELLED
                adj.cancelled_by = request.user.username
                adj.cancelled_at = timezone.now()
                adj.save(update_fields=["status", "cancelled_by", "cancelled_at"])
                messages.success(request, "Adjustment cancelled.")
            return redirect(f"/coordinator-adjustments/?start_date={week_start:%Y-%m-%d}")

    adjustments = LectureAdjustment.objects.filter(
        module=module,
        date__gte=week_start,
        date__lte=week_end,
    ).select_related("proxy_faculty", "created_by")
    adj_list = list(adjustments)
    if adj_list:
        day_set = {a.date.weekday() for a in adj_list}
        lecture_set = {a.lecture_no for a in adj_list}
        entries = list(
            TimetableEntry.objects.filter(
                module=module,
                day_of_week__in=list(day_set),
                lecture_no__in=list(lecture_set),
                is_active=True,
            )
        )
        entry_by_batch_fac = {}
        entry_by_fac = {}
        for e in entries:
            key = (e.day_of_week, e.lecture_no, (e.faculty or "").strip().lower(), (e.batch or "").strip().lower())
            entry_by_batch_fac[key] = e
            fac_key = (e.day_of_week, e.lecture_no, (e.faculty or "").strip().lower())
            entry_by_fac.setdefault(fac_key, []).append(e)
        for a in adj_list:
            if a.adjustment_type == LectureAdjustment.TYPE_PROXY:
                proxy_name = (a.proxy_faculty.name if a.proxy_faculty else "").strip().lower()
                proxy_subject = ""
                if proxy_name:
                    key = (a.date.weekday(), a.lecture_no, proxy_name, (a.batch or "").strip().lower())
                    match = entry_by_batch_fac.get(key)
                    if not match:
                        fac_list = entry_by_fac.get((a.date.weekday(), a.lecture_no, proxy_name), [])
                        match = fac_list[0] if fac_list else None
                    if match:
                        proxy_subject = match.subject or ""
                if not proxy_subject:
                    proxy_subject = a.subject or ""
                a.proxy_subject = proxy_subject
            else:
                a.proxy_subject = ""
            a.partner_label = ""
            if a.adjustment_type == LectureAdjustment.TYPE_SWAP and a.swap_batch and a.swap_lecture_no:
                a.partner_label = f"{a.swap_batch} · L{a.swap_lecture_no}"

    return render(
        request,
        "coordinator_adjustments.html",
        {
            "module": module,
            "adjustments": adj_list,
            "week_start": week_start,
            "week_end": week_end,
        },
    )


@login_required
def manage_rooms(request):
    if request.session.get("mentor"):
        return redirect("/mentor-dashboard/")

    module = _active_module(request)
    _ensure_active_timetable(module)
    timetable_rooms = sorted(
        {
            (r or "").strip()
            for r in TimetableEntry.objects.filter(module=module, is_active=True)
            .exclude(room="")
            .values_list("room", flat=True)
        }
    )
    timetable_rooms = [r for r in timetable_rooms if r]
    for name in timetable_rooms:
        Room.objects.update_or_create(module=module, name=name, defaults={"is_active": True})
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "add":
            name = (request.POST.get("name") or "").strip()
            if name:
                Room.objects.update_or_create(module=module, name=name, defaults={"is_active": True})
                messages.success(request, "Room added.")
        elif action == "update":
            room_id = request.POST.get("room_id")
            name = (request.POST.get("name") or "").strip()
            is_active = request.POST.get("is_active") == "1"
            room = Room.objects.filter(id=room_id, module=module).first()
            if room and name:
                room.name = name
                room.is_active = is_active
                room.save(update_fields=["name", "is_active"])
                messages.success(request, "Room updated.")
        elif action == "delete":
            room_id = request.POST.get("room_id")
            Room.objects.filter(id=room_id, module=module).delete()
            messages.success(request, "Room deleted.")
        return redirect("/manage-rooms/")

    rooms = Room.objects.filter(module=module).order_by("name")
    return render(
        request,
        "manage_rooms.html",
        {
            "module": module,
            "rooms": rooms,
        },
    )


def mentor_daily_absentees(request):
    mentor = _session_mentor_obj(request)
    if not mentor:
        return redirect("/")

    module = _active_module(request)
    date_val = _parse_date_param(request.GET.get("date"), timezone.localdate())
    calendar = _calendar_for_module(module)
    phase, week_no = week_for_date(calendar, date_val)
    day_no = date_val.weekday() + 1

    sessions = LectureSession.objects.filter(module=module, date=date_val)
    absences = (
        LectureAbsence.objects.filter(session__in=sessions, student__mentor=mentor)
        .select_related("student", "session")
        .order_by("student__roll_no", "session__lecture_no")
    )

    student_map = {}
    for a in absences:
        entry = student_map.setdefault(
            a.student_id,
            {
                "student": a.student,
                "count": 0,
                "details": [],
            },
        )
        entry["count"] += 1
        entry["details"].append(
            {
                "lecture_no": a.session.lecture_no,
                "subject": a.session.subject,
                "batch": a.session.batch,
            }
        )

    records = []
    for item in student_map.values():
        student = item["student"]
        call = (
            OtherCallRecord.objects.filter(
                mentor=mentor,
                student=student,
                call_category="less_attendance",
                created_at__date=date_val,
            )
            .order_by("-created_at")
            .first()
        )
        if not call:
            call = OtherCallRecord.objects.create(
                mentor=mentor,
                student=student,
                call_category="less_attendance",
            )
        records.append(
            {
                "call": call,
                "absent_count": item["count"],
                "details": item["details"],
            }
        )

    return render(
        request,
        "mentor_daily_absentees.html",
        {
            "mentor": mentor,
            "records": records,
            "selected_date": date_val,
            "week_no": week_no or 1,
            "day_no": day_no,
            "calendar": calendar,
        },
    )


@login_required
def attendance_fill_status(request):
    if request.session.get("mentor"):
        return redirect("/mentor-dashboard/")

    module = _active_module(request)
    _ensure_active_timetable(module)
    today = timezone.localdate()
    start_date = _parse_date_param(request.GET.get("start_date"))
    end_date = _parse_date_param(request.GET.get("end_date"))
    if not start_date or not end_date:
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=5)
    if start_date > end_date:
        start_date, end_date = end_date, start_date

    def _rows_for_date(date_val):
        block_reason = _attendance_block_reason(module, date_val)
        if block_reason:
            return {"date": date_val, "rows": [], "blocked_reason": block_reason}
        day_of_week = date_val.weekday()
        expected_entries = TimetableEntry.objects.filter(
            module=module, day_of_week=day_of_week, is_active=True
        )
        expected_map = {}
        for entry in expected_entries:
            faculty = (entry.faculty or "").strip().upper()
            expected_map.setdefault(faculty, []).append(entry)

        sessions = LectureSession.objects.filter(module=module, date=date_val)
        session_map = {}
        for sess in sessions:
            faculty = (sess.faculty or "").strip().upper()
            session_map.setdefault(faculty, []).append(sess)

        rows = []
        for faculty, entries in expected_map.items():
            marked = session_map.get(faculty, [])
            marked_keys = {(m.batch, m.lecture_no) for m in marked}
            missing = [e for e in entries if (e.batch, e.lecture_no) not in marked_keys]
            rows.append(
                {
                    "faculty": faculty,
                    "expected": len(entries),
                    "marked": len(marked),
                    "pending": len(missing),
                    "missing": missing,
                }
            )

        rows.sort(key=lambda r: (r["pending"], r["faculty"]))
        return {"date": date_val, "rows": rows, "blocked_reason": ""}

    week_rows = []
    cur = start_date
    while cur <= end_date:
        if cur.weekday() != 6:
            week_rows.append(_rows_for_date(cur))
        cur += timedelta(days=1)

    return render(
        request,
        "attendance_fill_status.html",
        {
            "module": module,
            "start_date": start_date,
            "end_date": end_date,
            "week_rows": week_rows,
        },
    )


@require_http_methods(["POST"])
def save_lecture_attendance(request):
    mentor = _session_mentor_obj(request)
    is_coordinator = bool(request.user.is_authenticated and not request.session.get("mentor"))
    if not mentor and not is_coordinator:
        return JsonResponse({"ok": False, "msg": "Unauthorized"}, status=401)

    module = None
    module_id = (request.POST.get("module_id") or "").strip()
    if mentor and module_id:
        module = allowed_modules_for_user(request).filter(id=module_id, is_active=True).first()
        if not module:
            return JsonResponse({"ok": False, "msg": "Module not allowed"}, status=403)
    if not module:
        module = _active_module(request)
    _ensure_active_timetable(module)
    date_val = _parse_date_param(request.POST.get("date"))
    batch = (request.POST.get("batch") or "").strip()
    lecture_no_raw = request.POST.get("lecture_no")
    if not date_val or not batch or not lecture_no_raw:
        return JsonResponse({"ok": False, "msg": "Missing required fields"}, status=400)
    if not _attendance_allowed_for_date(module, date_val):
        return JsonResponse({"ok": False, "msg": _attendance_block_reason(module, date_val)}, status=400)

    try:
        lecture_no = int(lecture_no_raw)
    except Exception:
        return JsonResponse({"ok": False, "msg": "Invalid lecture number"}, status=400)

    adjustment_id = request.POST.get("adjustment_id")
    adjustment = None
    entry = None
    if adjustment_id:
        adjustment = LectureAdjustment.objects.filter(
            id=adjustment_id,
            module=module,
            date=date_val,
            batch=batch,
            lecture_no=lecture_no,
            status=LectureAdjustment.STATUS_ACTIVE,
        ).first()
        if (
            adjustment
            and adjustment.adjustment_type == LectureAdjustment.TYPE_PROXY
            and mentor
            and (not adjustment.proxy_faculty or adjustment.proxy_faculty.name.lower() != mentor.name.lower())
        ):
            adjustment = None
        if not adjustment:
            return JsonResponse({"ok": False, "msg": "Adjustment not found"}, status=404)
    else:
        entry_qs = TimetableEntry.objects.filter(
            module=module,
            day_of_week=date_val.weekday(),
            lecture_no=lecture_no,
            batch=batch,
            is_active=True,
        )
        if mentor:
            entry_qs = entry_qs.filter(faculty__iexact=mentor.name)
        entry = entry_qs.first()
        if not entry:
            return JsonResponse({"ok": False, "msg": "Timetable entry not found"}, status=404)
        active_adj = LectureAdjustment.objects.filter(
            module=module,
            date=date_val,
            batch=batch,
            lecture_no=lecture_no,
            status=LectureAdjustment.STATUS_ACTIVE,
        ).first()
        if (
            active_adj
            and active_adj.adjustment_type == LectureAdjustment.TYPE_PROXY
            and mentor
            and (not active_adj.proxy_faculty or active_adj.proxy_faculty.name.lower() != mentor.name.lower())
        ):
            return JsonResponse({"ok": False, "msg": "Proxy assigned. Original faculty cannot mark attendance."}, status=403)

    session, _ = LectureSession.objects.update_or_create(
        module=module,
        date=date_val,
        lecture_no=lecture_no,
        batch=batch,
        defaults={
            "timetable_entry": entry or adjustment.timetable_entry if adjustment else entry,
            "day_of_week": date_val.weekday(),
            "time_slot": (adjustment.time_slot if adjustment else entry.time_slot),
            "subject": (adjustment.subject if adjustment else entry.subject),
            "faculty": (
                (
                    adjustment.proxy_faculty.name
                    if adjustment and adjustment.adjustment_type == LectureAdjustment.TYPE_PROXY and adjustment.proxy_faculty
                    else (mentor.name if mentor else entry.faculty)
                )
                if adjustment
                else entry.faculty
            ),
            "room": (adjustment.room if adjustment else entry.room),
            "marked_by": mentor if mentor else None,
        },
    )

    LectureAbsence.objects.filter(session=session).delete()
    absent_rolls = request.POST.getlist("absent_roll_numbers")
    roll_numbers = []
    for roll in absent_rolls:
        try:
            roll_no = int(str(roll).strip())
        except Exception:
            continue
        roll_numbers.append(roll_no)

    if roll_numbers:
        batch_key = _norm_batch_key(batch)
        students = list(Student.objects.filter(module=module, roll_no__in=roll_numbers))
        student_map = {}
        for s in students:
            if not batch_key or batch_key in _student_batch_keys(s):
                student_map[s.roll_no] = s
        absences = [
            LectureAbsence(session=session, student=student_map[rn], marked_by=mentor)
            for rn in roll_numbers
            if rn in student_map
        ]
        if absences:
            LectureAbsence.objects.bulk_create(absences)

    calendar = _calendar_for_module(module)
    phase, week_no = week_for_date(calendar, date_val)
    if phase and week_no:
        normalized_week = _normalize_week_no(phase, week_no)
        if not _attendance_lock_for_module_week(module, normalized_week) and not _has_manual_week(module, normalized_week):
            threading.Thread(
                target=_recompute_weekly_attendance_async,
                kwargs={"module_id": module.id, "phase": phase, "week_no": week_no},
                daemon=True,
            ).start()

    return JsonResponse({"ok": True})


def attendance_analytics(request):
    mentor = _session_mentor_obj(request)
    if not mentor and not request.user.is_authenticated:
        return redirect("/")

    module = _active_module(request)
    calendar = _calendar_for_module(module)
    phase = (request.GET.get("phase") or "T1").upper()
    week_param = (request.GET.get("week") or "all").strip().lower()
    week_no = None if week_param in {"all", ""} else int(week_param) + 1
    batch_filter = (request.GET.get("batch") or "").strip()
    batch_filter = (request.GET.get("batch") or "").strip()
    search = (request.GET.get("search") or "").strip()
    batch_choices = sorted(
        {
            ((student.division or "").strip() or (student.batch or "").strip())
            for student in Student.objects.filter(module=module).only("batch", "division")
            if ((student.division or "").strip() or (student.batch or "").strip())
        }
    )

    start, end = phase_range(calendar, phase)
    if not start or not end:
        return render(
            request,
            "attendance_analytics.html",
            {
                "module": module,
                "calendar": calendar,
                "phase": phase,
                "week_param": week_param,
                "batch_filter": batch_filter,
                "search": search,
                "rows": [],
                "week_list": [],
                "subject_list": [],
                "batch_choices": batch_choices,
                "message": "Set the academic calendar first.",
            },
        )

    end_date = end_date_for_week(calendar, phase, week_no)
    if end_date and end_date > end:
        end_date = end
    today = timezone.localdate()
    range_end = min(end_date or end, today)
    if range_end < start:
        return render(
            request,
            "attendance_analytics.html",
            {
                "module": module,
                "calendar": calendar,
                "phase": phase,
                "week_param": week_param,
                "batch_filter": batch_filter,
                "search": search,
                "rows": [],
                "week_list": [],
                "subject_list": [],
                "batch_choices": batch_choices,
                "message": "Attendance analytics are available only up to today.",
            },
        )

    batch_filter_key = _norm_batch_key(batch_filter)

    sessions_qs = LectureSession.objects.filter(
        module=module,
        date__gte=start,
        date__lte=range_end,
    )
    sessions = [
        s
        for s in sessions_qs
        if _attendance_allowed_for_date(module, s.date)
        and (not batch_filter_key or _norm_batch_key(s.batch) == batch_filter_key)
    ]
    alias_map = _subject_alias_map(module)

    session_ids = [s.id for s in sessions]
    absences = LectureAbsence.objects.filter(session_id__in=session_ids).select_related("session", "student")

    students_qs = Student.objects.filter(module=module)
    if search:
        search_q = Q(name__icontains=search) | Q(enrollment__icontains=search)
        if search.isdigit():
            search_q |= Q(roll_no=int(search))
        students_qs = students_qs.filter(search_q)
    students = [
        student
        for student in students_qs.order_by("roll_no", "name")
        if not batch_filter_key or batch_filter_key in _student_batch_keys(student)
    ]

    held_by_batch = {}
    held_by_batch_subject = {}
    held_by_batch_week = {}
    subject_set = set()
    week_set = set()

    for s in sessions:
        batch_key = _norm_batch_key(s.batch)
        held_by_batch.setdefault(batch_key, 0)
        held_by_batch[batch_key] += 1
        subj = _canonical_subject_name(module, s.subject, alias_map)
        subject_set.add(subj)
        held_by_batch_subject.setdefault(batch_key, {}).setdefault(subj, 0)
        held_by_batch_subject[batch_key][subj] += 1
        _, week_idx = week_for_date(calendar, s.date)
        if week_idx:
            week_set.add(week_idx)
            held_by_batch_week.setdefault(batch_key, {}).setdefault(week_idx, 0)
            held_by_batch_week[batch_key][week_idx] += 1

    absent_by_student = {}
    absent_by_student_subject = {}
    absent_by_student_week = {}
    for a in absences:
        sid = a.student_id
        absent_by_student[sid] = absent_by_student.get(sid, 0) + 1
        subj = _canonical_subject_name(module, a.session.subject, alias_map)
        absent_by_student_subject.setdefault(sid, {}).setdefault(subj, 0)
        absent_by_student_subject[sid][subj] += 1
        _, week_idx = week_for_date(calendar, a.session.date)
        if week_idx:
            absent_by_student_week.setdefault(sid, {}).setdefault(week_idx, 0)
            absent_by_student_week[sid][week_idx] += 1

    subject_list = [s for s in sorted(subject_set) if s]
    week_list = sorted(week_set)

    rows = []
    for student in students:
        batch_label = (student.division or student.batch or "").strip()
        batch_keys = _student_batch_keys(student)
        held_total = sum(held_by_batch.get(key, 0) for key in batch_keys)
        absent_total = absent_by_student.get(student.id, 0)
        attended_total = max(held_total - absent_total, 0)
        percent = round((attended_total / held_total) * 100, 1) if held_total else 0

        week_rows = []
        for wk in week_list:
            held_w = sum(held_by_batch_week.get(key, {}).get(wk, 0) for key in batch_keys)
            absent_w = absent_by_student_week.get(student.id, {}).get(wk, 0)
            attended_w = max(held_w - absent_w, 0)
            pct_w = round((attended_w / held_w) * 100, 1) if held_w else 0
            week_rows.append(
                {
                    "week": wk,
                    "held": held_w,
                    "attended": attended_w,
                    "percent": pct_w,
                }
            )

        subject_rows = []
        for subj in subject_list:
            held_s = sum(held_by_batch_subject.get(key, {}).get(subj, 0) for key in batch_keys)
            absent_s = absent_by_student_subject.get(student.id, {}).get(subj, 0)
            attended_s = max(held_s - absent_s, 0)
            pct_s = round((attended_s / held_s) * 100, 1) if held_s else 0
            subject_rows.append(
                {
                    "subject": subj,
                    "held": held_s,
                    "attended": attended_s,
                    "percent": pct_s,
                }
            )

        rows.append(
            {
                "student": student,
                "batch_label": batch_label,
                "held": held_total,
                "attended": attended_total,
                "percent": percent,
                "week_rows": week_rows,
                "subject_rows": subject_rows,
            }
        )

    return render(
        request,
        "attendance_analytics.html",
        {
            "module": module,
            "calendar": calendar,
            "phase": phase,
            "week_param": week_param,
            "batch_filter": batch_filter,
            "search": search,
            "rows": rows,
            "week_list": week_list,
            "subject_list": subject_list,
            "batch_choices": batch_choices,
        },
    )


def daily_absent_excel(request):
    mentor = _session_mentor_obj(request)
    if not mentor and not request.user.is_authenticated:
        return redirect("/")

    is_coordinator = bool(request.user.is_authenticated and not request.session.get("mentor") and not is_superadmin_user(request.user))
    module = None
    module_id = (request.GET.get("module_id") or "").strip()
    if mentor and module_id:
        module = allowed_modules_for_user(request).filter(id=module_id, is_active=True).first()
        if not module:
            return HttpResponse("Unauthorized", status=403)
    if not module:
        module = _active_module(request)
    date_val = _parse_date_param(request.GET.get("date"), timezone.localdate())
    batch_filter = (request.GET.get("batch") or "").strip()

    if not _attendance_allowed_for_date(module, date_val):
        return HttpResponse("Attendance is not allowed for this date.", status=400)

    if is_coordinator and not _attendance_fully_marked_for_date(module, date_val):
        return HttpResponse("Attendance is not fully marked for this date.", status=400)

    sessions_qs = LectureSession.objects.filter(module=module, date=date_val)
    if batch_filter:
        sessions_qs = sessions_qs.filter(batch=batch_filter)
    sessions = list(sessions_qs.order_by("batch", "lecture_no"))
    if not sessions:
        return HttpResponse("No attendance sessions found for selected date.", status=404)
    alias_map = _subject_alias_map(module)

    absences = LectureAbsence.objects.filter(session__in=sessions).select_related("student", "session")
    absences_by_session = {}
    for a in absences:
        absences_by_session.setdefault(a.session_id, []).append(a)

    batches = sorted({s.batch for s in sessions})
    sessions_by_batch = {}
    for s in sessions:
        sessions_by_batch.setdefault(s.batch, []).append(s)

    wb = Workbook()
    ws = wb.active
    day_str = date_val.strftime("%d-%b-%Y")
    ws.title = day_str[:31]
    title = f"L J Institute of Engineering and Technology\n{module.name} Daily Absent No.\n{day_str} ({date_val.strftime('%A')})"
    ws.cell(row=1, column=1, value=title)

    row_cursor = 2
    for i in range(0, len(batches), 2):
        left = batches[i]
        right = batches[i + 1] if i + 1 < len(batches) else None

        def batch_label(code):
            raw = code.strip()
            if raw.upper().startswith("B-"):
                num = raw.split("-", 1)[1]
                return f"Batch-{num} ({raw})"
            return f"Batch ({raw})"

        ws.cell(row=row_cursor, column=1, value=batch_label(left))
        if right:
            ws.cell(row=row_cursor, column=7, value=batch_label(right))
        row_cursor += 1

        headers = ["Lec No.", "Total no of absent", "Subject", "Faculty", "Absent Nos."]
        for idx, label in enumerate(headers):
            ws.cell(row=row_cursor, column=1 + idx, value=label)
            if right:
                ws.cell(row=row_cursor, column=7 + idx, value=label)
        row_cursor += 1

        left_sessions = sessions_by_batch.get(left, [])
        right_sessions = sessions_by_batch.get(right, []) if right else []
        max_len = max(len(left_sessions), len(right_sessions))

        for j in range(max_len):
            if j < len(left_sessions):
                s = left_sessions[j]
                abs_list = absences_by_session.get(s.id, [])
                rolls = sorted([a.student.roll_no for a in abs_list if a.student.roll_no is not None])
                absent_str = ", ".join(str(x) for x in rolls) if rolls else "NIL"
                ws.cell(row=row_cursor, column=1, value=s.lecture_no)
                ws.cell(row=row_cursor, column=2, value=len(rolls))
                ws.cell(row=row_cursor, column=3, value=_canonical_subject_name(module, s.subject, alias_map))
                ws.cell(row=row_cursor, column=4, value=s.faculty)
                ws.cell(row=row_cursor, column=5, value=absent_str)
            if right and j < len(right_sessions):
                s = right_sessions[j]
                abs_list = absences_by_session.get(s.id, [])
                rolls = sorted([a.student.roll_no for a in abs_list if a.student.roll_no is not None])
                absent_str = ", ".join(str(x) for x in rolls) if rolls else "NIL"
                ws.cell(row=row_cursor, column=7, value=s.lecture_no)
                ws.cell(row=row_cursor, column=8, value=len(rolls))
                ws.cell(row=row_cursor, column=9, value=_canonical_subject_name(module, s.subject, alias_map))
                ws.cell(row=row_cursor, column=10, value=s.faculty)
                ws.cell(row=row_cursor, column=11, value=absent_str)
            row_cursor += 1

        row_cursor += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"Daily_Absent_{date_val:%Y-%m-%d}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _daily_absent_cards(module, date_val, batch_filter=""):
    sessions_qs = LectureSession.objects.filter(module=module, date=date_val)
    if batch_filter:
        sessions_qs = sessions_qs.filter(batch=batch_filter)
    sessions = list(sessions_qs.order_by("batch", "lecture_no", "time_slot"))
    if not sessions:
        return [], []
    alias_map = _subject_alias_map(module)

    absences = LectureAbsence.objects.filter(session__in=sessions).select_related("student", "session")
    absences_by_session = {}
    for a in absences:
        absences_by_session.setdefault(a.session_id, []).append(a)

    batches = sorted({s.batch for s in sessions})
    totals = {}
    for student in Student.objects.filter(module=module):
        for key in {_norm_batch_key(student.batch), _norm_batch_key(student.division)}:
            if key:
                totals[key] = totals.get(key, 0) + 1

    sessions_by_batch = {}
    for s in sessions:
        sessions_by_batch.setdefault(s.batch, []).append(s)

    batch_cards = []
    for batch in batches:
        batch_sessions = sessions_by_batch.get(batch, [])
        total_lectures = len(batch_sessions)
        absent_count_by_roll = {}
        for sess in batch_sessions:
            for absence in absences_by_session.get(sess.id, []):
                if absence.student and absence.student.roll_no is not None:
                    absent_count_by_roll[absence.student.roll_no] = absent_count_by_roll.get(absence.student.roll_no, 0) + 1
        lectures = []
        for sess in batch_sessions:
            abs_list = absences_by_session.get(sess.id, [])
            rolls = sorted([a.student.roll_no for a in abs_list if a.student.roll_no is not None])
            roll_items = [
                {
                    "roll_no": roll_no,
                    "is_partial": absent_count_by_roll.get(roll_no, 0) < total_lectures,
                }
                for roll_no in rolls
            ]
            total_students = totals.get(_norm_batch_key(batch), 0)
            absent_count = len(rolls)
            present_count = max(total_students - absent_count, 0) if total_students else 0
            lectures.append(
                {
                    "lecture_no": sess.lecture_no,
                    "time_slot": sess.time_slot,
                    "subject": _canonical_subject_name(module, sess.subject, alias_map),
                    "faculty": sess.faculty,
                    "absent_rolls": rolls,
                    "absent_roll_items": roll_items,
                    "absent_text": ", ".join(str(x) for x in rolls) if rolls else "NIL",
                    "absent_count": absent_count,
                    "present_count": present_count,
                }
            )
        lectures.sort(key=lambda r: (r["lecture_no"], r["time_slot"]))
        batch_cards.append(
            {
                "batch": batch,
                "total_students": totals.get(_norm_batch_key(batch), 0),
                "lectures": lectures,
            }
        )
    return batch_cards, batches


def _daily_absent_cards_for_pdf(module, date_val, batch_filter=""):
    day_of_week = date_val.weekday()
    alias_map = _subject_alias_map(module)
    timetable_qs = TimetableEntry.objects.filter(module=module, day_of_week=day_of_week, is_active=True)
    if batch_filter:
        timetable_qs = timetable_qs.filter(batch=batch_filter)
    timetable_entries = list(timetable_qs.order_by("batch", "lecture_no", "time_slot"))
    if not timetable_entries:
        return [], []

    sessions = list(
        LectureSession.objects.filter(
            module=module,
            date=date_val,
            batch__in=[entry.batch for entry in timetable_entries],
        ).order_by("batch", "lecture_no", "time_slot")
    )
    session_map = {(session.batch, session.lecture_no): session for session in sessions}
    absences = LectureAbsence.objects.filter(session__in=sessions).select_related("student", "session")
    absences_by_session = {}
    for absence in absences:
        absences_by_session.setdefault(absence.session_id, []).append(absence)

    totals = {}
    for student in Student.objects.filter(module=module):
        for key in {_norm_batch_key(student.batch), _norm_batch_key(student.division)}:
            if key:
                totals[key] = totals.get(key, 0) + 1

    entries_by_batch = {}
    for entry in timetable_entries:
        entries_by_batch.setdefault(entry.batch, []).append(entry)

    batch_cards = []
    for batch in sorted(entries_by_batch.keys()):
        batch_entries = entries_by_batch.get(batch, [])
        total_lectures = len(batch_entries)
        absent_count_by_roll = {}
        for entry in batch_entries:
            session = session_map.get((batch, entry.lecture_no))
            if not session:
                continue
            for absence in absences_by_session.get(session.id, []):
                if absence.student and absence.student.roll_no is not None:
                    absent_count_by_roll[absence.student.roll_no] = absent_count_by_roll.get(absence.student.roll_no, 0) + 1

        lectures = []
        for entry in batch_entries:
            session = session_map.get((batch, entry.lecture_no))
            abs_list = absences_by_session.get(session.id, []) if session else []
            rolls = sorted([a.student.roll_no for a in abs_list if a.student.roll_no is not None])
            roll_items = [
                {
                    "roll_no": roll_no,
                    "is_partial": absent_count_by_roll.get(roll_no, 0) < total_lectures,
                }
                for roll_no in rolls
            ]
            total_students = totals.get(_norm_batch_key(batch), 0)
            absent_count = len(rolls) if session else None
            present_count = (max(total_students - len(rolls), 0) if total_students else 0) if session else None
            lectures.append(
                {
                    "lecture_no": entry.lecture_no,
                    "time_slot": entry.time_slot,
                    "subject": _canonical_subject_name(module, entry.subject, alias_map),
                    "faculty": (session.faculty if session and session.faculty else entry.faculty),
                    "absent_rolls": rolls,
                    "absent_roll_items": roll_items,
                    "absent_text": ", ".join(str(x) for x in rolls) if rolls else "",
                    "absent_count": absent_count,
                    "present_count": present_count,
                    "is_filled": bool(session),
                }
            )
        batch_cards.append(
            {
                "batch": batch,
                "total_students": totals.get(_norm_batch_key(batch), 0),
                "lectures": lectures,
            }
        )
    return batch_cards, sorted(entries_by_batch.keys())


def daily_absent_live(request):
    mentor = _session_mentor_obj(request)
    if not mentor and not request.user.is_authenticated:
        return redirect("/")

    is_coordinator = bool(request.user.is_authenticated and not request.session.get("mentor") and not is_superadmin_user(request.user))
    module = None
    module_id = (request.GET.get("module_id") or "").strip()
    if mentor and module_id:
        module = allowed_modules_for_user(request).filter(id=module_id, is_active=True).first()
        if not module:
            return HttpResponse("Unauthorized", status=403)
    if not module:
        module = _active_module(request)

    date_val = _parse_date_param(request.GET.get("date"), timezone.localdate())
    batch_filter = (request.GET.get("batch") or "").strip()

    if not _attendance_allowed_for_date(module, date_val):
        return HttpResponse("Attendance is not allowed for this date.", status=400)

    if is_coordinator and not _attendance_fully_marked_for_date(module, date_val):
        return HttpResponse("Attendance is not fully marked for this date.", status=400)

    batch_cards, batches = _daily_absent_cards(module, date_val, batch_filter)
    return render(
        request,
        "daily_absent_live.html",
        {
            "module": module,
            "selected_date": date_val,
            "batch_filter": batch_filter,
            "batch_cards": batch_cards,
            "batches": batches,
        },
    )


def daily_absent_live_pdf(request):
    mentor = _session_mentor_obj(request)
    if not mentor and not request.user.is_authenticated:
        return redirect("/")

    is_coordinator = bool(
        request.user.is_authenticated and not request.session.get("mentor") and not is_superadmin_user(request.user)
    )
    module = None
    module_id = (request.GET.get("module_id") or "").strip()
    if mentor and module_id:
        module = allowed_modules_for_user(request).filter(id=module_id, is_active=True).first()
        if not module:
            return HttpResponse("Unauthorized", status=403)
    if not module:
        module = _active_module(request)

    date_val = _parse_date_param(request.GET.get("date"), timezone.localdate())
    batch_filter = (request.GET.get("batch") or "").strip()

    if not _attendance_allowed_for_date(module, date_val):
        return HttpResponse("Attendance is not allowed for this date.", status=400)

    if is_coordinator and not _attendance_fully_marked_for_date(module, date_val):
        return HttpResponse("Attendance is not fully marked for this date.", status=400)

    batch_cards, _ = _daily_absent_cards_for_pdf(module, date_val, batch_filter)
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="Daily_Absent_{date_val:%Y-%m-%d}.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=24,
        rightMargin=24,
        topMargin=18,
        bottomMargin=18,
    )
    styles = getSampleStyleSheet()
    absent_style = styles["BodyText"]
    absent_style.fontSize = 8.5
    absent_style.leading = 10
    title_text = (
        "L J Institute of Engineering and Technology<br/>"
        f"{module.name} Lecture Daily Absent No.<br/>"
        f"{date_val.strftime('%d-%B-%Y (%A)')}"
    )
    title_style = styles["Heading4"]
    title_style.alignment = 1
    title_style.textColor = colors.black
    title_style.leading = 14
    title_table = Table(
        [[Paragraph(title_text, title_style)]],
        colWidths=[doc.width],
    )
    title_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#9FE3E8")),
                ("BOX", (0, 0), (-1, -1), 1.5, colors.black),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story = [title_table, Spacer(1, 10)]

    for card in batch_cards:
        story.append(Paragraph(f"Batch: {card['batch']}", styles["Heading5"]))
        table_data = [
            ["Lec", "Time", "Subject", "Faculty", "Absent Nos.", "Absent", "Present"]
        ]
        for row in card["lectures"]:
            if row["is_filled"] and row["absent_roll_items"]:
                absent_parts = []
                for item in row["absent_roll_items"]:
                    roll_text = str(item["roll_no"])
                    if item.get("is_partial"):
                        absent_parts.append(f"<font color='#dc3545'><b>{roll_text}</b></font>")
                    else:
                        absent_parts.append(roll_text)
                absent_text = Paragraph(", ".join(absent_parts), absent_style)
            elif row["is_filled"]:
                absent_text = "NIL"
            else:
                absent_text = ""
            table_data.append(
                [
                    row["lecture_no"],
                    row["time_slot"],
                    row["subject"],
                    row["faculty"],
                    absent_text,
                    (row["absent_count"] if row["absent_count"] is not None else ""),
                    (row["present_count"] if row["present_count"] is not None else ""),
                ]
            )
        table = Table(
            table_data,
            repeatRows=1,
            colWidths=[20, 54, 60, 40, 233, 56, 60],
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f3f5")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd3da")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                    ("LEFTPADDING", (0, 0), (-1, -1), 2),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 8))

    doc.build(story)
    return response


def weekly_attendance_live(request):
    mentor = _session_mentor_obj(request)
    if not mentor and not request.user.is_authenticated:
        return redirect("/")

    is_coordinator = bool(request.user.is_authenticated and not request.session.get("mentor") and not is_superadmin_user(request.user))
    module = None
    module_id = (request.GET.get("module_id") or "").strip()
    if mentor and module_id:
        module = allowed_modules_for_user(request).filter(id=module_id, is_active=True).first()
        if not module:
            return HttpResponse("Unauthorized", status=403)
    if not module:
        module = _active_module(request)

    calendar = _calendar_for_module(module)
    phase = (request.GET.get("phase") or "T1").upper()
    week_param = (request.GET.get("week") or "all").strip().lower()
    week_no = None if week_param in {"all", ""} else int(week_param) + 1

    start, end = phase_range(calendar, phase)
    if not start or not end:
        return render(
            request,
            "weekly_attendance_live.html",
            {
                "module": module,
                "calendar": calendar,
                "phase": phase,
                "week_param": week_param,
                "subject_list": [],
                "rows": [],
                "message": "Academic calendar not configured.",
            },
        )

    end_date = end_date_for_week(calendar, phase, week_no)
    if end_date and end_date > end:
        end_date = end
    today = timezone.localdate()
    range_end = min(end_date or end, today)
    if range_end < start:
        return render(
            request,
            "weekly_attendance_live.html",
            {
                "module": module,
                "calendar": calendar,
                "phase": phase,
                "week_param": week_param,
                "subject_list": [],
                "rows": [],
                "message": "Weekly attendance is available only up to today.",
            },
        )

    if is_coordinator and not _attendance_fully_marked_for_range(module, start, range_end):
        return HttpResponse("Attendance is not fully marked for the selected range.", status=400)

    sessions = list(
        LectureSession.objects.filter(module=module, date__gte=start, date__lte=range_end)
    )
    sessions = [s for s in sessions if _attendance_allowed_for_date(module, s.date)]
    session_ids = [s.id for s in sessions]
    absences = LectureAbsence.objects.filter(session_id__in=session_ids).select_related("session", "student")

    alias_map = _subject_alias_map(module)
    subject_list = sorted(
        {_canonical_subject_name(module, s.subject, alias_map) for s in sessions if (s.subject or "").strip()}
    )
    week_set = set()
    for s in sessions:
        _, week_idx = week_for_date(calendar, s.date)
        if week_idx:
            week_set.add(week_idx)
    week_list = sorted(week_set)

    held_by_batch_subject = {}
    held_by_batch = {}
    for s in sessions:
        batch_key = _norm_batch_key(s.batch)
        held_by_batch.setdefault(batch_key, 0)
        held_by_batch[batch_key] += 1
        subj = _canonical_subject_name(module, s.subject, alias_map)
        held_by_batch_subject.setdefault(batch_key, {}).setdefault(subj, 0)
        held_by_batch_subject[batch_key][subj] += 1

    absent_by_student_subject = {}
    absent_by_student = {}
    for a in absences:
        sid = a.student_id
        absent_by_student[sid] = absent_by_student.get(sid, 0) + 1
        subj = _canonical_subject_name(module, a.session.subject, alias_map)
        absent_by_student_subject.setdefault(sid, {}).setdefault(subj, 0)
        absent_by_student_subject[sid][subj] += 1

    students = list(Student.objects.filter(module=module).order_by("roll_no", "name"))

    rows = []
    for student in students:
        batch_keys = _student_batch_keys(student)
        held_total = sum(held_by_batch.get(key, 0) for key in batch_keys)
        absent_total = absent_by_student.get(student.id, 0)
        attended_total = max(held_total - absent_total, 0)
        overall_pct = round((attended_total / held_total) * 100, 2) if held_total else 0

        subject_rows = []
        for subj in subject_list:
            held = sum(held_by_batch_subject.get(key, {}).get(subj, 0) for key in batch_keys)
            absent = absent_by_student_subject.get(student.id, {}).get(subj, 0)
            attended = max(held - absent, 0)
            pct = round((attended / held) * 100, 2) if held else 0
            subject_rows.append({"attended": attended, "held": held, "percent": pct})

        rows.append(
            {
                "student": student,
                "subjects": subject_rows,
                "attended_total": attended_total,
                "held_total": held_total,
                "overall_pct": overall_pct,
            }
        )

    return render(
        request,
        "weekly_attendance_live.html",
        {
            "module": module,
            "calendar": calendar,
            "phase": phase,
            "week_param": week_param,
            "week_list": week_list,
            "subject_list": subject_list,
            "rows": rows,
            "range_start": start,
            "range_end": range_end,
        },
    )


def _safe_sheet_title(title):
    cleaned = re.sub(r"[:\\/?*\[\]]", "_", str(title or "").strip())
    return (cleaned or "Sheet")[:31]


def _highlight_low_percent(cell, ratio_value):
    if ratio_value is None:
        return
    try:
        numeric = float(ratio_value)
    except Exception:
        return
    if numeric < 0.8:
        cell.fill = PatternFill(fill_type="solid", fgColor="F8CBAD")
        cell.font = Font(color="9C0006", bold=True)


def _student_batch_keys(student):
    return {
        key
        for key in {
            _norm_batch_key(student.batch),
            _norm_batch_key(student.division),
        }
        if key
    }


def _ordered_subject_names_for_module(module, sessions):
    alias_map = _subject_alias_map(module)
    session_subjects = {
        _canonical_subject_name(module, session.subject, alias_map)
        for session in sessions
        if (session.subject or "").strip()
    }
    ordered = []
    seen = set()
    for subject in Subject.objects.filter(module=module, is_active=True).order_by("display_order", "name"):
        name = (subject.name or "").strip()
        canonical = _canonical_subject_name(module, name, alias_map)
        if canonical and canonical in session_subjects and canonical not in seen:
            ordered.append(canonical)
            seen.add(canonical)
        short = (subject.short_name or "").strip()
        canonical_short = _canonical_subject_name(module, short, alias_map) if short else ""
        if canonical_short and canonical_short in session_subjects and canonical_short not in seen:
            ordered.append(canonical_short)
            seen.add(canonical_short)
    for name in sorted(session_subjects):
        if name and name not in seen:
            ordered.append(name)
            seen.add(name)
    return ordered


def _weekly_export_data(module, calendar, phase, week_no, batch_filter=""):
    start, end = phase_range(calendar, phase)
    if not start or not end:
        return None

    end_date = end_date_for_week(calendar, phase, week_no)
    if end_date and end_date > end:
        end_date = end
    range_end = min(end_date or end, timezone.localdate())
    if range_end < start:
        return None

    batch_filter_key = _norm_batch_key(batch_filter)

    students = list(
        Student.objects.filter(module=module)
        .select_related("mentor")
        .order_by("batch", "roll_no", "name")
    )
    if batch_filter_key:
        students = [student for student in students if batch_filter_key in _student_batch_keys(student)]
    student_keys = {student.id: _student_batch_keys(student) for student in students}
    students_by_batch = {}
    for student in students:
        batch_label = (student.batch or student.division or "").strip() or "Unassigned"
        students_by_batch.setdefault(batch_label, []).append(student)

    sessions = list(
        LectureSession.objects.filter(module=module, date__gte=start, date__lte=range_end)
        .order_by("date", "lecture_no", "batch", "time_slot")
    )
    sessions = [session for session in sessions if _attendance_allowed_for_date(module, session.date)]
    alias_map = _subject_alias_map(module)
    if batch_filter_key:
        sessions = [session for session in sessions if _norm_batch_key(session.batch) == batch_filter_key]
    session_ids = [session.id for session in sessions]
    absences = list(
        LectureAbsence.objects.filter(session_id__in=session_ids)
        .select_related("student", "session")
    )
    absences_by_session = {}
    for absence in absences:
        absences_by_session.setdefault(absence.session_id, set()).add(absence.student_id)

    per_student = {
        student.id: {
            "held_total": 0,
            "attended_total": 0,
            "subject": {},
            "week": {},
            "session": {},
        }
        for student in students
    }
    session_entries = []
    batch_session_map = {}
    week_numbers = set()

    for session in sessions:
        batch_key = _norm_batch_key(session.batch)
        matched_students = []
        for student in students:
            if batch_key and batch_key in student_keys.get(student.id, set()):
                matched_students.append(student)
        session_absent_ids = absences_by_session.get(session.id, set())
        _, session_week_no = week_for_date(calendar, session.date)
        if session_week_no:
            week_numbers.add(session_week_no)
        subject_name = _canonical_subject_name(module, session.subject, alias_map)
        batch_session_map.setdefault(session.batch, []).append(session)

        session_entry = {
            "session": session,
            "student_ids": [student.id for student in matched_students],
            "subject": subject_name,
            "week_no": session_week_no,
        }
        session_entries.append(session_entry)

        for student in matched_students:
            stats = per_student[student.id]
            stats["held_total"] += 1
            attended = student.id not in session_absent_ids
            if attended:
                stats["attended_total"] += 1
            if subject_name:
                subj_stats = stats["subject"].setdefault(subject_name, {"held": 0, "attended": 0})
                subj_stats["held"] += 1
                if attended:
                    subj_stats["attended"] += 1
            if session_week_no:
                wk_stats = stats["week"].setdefault(session_week_no, {"held": 0, "attended": 0})
                wk_stats["held"] += 1
                if attended:
                    wk_stats["attended"] += 1
            stats["session"][session.id] = attended

    ordered_subjects = _ordered_subject_names_for_module(module, sessions)
    ordered_batches = sorted(students_by_batch.keys())
    ordered_weeks = sorted(week_numbers)

    return {
        "start": start,
        "end": range_end,
        "students": students,
        "students_by_batch": students_by_batch,
        "ordered_batches": ordered_batches,
        "sessions": sessions,
        "batch_session_map": batch_session_map,
        "session_entries": session_entries,
        "subjects": ordered_subjects,
        "weeks": ordered_weeks,
        "per_student": per_student,
        "alias_map": alias_map,
        "module": module,
    }


def _autosize_sheet(ws, min_width=10, max_width=28):
    for column_cells in ws.columns:
        lengths = [len(str(cell.value or "")) for cell in column_cells]
        width = max(lengths or [min_width]) + 2
        letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[letter].width = max(min_width, min(width, max_width))


def _write_compiled_sheet(ws, module, phase, export_data):
    ws.title = "Compiled"
    ws.cell(row=1, column=1, value="L.J. Institute of Engineering and Technology")
    ws.cell(row=2, column=1, value=module.name)
    ws.cell(row=3, column=1, value=f"Weekly Attendance - {phase}")
    headers = ["Roll No", "Batch", "Division", "Student Name", "Enrollment No", "Mentor"]
    row_no = 6
    col_no = 1
    for label in headers:
        ws.cell(row=row_no, column=col_no, value=label)
        ws.cell(row=row_no, column=col_no).font = Font(bold=True)
        col_no += 1
    for subject_name in export_data["subjects"]:
        ws.cell(row=row_no, column=col_no, value=subject_name)
        ws.cell(row=row_no + 1, column=col_no, value="Att")
        ws.cell(row=row_no + 1, column=col_no + 1, value="Held")
        ws.cell(row=row_no + 1, column=col_no + 2, value="%")
        col_no += 3
    ws.cell(row=row_no, column=col_no, value="Overall")
    ws.cell(row=row_no + 1, column=col_no, value="Att")
    ws.cell(row=row_no + 1, column=col_no + 1, value="Held")
    ws.cell(row=row_no + 1, column=col_no + 2, value="%")

    data_row = 8
    for student in export_data["students"]:
        stats = export_data["per_student"][student.id]
        row = [
            student.roll_no,
            student.batch or "",
            student.division or "",
            student.name,
            student.enrollment,
            student.mentor.name if student.mentor else "",
        ]
        for idx, value in enumerate(row, start=1):
            ws.cell(row=data_row, column=idx, value=value)
        col_no = 7
        for subject_name in export_data["subjects"]:
            subject_stats = stats["subject"].get(subject_name, {"held": 0, "attended": 0})
            held = subject_stats["held"]
            attended = subject_stats["attended"]
            ratio = (attended / held) if held else None
            ws.cell(row=data_row, column=col_no, value=attended)
            ws.cell(row=data_row, column=col_no + 1, value=held)
            pct_cell = ws.cell(row=data_row, column=col_no + 2, value=ratio if ratio is not None else "")
            if ratio is not None:
                pct_cell.number_format = "0.0%"
                _highlight_low_percent(pct_cell, ratio)
            col_no += 3
        held_total = stats["held_total"]
        attended_total = stats["attended_total"]
        overall_ratio = (attended_total / held_total) if held_total else None
        ws.cell(row=data_row, column=col_no, value=attended_total)
        ws.cell(row=data_row, column=col_no + 1, value=held_total)
        pct_cell = ws.cell(row=data_row, column=col_no + 2, value=overall_ratio if overall_ratio is not None else "")
        if overall_ratio is not None:
            pct_cell.number_format = "0.0%"
            _highlight_low_percent(pct_cell, overall_ratio)
        data_row += 1
    _autosize_sheet(ws)


def _write_batchwise_sheets(workbook, export_data):
    for batch in export_data["ordered_batches"]:
        ws = workbook.create_sheet(title=_safe_sheet_title(batch))
        ws.append(["Roll No", "Student Name", "Enrollment No", "Mentor", "Attended", "Held", "%"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for student in export_data["students_by_batch"].get(batch, []):
            stats = export_data["per_student"][student.id]
            held_total = stats["held_total"]
            attended_total = stats["attended_total"]
            ratio = (attended_total / held_total) if held_total else None
            ws.append([
                student.roll_no,
                student.name,
                student.enrollment,
                student.mentor.name if student.mentor else "",
                attended_total,
                held_total,
                ratio if ratio is not None else "",
            ])
            if ratio is not None:
                pct_cell = ws.cell(row=ws.max_row, column=7)
                pct_cell.number_format = "0.0%"
                _highlight_low_percent(pct_cell, ratio)
        _autosize_sheet(ws)


def _write_subjectwise_sheets(workbook, export_data):
    for subject_name in export_data["subjects"]:
        ws = workbook.create_sheet(title=_safe_sheet_title(subject_name))
        ws.append(["Roll No", "Student Name", "Batch", "Enrollment No", "Attended", "Held", "%", "Overall %"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for student in export_data["students"]:
            stats = export_data["per_student"][student.id]
            subject_stats = stats["subject"].get(subject_name, {"held": 0, "attended": 0})
            held = subject_stats["held"]
            attended = subject_stats["attended"]
            ratio = (attended / held) if held else None
            overall_ratio = (stats["attended_total"] / stats["held_total"]) if stats["held_total"] else None
            ws.append([
                student.roll_no,
                student.name,
                student.batch or student.division or "",
                student.enrollment,
                attended,
                held,
                ratio if ratio is not None else "",
                overall_ratio if overall_ratio is not None else "",
            ])
            if ratio is not None:
                pct_cell = ws.cell(row=ws.max_row, column=7)
                pct_cell.number_format = "0.0%"
                _highlight_low_percent(pct_cell, ratio)
            if overall_ratio is not None:
                pct_cell = ws.cell(row=ws.max_row, column=8)
                pct_cell.number_format = "0.0%"
                _highlight_low_percent(pct_cell, overall_ratio)
        _autosize_sheet(ws)


def _write_register_sheets(workbook, export_data):
    module = export_data.get("module")
    alias_map = export_data.get("alias_map") or ( _subject_alias_map(module) if module else {} )
    for batch in export_data["ordered_batches"]:
        sessions = [
            session
            for session in export_data["batch_session_map"].get(batch, [])
            if _norm_batch_key(session.batch) == _norm_batch_key(batch)
        ]
        ws = workbook.create_sheet(title=_safe_sheet_title(f"{batch} Register"))
        headers = ["Roll No", "Student Name", "Enrollment No"]
        session_labels = [
            f"{session.date:%d-%b} L{session.lecture_no} {(_canonical_subject_name(module, session.subject, alias_map) or 'Lecture')}"
            for session in sessions
        ]
        ws.append(headers + session_labels + ["Attended", "Held", "%"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for student in export_data["students_by_batch"].get(batch, []):
            stats = export_data["per_student"][student.id]
            row = [student.roll_no, student.name, student.enrollment]
            attended_count = 0
            held_count = 0
            for session in sessions:
                attended = stats["session"].get(session.id)
                if attended is None:
                    row.append("")
                    continue
                held_count += 1
                if attended:
                    attended_count += 1
                    row.append("P")
                else:
                    row.append("A")
            ratio = (attended_count / held_count) if held_count else None
            row.extend([attended_count, held_count, ratio if ratio is not None else ""])
            ws.append(row)
            if ratio is not None:
                pct_cell = ws.cell(row=ws.max_row, column=len(row))
                pct_cell.number_format = "0.0%"
                _highlight_low_percent(pct_cell, ratio)
        _autosize_sheet(ws, min_width=8, max_width=18)


def weekly_attendance_excel(request):
    if request.session.get("mentor"):
        return HttpResponse("Forbidden", status=403)
    if not request.user.is_authenticated:
        return redirect("/")

    module = _active_module(request)
    is_coordinator = bool(request.user.is_authenticated and not request.session.get("mentor") and not is_superadmin_user(request.user))
    calendar = _calendar_for_module(module)
    phase = (request.GET.get("phase") or "T1").upper()
    week_param = (request.GET.get("week") or "all").strip().lower()
    week_no = None if week_param in {"all", ""} else int(week_param) + 1

    start, end = phase_range(calendar, phase)
    if not start or not end:
        return HttpResponse("Academic calendar not configured.", status=400)

    end_date = end_date_for_week(calendar, phase, week_no)
    if end_date and end_date > end:
        end_date = end
    range_end = min(end_date or end, timezone.localdate())
    if range_end < start:
        return HttpResponse("Weekly attendance is available only up to today.", status=400)

    if is_coordinator and not _attendance_fully_marked_for_range(module, start, range_end):
        return HttpResponse("Attendance is not fully marked for the selected range.", status=400)

    export_format = (request.GET.get("format") or "compiled").strip().lower()
    export_data = _weekly_export_data(module, calendar, phase, week_no, batch_filter=batch_filter)
    if export_data is None:
        return HttpResponse("Academic calendar not configured.", status=400)

    wb = Workbook()
    primary_ws = wb.active
    _write_compiled_sheet(primary_ws, module, phase, export_data)
    if export_format == "batchwise":
        _write_batchwise_sheets(wb, export_data)
        if wb.sheetnames and wb.sheetnames[0] == "Compiled":
            del wb["Compiled"]
    elif export_format == "subjectwise":
        _write_subjectwise_sheets(wb, export_data)
        if wb.sheetnames and wb.sheetnames[0] == "Compiled":
            del wb["Compiled"]
    elif export_format == "register":
        _write_register_sheets(wb, export_data)
        if wb.sheetnames and wb.sheetnames[0] == "Compiled":
            del wb["Compiled"]
    else:
        _write_batchwise_sheets(wb, export_data)
        _write_subjectwise_sheets(wb, export_data)
        _write_register_sheets(wb, export_data)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"Weekly_Attendance_{phase}_{export_format}_{date.today():%Y-%m-%d}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_http_methods(["POST"])
def recompute_week_from_daily(request):
    if request.session.get("mentor"):
        return JsonResponse({"ok": False, "msg": "Forbidden"}, status=403)
    module = _active_module(request)
    phase = (request.POST.get("phase") or "T1").upper()
    week_raw = request.POST.get("week_no")
    try:
        week_no = int(week_raw)
    except Exception:
        return JsonResponse({"ok": False, "msg": "Invalid week number"}, status=400)

    normalized_week = _normalize_week_no(phase, week_no)
    if _attendance_lock_for_module_week(module, normalized_week):
        return JsonResponse({"ok": False, "msg": "Week is locked."}, status=400)
    if _has_manual_week(module, normalized_week):
        return JsonResponse({"ok": False, "msg": "Manual upload exists for this week."}, status=400)

    updated, err = recompute_weekly_attendance_from_daily(module, phase, week_no)
    if err:
        return JsonResponse({"ok": False, "msg": err}, status=400)
    return JsonResponse({"ok": True, "updated": updated})
